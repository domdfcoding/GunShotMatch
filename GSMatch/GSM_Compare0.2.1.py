#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
#  GSM_Compare0.1.py
program_name ="GunShotMatch Comparison"
_version = '0.2'

#  Copyright 2018 Dominic Davis-Foster <domdf@dominic-desktop>
#  
#  This program is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 2 of the License, or
#  (at your option) any later version.
#  
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#  
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#  
#  

from GSMatch.utils import timing # Times the program
import sys, os

from GSMatch.utils import clear, check_dependencies
clear()		#clear the display

missing_modules = check_dependencies(["progressbar", "openpyxl", "matplotlib"], prt=False)
if len(missing_modules) > 0:
	for mod in missing_modules:
		print("{} is not installed. Please install it and try again".format(mod))
		sys.exit(1)
		
import progressbar

print("""\r{} Version {} is loading. Please wait...
Copyright 2018 Dominic Davis-Foster""".format(program_name,_version))


import traceback

if "-h" not in str(sys.argv):
	bar = progressbar.ProgressBar(max_value=27) #progressbar for imports

	"""Imports"""
	import time ; time.sleep(0.1); bar.update(1)
	import argparse	; time.sleep(0.1); bar.update(2)
	import csv		; time.sleep(0.1); bar.update(3)
	import datetime	; time.sleep(0.1); bar.update(4)
	import warnings	; time.sleep(0.1); bar.update(5)
	import re		; time.sleep(0.1); bar.update(6)
	import numpy	; time.sleep(0.1); bar.update(7)
	import math		; time.sleep(0.1); bar.update(8)
	import platform	; time.sleep(0.1); bar.update(9)
	import shutil	; time.sleep(0.1); bar.update(10)
	import atexit	; time.sleep(0.1); bar.update(11)
	import ConfigParser	; time.sleep(0.1); bar.update(12)
	import tarfile	; time.sleep(0.1); bar.update(13)

	from openpyxl import  Workbook, worksheet, load_workbook	# https://openpyxl.readthedocs.io/en/default/
	time.sleep(0.1); bar.update(14)
	from openpyxl.styles import Font, Fill, Alignment
	time.sleep(0.1); bar.update(15)
	from openpyxl.chart import BarChart, RadarChart, Series, Reference
	time.sleep(0.1); bar.update(16)
	from openpyxl.utils import get_column_letter, column_index_from_string
	time.sleep(0.1); bar.update(17)

	from shutil import copyfile					; time.sleep(0.1); bar.update(18)
	from decimal import Decimal, ROUND_HALF_UP	; time.sleep(0.1); bar.update(19)
	from collections import Counter				; time.sleep(0.1); bar.update(20)
	from itertools import chain, izip, permutations, product	; time.sleep(0.1); bar.update(21)
	from subprocess import Popen, PIPE, STDOUT	; time.sleep(0.1); bar.update(22)
	from scipy import stats	; time.sleep(0.1); bar.update(23)

	from GSMatch.utils import entry, parent_path, RepresentsInt, rounders, mean_none
	time.sleep(0.1); bar.update(24)
	from GSMatch.utils import append_to_xlsx, load_config, stat_format
	time.sleep(0.1); bar.update(25)
	from GSMatch.utils import *
	time.sleep(0.1); bar.update(26)
	
	col_list = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ','BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ','CA','CB','CC','CD','CE','CF','CG','CH','CI','CJ','CK','CL','CM','CN','CO','CP','CQ','CR','CS','CT','CU','CV','CW','CX','CY','CZ','DA','DB','DC','DE','DF'] #Only Supports 12 samples

else:
	import argparse
	import warnings


"""From https://stackoverflow.com/questions/21129020/how-to-fix-unicodedecodeerror-ascii-codec-cant-decode-byte
Big shout out to fisherman https://stackoverflow.com/users/2309581/fisherman"""
reload(sys)  
sys.setdefaultencoding('utf8')
time.sleep(0.1); bar.update(27)


verbose = False
def verbosePrint(text):	#print only if the --verbose flag is set
	if verbose:
		print(str(text))


def comparison():
	
	"""Open the output file"""
	#output_filename = "./COMPARISON_{}".format(str(datetime.datetime.now())[:-7].replace("-","").replace(":","").replace(" ",""))
	output_filename = "./{}_v_{}_COMPARISON_{}".format(left_sample[:-15],right_sample[:-15],str(datetime.datetime.now())[:-7].replace("-","").replace(":","").replace(" ",""))
	while True:
		try:
			outputCSV = open(output_filename + ".CSV","w")
			outputCSV.write(";;;;{};;;;;;;;{};;;;;;;;t-tests;;;;;;;;\n".format(left_sample, right_sample))
			outputCSV.write(";;;;Retention Time;;;;Peak Area;;;;Retention Time;;;;Peak Area;;;;Retention Time;;;;Peak Area;;;;Welch's t-test Peak Area;;;;MS Comparison\n") 
			outputCSV.write("Name;CAS Number;;;Mean;STDEV;%RSD;;Mean;STDEV;%RSD;;Mean;STDEV;%RSD;;Mean;STDEV;%RSD;;t-statistic;p-value;Result;;t-statistic;p-value;Result;;t-statistic;p-value;Result;;Mean;STDEV;%RSD\n")
		
			outputCSV.close()
			break
		except IOError:
			print "The file \"" + output_filename + "\" is locked for editing in another program. Press any key to try again."
			raw_input(">")
	
	outputCSV = open(output_filename + ".CSV","a")

	
	if not os.path.exists("./.temp"):
		os.makedirs("./.temp")

	"""get list of LEFT samples and statistics data"""
	left_comparison = []
	left_statistics = []
	
#	print left_sample
#	print left_sample[:-15]
	
	"""from https://stackoverflow.com/questions/31163668/how-do-i-extract-a-tar-file-using-python-2-4"""
	tar = tarfile.open("{}.tar.gz".format(left_sample))
	tar.extractall(path='./.temp/{}'.format(left_sample[:-15]))
	tar.close()
	
	"""From http://openpyxl.readthedocs.io/en/default/"""
	wb = load_workbook("./.temp/{0}/{0}_FINAL.xlsx".format(left_sample[:-15]))
	
	combined_statistics=[]
	
	#Get samples list
	ws = wb["Index"]
	#for row in ws.iter_rows('B13:B{}'.format(ws.max_row)):
	for row in ws.iter_rows('F2:AF2'):
		for cell in row:
			if cell.value:
				#print cell.value
				if len(cell.value) > 0:
#					left_comparison.append(cell.value.split(",")[1].replace('"','')[:-1])
					left_comparison.append(cell.value)

	
#	print left_comparison
	
	#get statistics
	ws = wb["Statistics"]
	for row_index, row in enumerate(ws.iter_rows('A2:L{}'.format(ws.max_row))):
		if row_index > 0:
			left_statistics.append([cell.value for cell in row]+['','','','','','','',''])
	
	for row in left_statistics:
		row=[row[4]]+row
		combined_statistics.append(row)
	
#	for row in left_statistics:
#		print row
	
	
	"""get list of RIGHT samples and statistics data"""
	right_comparison = []
	right_statistics = []
	
#	print right_sample
#	print right_sample[:-15]
	
	"""from https://stackoverflow.com/questions/31163668/how-do-i-extract-a-tar-file-using-python-2-4"""
	tar = tarfile.open("{}.tar.gz".format(right_sample))
	tar.extractall(path='./.temp/{}'.format(right_sample[:-15]))
	tar.close()
	
	"""From http://openpyxl.readthedocs.io/en/default/"""
	wb = load_workbook("./.temp/{0}/{0}_FINAL.xlsx".format(right_sample[:-15]))
	
	#Get samples list
	ws = wb["Index"]
	#for row in ws.iter_rows('B13:B{}'.format(ws.max_row)):
	for row in ws.iter_rows('F2:AF2'):
		for cell in row:
			if cell.value:
#				print cell.value
#				print "'{}'".format(cell.value)
				if len(cell.value) > 0:
				#	right_comparison.append(cell.value.split(",")[1].replace('"','')[:-1])
					right_comparison.append(cell.value)
	
#	print right_comparison
	
	#get statistics
	ws = wb["Statistics"]
	temp_statistics = []
	for row_index, row in enumerate(ws.iter_rows('A2:L{}'.format(ws.max_row))):
		if row_index > 0:
				temp_statistics.append([cell.value for cell in row])

	for row in temp_statistics:
		try:
			row.insert(4,'')
			row.insert(4,'')
			row.insert(4,'')
			row.insert(4,'')
			row.insert(4,'')
			row.insert(4,'')
			row.insert(4,'')
			row.insert(4,'')
		except AttributeError:
			pass
		right_statistics.append(row)
		
	for row in right_statistics:
		row=[row[12]]+row
		combined_statistics.append(row)	
	
	
	from operator import itemgetter
	name_sorted_combined_statistics = sorted(combined_statistics, key=itemgetter(0))
	
#	for row in name_sorted_combined_statistics:
#		print row
	
	merged_hit_list = []
	
	for row_index, row in enumerate(name_sorted_combined_statistics):
		if row_index == len(name_sorted_combined_statistics)-1:
			break
		elif row[1] == name_sorted_combined_statistics[row_index+1][1] and within1min(row[5],name_sorted_combined_statistics[row_index+1][13]):
			if name_sorted_combined_statistics[row_index+1][5] == '':
		
				row[13] = name_sorted_combined_statistics[row_index+1][13]
				row[14] = name_sorted_combined_statistics[row_index+1][14]
				row[15] = name_sorted_combined_statistics[row_index+1][15]
								
				row[17] = name_sorted_combined_statistics[row_index+1][17]
				row[18] = name_sorted_combined_statistics[row_index+1][18]
				row[19] = name_sorted_combined_statistics[row_index+1][19]
								
				#print row
				name_sorted_combined_statistics[row_index] = row
				del name_sorted_combined_statistics[row_index+1]
				
		elif row[1] == name_sorted_combined_statistics[row_index+1][1] and within1min(row[13],name_sorted_combined_statistics[row_index+1][5]):
			if name_sorted_combined_statistics[row_index+1][13] == '':
		
				row[5] = name_sorted_combined_statistics[row_index+1][5]
				row[6] = name_sorted_combined_statistics[row_index+1][6]
				row[7] = name_sorted_combined_statistics[row_index+1][7]
								
				row[9] = name_sorted_combined_statistics[row_index+1][9]
				row[10] = name_sorted_combined_statistics[row_index+1][10]
				row[11] = name_sorted_combined_statistics[row_index+1][11]
								
				#print row
				name_sorted_combined_statistics[row_index] = row
				del name_sorted_combined_statistics[row_index+1]
	
	RT_sorted_combined_statistics = sorted(name_sorted_combined_statistics, key=itemgetter(0))

	"""Calculate t statistics from MEAN and STDEV"""
	a_value = 0.01
	
	for row in RT_sorted_combined_statistics:
		
		del row[-1]
		#print row[0]
		
		"""Retention Time"""
		l_mean = (row[5] if row[5]!='' else 0.0)
		l_std = (row[6] if row[6]!='' else 0.0)
		l_nobs = len(left_comparison)
		#print left_comparison
		#print len(left_comparison)
		
		r_mean = (row[13] if row[13]!='' else 0.0)
		r_std = (row[14] if row[14]!='' else 0.0)
		r_nobs = len(right_comparison)
		#print right_comparison
		#print len(right_comparison)

#		print l_mean, ' ', l_std, ' ', l_nobs		
#		print r_mean, ' ', r_std, ' ', r_nobs
		
		row.append('')
		#raw_input(">")
		
		if not any(x == 0.0 for x in [l_mean,l_nobs,r_mean,r_nobs]):
#			print "Top Banana! {}".format(l_mean)
			rt_t_stat, rt_p_val = stats.ttest_ind_from_stats(l_mean,l_std,l_nobs,r_mean,r_std,r_nobs)
		
#			print "RT t-Statistic = {}".format(rt_t_stat)
#			print "RT p-value = {}".format(rt_p_val)
			
			if str(rt_t_stat) == "inf":
				rt_t_stat = "#"
			
			row.append(rt_t_stat)
			row.append(rt_p_val)
			
			if rt_p_val <= a_value:
	#			print "Reject Null Hypothesis; Means are Different"
				row.append("Diff")
			else:
	#			print "Don't Reject Null Hypothesis; Means are Same"		
				row.append("Same")
				
		
		else:
			row.append("")
			row.append("")
			row.append("Diff")
		

				
		"""Peak Area"""#	8, 9 and 16, 17
		l_mean = (row[9] if row[9]!='' else 0.0)
		l_std = (row[10] if row[10]!='' else 0.0)
		l_nobs = len(left_comparison)
		
		r_mean = (row[17] if row[17]!='' else 0.0)
		r_std = (row[18] if row[18]!='' else 0.0)
		r_nobs = len(right_comparison)

#		print l_mean, ' ', l_std, ' ', l_nobs		
#		print r_mean, ' ', r_std, ' ', r_nobs
		
		row.append('')
		
		# Independant 2 samples t-test
		if not any(x == 0.0 for x in [l_mean,l_nobs,r_mean,r_nobs]):
			pa_t_stat, pa_p_val = stats.ttest_ind_from_stats(l_mean,l_std,l_nobs,r_mean,r_std,r_nobs)
	#		print "Independant samples t-test"
	#		print "PA t-Statistic = {}".format(pa_t_stat)
	#		print "PA p-value = {}".format(pa_p_val)
		
			row.append(pa_t_stat)
			row.append(pa_p_val)
			
			if pa_p_val <= a_value:
	#			print "Reject Null Hypothesis; Means are Different"
				row.append("Diff")
			else:
	#			print "Don't Reject Null Hypothesis; Means are Same"		
				row.append("Same")
		
		else:
			row.append("")
			row.append("")
			row.append("Diff")
		
#		print row
#		raw_input(">")
#		clear()

		row.append('')

		# Welch's t-test
		if not any(x == 0.0 for x in [l_mean,l_nobs,r_mean,r_nobs]):
			pa_t_stat, pa_p_val = stats.ttest_ind_from_stats(l_mean,l_std,l_nobs,r_mean,r_std,r_nobs, False)
	#		print "Welch's t-test"		
	#		print "PA t-Statistic = {}".format(pa_t_stat)
	#		print "PA p-value = {}".format(pa_p_val)
		
			row.append(pa_t_stat)
			row.append(pa_p_val)
			
			if pa_p_val <= a_value:
	#			print "Reject Null Hypothesis; Means are Different"
				row.append("Diff")
			else:
	#			print "Don't Reject Null Hypothesis; Means are Same"		
				row.append("Same")
		
		else:
			row.append("")
			row.append("")
			row.append("Diff")


		
	"""MS Comparison"""
	rt_list = []
	for row in RT_sorted_combined_statistics:
		if row[5] not in ['', 0.0,None] and row[13] not in ['', 0.0,None]:
			rt_list.append(row[5])
			rt_list.append(row[13])
	
	generate_nist_library([row[5],row[13]])		#remember to uncomment
	
	for row in RT_sorted_combined_statistics:
		if row[5] not in ['', 0.0,None] and row[13] not in ['', 0.0,None]:
			print left_comparison, right_comparison, row[5], row[13]
			MF_mean, MF_std, MF_rsd = nist_ms_comparison(left_comparison, right_comparison, row[5], row[13])
			row.append("")
			row.append(MF_mean)
			row.append(MF_std)
			row.append(MF_rsd)
			time.sleep(0.5)
	
	"""Write to CSV"""
	for row in RT_sorted_combined_statistics:
		#print row
		
		"""from https://stackoverflow.com/questions/19363881/replace-none-value-in-list"""
		outputCSV.write(";".join(['' if v is None else str(v) for v in row[1:]]))
		outputCSV.write("\n")
	
	outputCSV.close()

	time.sleep(3)
	append_to_xlsx(output_filename + ".CSV", output_filename + ".xlsx", "Comparison", overwrite=True,seperator=";")
	comparisonFormat(output_filename + ".xlsx", a_value)
	shutil.move(output_filename + ".CSV", OUTPUT_DIRECTORY)
	try:
		os.unlink("ms_comparisons_COMPARISON")
	except:
		pass
	try:
		os.unlink("ms_comparisons_COMPARISON")
	except:
		pass
			

def within1min(value1, value2):
	if value1 not in [0, None, ''] and value2 not in [0, None, '']:
		return (float(value1)-1) < (float(value2)) < (float(value1)+1)
	else:
		return False

def comparisonFormat(inputFile, a_value = 0.05):	# Formatting for Final Output
	widgets = [progressbar.widgets.AnimatedMarker(), ' ','Generating XLSX Output.', ' ', progressbar.Timer(),]
	bar = progressbar.ProgressBar(widgets=widgets, max_value=progressbar.UnknownLength).start()
	bar_step = 1	
	
	"""From http://openpyxl.readthedocs.io/en/default/"""
	wb = load_workbook(inputFile)
	
	# grab the active worksheet
	ws = wb["Comparison"]

	stat_format(ws,sheet_type="Comparison")
	
	ws.cell("A1").value = u"Threshold α={}".format(a_value)
	
	#Additional formatting
#	ws.merge_cells("E2:G2")
#	ws.merge_cells("I1:K1")
#			wsX.merge_cells("M1:O1")
#			wsX.merge_cells("Q1:S1")
#			wsX.merge_cells("U1:W1")
#			wsX.merge_cells("Y1:AA1")

	# Save the file
	wb.save(inputFile)
			
def copytree(src, dst, symlinks=False, ignore=None):
	"""from https://stackoverflow.com/questions/1868714/how-do-i-copy-an-entire-directory-of-files-into-an-existing-directory-using-pyth
		because shutil.copytree is borked"""
	for item in os.listdir(src):
		s = os.path.join(src, item)
		d = os.path.join(dst, item)
		if os.path.isdir(s):
			shutil.copytree(s, d, symlinks, ignore)
		else:
			shutil.copy2(s, d)

def generate_nist_library(retention_times):		#pass as many RTs as you want to compare overall
	if type(retention_times) == str:
		retention_times = [retention_times]
	
	if os.path.exists("./.temp/MSP/"):
		shutil.rmtree("./.temp/MSP/")
	os.makedirs("./.temp/MSP/")
	
	"""Merge MSP Files together for create NIST Library"""
	copytree('./.temp/{}/MSP/'.format(left_sample[:-15]),"./.temp/MSP/")
	copytree('./.temp/{}/MSP/'.format(right_sample[:-15]),"./.temp/MSP/")
		
	
	"""From https://stackoverflow.com/questions/13613336/python-concatenate-text-files"""
	with open("./.temp/comparison.msp", 'w') as outfile:
		for filename in os.listdir("./.temp/MSP/"):
			if any(str(x) in filename for x in retention_times):# in filename:
				with open(os.path.join("./.temp/MSP",filename)) as infile:
					for line in infile:
						outfile.write(line)
	
#	print lib2nist_path
	
	#Generate NIST Library
	print("\n\nClick exit once the library has been generated. If nothing happens just click exit anyway.")
	print("On Linux you need to navigate to the .MSP file manually. It should be located at {}".format(os.path.join(os.getcwd(),".temp/comparison.msp")))
	os.system("{}".format("wine " if platform.system() == "Linux" else "") + lib2nist_path + "lib2nist.exe " + "./.temp/comparison.msp")

	#Copy library to NIST directory
	if os.path.isdir(nist_path+"comparison"):
		shutil.rmtree(nist_path+"comparison") #delete existing directory
	shutil.copytree(lib2nist_path + "comparison", nist_path + "comparison")

def nist_ms_comparison(left_comparison, right_comparison, left_rt, right_rt):	#pass one pair of retention times
	
#	print left_comparison
#	print right_comparison
#	print left_rt
#	print right_rt
	#raw_input(">")
	
	left_rt = rounders(left_rt,"0.000")
	right_rt = rounders(right_rt,"0.000")
	
	print left_rt
	print right_rt
	
	
	data_dict = {}

	try:
		generate_ini(nist_path, "comparison", 50)
		
		perms = []
		for i in product(left_comparison, right_comparison):
#			print i
			perms.append(i)

		
		def remove_chars(input_string):
			return input_string.replace("Hit 10",""
				).replace("Hit 11","").replace("Hit 12","").replace("Hit 13","").replace("Hit 14","").replace("Hit 15",""
				).replace("Hit 16","").replace("Hit 17","").replace("Hit 18","").replace("Hit 19","").replace("Hit 20",""
				).replace("Hit 21","").replace("Hit 22","").replace("Hit 23","").replace("Hit 24","").replace("Hit 25",""
				).replace("Hit 26","").replace("Hit 27","").replace("Hit 28","").replace("Hit 29","").replace("Hit 30",""
				).replace("Hit 31","").replace("Hit 32","").replace("Hit 33","").replace("Hit 34","").replace("Hit 35",""
				).replace("Hit 36","").replace("Hit 37","").replace("Hit 38","").replace("Hit 39","").replace("Hit 40",""
				).replace("Hit 41","").replace("Hit 42","").replace("Hit 43","").replace("Hit 44","").replace("Hit 45",""
				).replace("Hit 46","").replace("Hit 47","").replace("Hit 48","").replace("Hit 49","").replace("Hit 50",""
				).replace("Hit 1","").replace("Hit 2","").replace("Hit 3","").replace("Hit 4","").replace("Hit 5",""
				).replace("Hit 6","").replace("Hit 7","").replace("Hit 8","").replace("Hit 9","").replace("MF:",""
				).replace(":","").replace("<","").replace(">","").replace(" ","")#.replace("_"+rt_list[index],"")
		
		print("\nMass Spectrum Comparison in progress. Please wait.")

		with open("./ms_comparisons_{}.txt".format("COMPARISON"),"w") as f:
			f.write("Left RT = {}; Right RT = {}\n".format(left_rt, right_rt))
			
			compare_MF = []
			
			for left_sample_name in left_comparison:
				#print left_sample_name
				
				if not os.path.exists(os.path.join("./.temp/MSP",("{}_{}.MSP".format(left_sample_name,rounders(left_rt,"0.000"))))):
					if os.path.exists(os.path.join("./.temp/MSP",("{}_{}.MSP".format(left_sample_name,rounders(left_rt+Decimal(0.001),"0.000"))))):
						raw_output = nist_db_connector(nist_path,os.path.join("./.temp/MSP",("{}_{}.MSP".format(left_sample_name,rounders(left_rt+Decimal(0.001),"0.000")))))
					elif os.path.exists(os.path.join("./.temp/MSP",("{}_{}.MSP".format(left_sample_name,rounders(left_rt-Decimal(0.001),"0.000"))))):
						raw_output = nist_db_connector(nist_path,os.path.join("./.temp/MSP",("{}_{}.MSP".format(left_sample_name,rounders(left_rt-Decimal(0.001),"0.000")))))
				else:
					raw_output = nist_db_connector(nist_path,os.path.join("./.temp/MSP",("{}_{}.MSP".format(left_sample_name,rounders(left_rt,"0.000")))))	##here
#				print raw_output
							
				search_results = [left_sample_name]

				for row_index, row in enumerate(raw_output.split("\n")):

					if row_index > 0:
						if str(right_rt) in row:
							if left_sample_name not in row:
								if any(x in row for x in right_comparison):
									search_results.append(remove_chars(row.split(";")[0]).replace("_"+str(right_rt),""))
									search_results.append(remove_chars(row.split(";")[2]).replace("_"+str(right_rt),""))
			
				
#	for diagnostics				if len(search_results) != 5:
#						print "RT = {}".format(rt_list[index])
#						print ""
#						print raw_output
#						print len(raw_output)
#						print""
#						print search_results
#						raw_input(">")
				
#				print search_results
				
				for combo in perms:
					if combo[0] == search_results[0]:
						if combo[1] in search_results:
							f.write(str(combo) + " " + search_results[1+search_results.index(combo[1])]+"\n")
							compare_MF.append(int(search_results[1+search_results.index(combo[1])]))
						else:
							f.write(str(combo) + " Not Found\n")
							compare_MF.append(0)
				time.sleep(0.5)
				
			MF_mean = numpy.mean(compare_MF)
			MF_std = numpy.std(compare_MF, ddof=1)
			MF_rsd = numpy.std(compare_MF, ddof=1)/numpy.mean(compare_MF)
			f.write(str(MF_mean)+"\n")
			f.write((str(MF_std))+"\n")
			f.write(str(MF_rsd)+"\n")
			f.write("\n")
			f.write("\n")
	except:
		traceback.print_exc()	#print the error
		reload_ini(nist_path)
		sys.exit(1)
	
	print "\n"
	reload_ini(nist_path)
#	print MF_mean, MF_std, MF_rsd
	return MF_mean, MF_std, MF_rsd

lib2nist_path, nist_path, NULL, NULL, OUTPUT_DIRECTORY, NULL, XY_DIRECTORY, MSP_DIRECTORY, NULL, NULL, NULL, NULL, NULL = load_config("./config.ini")
	

"""Define Exit Functions"""
if "-h" not in str(sys.argv):
	#atexit.register(final_cleanup)
	atexit.register(reload_ini, nist_path)

if __name__ == '__main__':
	"""from https://stackoverflow.com/questions/14463277/how-to-disable-python-warnings"""
	with warnings.catch_warnings():
		warnings.simplefilter("ignore")

		parser = argparse.ArgumentParser()
		
		#Command line switches
		parser.add_argument("--info",help="Show program info.", action='store_true')
		parser.add_argument("-v","--verbose",help="Turns on verbosity for diagnostics", action='store_true')
		parser.add_argument("-l","--left",help="LEFT sample to compare.",nargs='+')
		parser.add_argument("-r","--right",help="RIGHT sample to compare.",nargs='+')
		
		args = parser.parse_args()
		
		if args.info:
			parser.print_help()	#show help and info
			info()
			sys.exit(0) 
			
		"""Sample List"""
		samples_error = "No samples specified. Please enter samples in list.txt or following the --samples argument\n."
	
		left_sample, right_sample = '',''
	
		if args.left and args.right:
			left_sample = args.left[0]
			right_sample = args.right[0]
		
		elif os.path.isfile("comparison_list.txt"):
			with open("./comparison_list.txt","r") as f:
				prefixList = f.read().splitlines()
				left_sample = prefixList[0]
				right_sample = prefixList[1]
				"""From https://stackoverflow.com/questions/12330522/reading-a-file-without-newlines"""
		
		else:
			print(samples_error)
			sys.exit(1)
		
		if len(right_sample) == 0:
			print(samples_error)
			sys.exit(1)
		if len(left_sample) == 0:
			print(samples_error)
			sys.exit(1)
			
		time.sleep(0.1); bar.update(25)
		
		print("\nReady")
		print("Samples to compare: {} and {}".format(left_sample,right_sample))

		comparison()	# Actual Comparison
			
	#	make_archive()
	#	final_cleanup()
		print("\nComplete.")

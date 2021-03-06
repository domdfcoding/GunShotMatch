#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
#  GSMatch_Comparison.py
"""
GunShotMatch Project Comparison
"""
#
#  This file is part of GunShotMatch
#
#  Copyright 2017-2020 Dominic Davis-Foster <dominic@davis-foster.co.uk>
#
#  This program is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 3 of the License, or
#  (at your option) any later version.
#
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#


# stdlib
import json
import operator
import os
import sys
import time
from itertools import product

# 3rd party
import numpy
import pandas
from chemistry_tools import spectrum_similarity
from domdf_python_tools.terminal import br, clear
from domdf_spreadsheet_tools import append_to_xlsx, format_header, format_sheet, make_column_property_dict
from mathematical.data_frames import df_count, df_mean
from mathematical.utils import rounders
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pyms.DPA.Alignment import exprl2alignment
from pyms.DPA.PairwiseAlignment import align, align_with_tree, PairwiseAlignment
from pyms.Experiment import load_expr

# this package
from GSMatch.GSMatch_Core.charts import (
	box_whisker_wrapper, bw_default_colours, bw_default_styles, default_colours, default_filetypes,
	PrincipalComponentAnalysis, radar_chart_wrapper,  # radar_chart, box_whisker,
	)
from GSMatch.GSMatch_Core.PeakAlignment import get_ms_alignment, get_peak_alignment


__author__ = "Dominic Davis-Foster"
__copyright__ = "Copyright 2017-2019 Dominic Davis-Foster"
__license__ = "GPLv3"
__version__ = "0.1.0"
__email__ = "dominic@davis-foster.co.uk"

program_name = "GunShotMatch Project Comparison"


class GSMCompare(object):
	def __init__(self, left_sample, right_sample, config=None):
		from domdf_python_tools.paths import maybe_make
		
		if config is None:
			from GSMatch.GSMatch_Core.Config import GSMConfig
			self.config = GSMConfig("config.ini")
		else:
			self.config = config  # GSMConfig object
		
		maybe_make(os.path.join(self.config.charts_dir, "Comparison"))
		
		with open(os.path.join(left_sample), "r") as info_file:
			self.left_prefixList = [x.rstrip("\r\n") for x in info_file.readlines()]
		
		with open(os.path.join(right_sample), "r") as info_file:
			self.right_prefixList = [x.rstrip("\r\n") for x in info_file.readlines()]
		
		self.left_sample = os.path.splitext(os.path.split(left_sample)[-1])[0]
		self.right_sample = os.path.splitext(os.path.split(right_sample)[-1])[0]
		
		print(self.left_sample, self.left_prefixList)
		print(self.right_sample, self.right_prefixList)
		
		self.comparison_name = f"{self.left_sample} v {self.right_sample}"
	
	def setup_data(self):
		import pandas
		from mathematical.data_frames import df_count
		
		"""Chart Data"""
		chart_data = pandas.concat(
				[
						pandas.read_csv(
								os.path.join(
										self.config.csv_dir,
										f"{self.left_sample}_CHART_DATA.csv"
										),
								sep=";",
								index_col=0),
						pandas.read_csv(
								os.path.join(
										self.config.csv_dir,
										f"{self.right_sample}_CHART_DATA.csv"
										),
								sep=";",
								index_col=0
								)
						], axis=1, sort=False)
		
		chart_data.drop("Compound Names", axis=1, inplace=True)
		chart_data['Compound Names'] = chart_data.index
		
		# determine order of compounds on graph
		for compound in chart_data.index.values:
			chart_data["Count"] = chart_data.apply(
					df_count,
					args=([f"{sample} Peak Area" for sample in [self.left_sample, self.right_sample]],),
					axis=1)
		
		chart_data['Compound Names'] = chart_data.index
		self.chart_data = chart_data.sort_values(['Count', 'Compound Names'])
		self.chart_data.fillna(0, inplace=True)
	
	def setup_charts(
			self,
			display=False,
			styles=bw_default_styles,
			bw_colours=bw_default_colours,
			colours=default_colours,
			filetypes=default_filetypes
			):
		if display:
			self.radar_mode = "display"
			self.box_whisker_mode = "display"
			self.pca_mode = "display"
		else:
			self.radar_mode = os.path.join(
					self.config.charts_dir,
					"Comparison",
					f"{self.comparison_name}_RADAR"
					)
			self.box_whisker_mode = os.path.join(
					self.config.charts_dir, "Comparison",
					f"{self.comparison_name}_BOX_WHISKER"
					)
			self.pca_mode = os.path.join(
					self.config.charts_dir, "Comparison",
					f"{self.comparison_name}_PCA"
					)
		
		self.colours = colours
		self.bw_colours = bw_colours
		self.bw_styles = styles
		self.filetypes = filetypes
	
	def radar_chart(self, figsize=(4, 9), use_log=10):
		
		radar_chart_wrapper(
				self.chart_data,
				[self.left_sample, self.right_sample],
				use_log=use_log,
				colours=self.colours,
				figsize=figsize,
				mode=self.radar_mode,
				filetypes=self.filetypes
				)
	
	def box_whisker_chart(
			self,
			figsize=None,
			show_outliers=True,
			show_raw_data=False,
			err_bar="range",
			outlier_mode="2stdev",
			leg_cols=1,
			column_width=4
			):
		
		box_whisker_wrapper(
				self.chart_data, [
						(self.left_sample, self.left_prefixList),
						(self.right_sample, self.right_prefixList)
						],
				show_outliers=show_outliers,
				show_raw_data=show_raw_data,
				err_bar=err_bar,
				outlier_mode=outlier_mode,
				leg_cols=leg_cols,
				mode=self.box_whisker_mode,
				figsize=figsize,
				column_width=column_width,
				styles=self.bw_styles,
				colours=self.bw_colours,
				filetypes=self.filetypes
				)
	
	def pca(self, figsize=(8, 8)):
		"""Principal Component Analysis"""
		
		pca_data = {
				"target": [self.left_sample] * len(self.left_prefixList) + [self.right_sample] * len(
						self.right_prefixList)
				}
		features = []
		targets = [self.left_sample, self.right_sample]
		
		for compound in self.chart_data.index.values:
			area_list = []
			for prefix in self.left_prefixList + self.right_prefixList:
				area_list.append(self.chart_data.loc[compound, prefix])
			pca_data[compound] = area_list
			features.append(compound)
		
		pca_data = pandas.DataFrame(data=pca_data)
		
		pca_chart = PrincipalComponentAnalysis()
		self.pca = pca_chart.setup_data(pca_data, features, targets)
		pca_chart.setup_subplots(figsize)
		pca_chart.setup_datapoints(self.colours)
		pca_chart.create_chart()
		pca_chart.create_legend()
		pca_chart.fig.tight_layout()
		if self.pca_mode == "display":
			pca_chart.show_chart()
		else:
			pca_chart.save_chart(self.pca_mode)
	
	def peak_comparison(self, a_value=0.05):
		"""
		Open the output file
		
		:param a_value:
		:type a_value:
		:return:
		:rtype:
		"""
		output_filename = os.path.join(self.config.results_dir, f"{self.comparison_name}_COMPARISON")
		
		while True:
			try:
				outputCSV = open(output_filename + ".CSV", "w")
				outputCSV.write(
						f"Explained Variance Ratio: {rounders(self.pca[0], '0.0000')}, "
						f"{rounders(self.pca[1], '0.0000')};;;;{self.left_sample};;;;;;;;"
						f"{self.right_sample};;;;;;;;t-tests;;;;;;;;\n")
				outputCSV.write(
						f"t-test Threshold α={a_value};;;;Retention Time;;;;Peak Area;;;;"
						f"Retention Time;;;;Peak Area;;;;Retention Time;;;;Peak Area;;;;"
						f"Welch's t-test Peak Area;;;;MS Comparison\n")
				outputCSV.write(
						"Name;CAS Number;;;Mean;STDEV;%RSD;;Mean;STDEV;%RSD;;Mean;STDEV;"
						"%RSD;;Mean;STDEV;%RSD;;t-statistic;p-value;Result;;t-statistic;"
						"p-value;Result;;t-statistic;p-value;Result;;Mean;STDEV;%RSD\n")
				
				break
			except IOError:
				print(f"The file '{output_filename}' is locked for editing in another program.")
				input("Press any key to try again.")
		
		"""Peak Data"""
		left_peak_data = []
		with open(os.path.join(self.config.csv_dir, f"{self.left_sample}_peak_data.json"), "r") as jsonfile:
			for peak in jsonfile.readlines():
				left_peak_data.append(json.loads(peak))
		
		right_peak_data = []
		with open(os.path.join(self.config.csv_dir, f"{self.right_sample}_peak_data.json"), "r") as jsonfile:
			for peak in jsonfile.readlines():
				right_peak_data.append(json.loads(peak))
		
		"""Alignment Data"""
		# define the input experiments list
		left_expr_list = []
		for prefix in self.left_prefixList:
			file_name = os.path.join(self.config.expr_dir, f"{prefix}.expr")
			left_expr_list.append(load_expr(file_name))
		
		right_expr_list = []
		for prefix in self.right_prefixList:
			file_name = os.path.join(self.config.expr_dir, f"{prefix}.expr")
			right_expr_list.append(load_expr(file_name))
		
		print("\nAligning\n")
		left_F1 = exprl2alignment(left_expr_list)
		left_T1 = PairwiseAlignment(
				left_F1, self.config.comparison_rt_modulation, self.config.comparison_gap_penalty)
		left_A1 = align_with_tree(left_T1, min_peaks=self.config.comparison_min_peaks)
		
		right_F2 = exprl2alignment(right_expr_list)
		right_T2 = PairwiseAlignment(
				right_F2, self.config.comparison_rt_modulation, self.config.comparison_gap_penalty)
		right_A2 = align_with_tree(right_T2, min_peaks=self.config.comparison_min_peaks)
		
		both_alignment = align(
				left_A1, right_A2, self.config.comparison_rt_modulation, self.config.comparison_gap_penalty)
		
		# print(score_matrix(left_A1, right_A2, Dw))
		
		rt_alignment = get_peak_alignment(both_alignment)
		
		if not rt_alignment.empty:
			rt_alignment[self.left_sample] = rt_alignment.apply(
					df_mean, axis=1,
					args=([prefix for prefix in self.left_prefixList],),
					)
			rt_alignment[self.right_sample] = rt_alignment.apply(
					df_mean, axis=1,
					args=([prefix for prefix in self.right_prefixList],),
					)
			
			ms_alignment = get_ms_alignment(both_alignment)
			
			left_aligned_peaks = find_aligned_peaks(
					self.left_sample, self.left_prefixList, left_peak_data, rt_alignment, ms_alignment)
			right_aligned_peaks = find_aligned_peaks(
					self.right_sample, self.right_prefixList, right_peak_data, rt_alignment, ms_alignment)
			
			# # print(f"{left_sample} Peaks")
			# for index, aligned_peak in enumerate(rt_alignment[self.left_sample]):
			# 	found_peak = False
			# 	for peak in left_peak_data:
			# 		# print(aligned_peak, peak["average_rt"])
			# 		if peak["average_rt"] == aligned_peak:
			# 			# print(peak)
			# 			# for key in peak:
			# 			# 	print(f"{key}: {peak[key]}")
			# 			peak["ms_list"] = [ms_alignment.iloc[index][prefix] for prefix in self.left_prefixList]
			# 			left_aligned_peaks.append(peak)
			# 			found_peak = True
			# 			break
			# 	if not found_peak:
			# 		left_aligned_peaks.append(None)
			#
			# # print(f"{right_sample} Peaks")
			# for index, aligned_peak in enumerate(rt_alignment[self.right_sample]):
			# 	found_peak = False
			# 	for peak in right_peak_data:
			# 		if peak["average_rt"] == aligned_peak:
			# 			# print(peak)
			# 			# for key in peak:
			# 			# 	print(f"{key}: {peak[key]}")
			# 			peak["ms_list"] = [ms_alignment.iloc[index][prefix] for prefix in self.right_prefixList]
			# 			right_aligned_peaks.append(peak)
			# 			found_peak = True
			# 			break
			# 	if not found_peak:
			# 		right_aligned_peaks.append(None)
			
			aligned_non_matching_peaks = []
			
			for left_peak, right_peak in zip(left_aligned_peaks, right_aligned_peaks):
				if not any([left_peak is None, right_peak is None]):
					# print(f"{left_peak['average_rt']}		{right_peak['average_rt']}")
					if f"{left_peak['hits'][0]['CAS']}" == f"{right_peak['hits'][0]['CAS']}":
						# The top hit for each project is the same
						# print(f"{left_peak['hits'][0]['Name']}		{right_peak['hits'][0]['Name']}")
						left_hit_number, right_hit_number = 0, 0
					else:  # Check if there is a match in the other four hits
						left_hit_dict = {}
						right_hit_dict = {}
						for hit_num in range(0, 5):
							left_hit_dict[f"{left_peak['hits'][hit_num]['CAS']}"] = hit_num
							right_hit_dict[f"{right_peak['hits'][hit_num]['CAS']}"] = hit_num
						left_hit_set = set(left_hit_dict)
						right_hit_set = set(right_hit_dict)
						
						results_list = []
						for CAS in left_hit_set.intersection(right_hit_set):
							# CAS Number and Hit Numbers of hits in common
							left_hit_num = left_hit_dict[CAS]
							right_hit_num = right_hit_dict[CAS]
							left_hit_mf = left_peak["hits"][left_hit_num]["average_MF"]
							right_hit_mf = right_peak["hits"][right_hit_num]["average_MF"]
							
							results_list.append([
									CAS, left_hit_num, right_hit_num,
									numpy.mean([left_hit_num, right_hit_num]),
									numpy.mean([left_hit_mf, right_hit_mf])
									])
						
						results_list = sorted(results_list, key=operator.itemgetter(3, 4))
						# print(results_list[0])
						# print(results_list)
						
						if not results_list:
							aligned_non_matching_peaks.append((left_peak, right_peak))
							continue
						
						left_hit_number, right_hit_number = results_list[0][1:3]
					
					# print(f"{left_peak['hits'][left_hit_number]['Name']}		{right_peak['hits'][right_hit_number]['Name']}")
					
					# Write output data
					name = left_peak['hits'][left_hit_number]['Name']
					CAS = left_peak['hits'][left_hit_number]['CAS']
					
					mf_mean, mf_stdev = ms_comparisons(left_peak["ms_list"], right_peak["ms_list"])
					
					write_peak(
							outputCSV, name, CAS,
							*get_peak_rt_stats(left_peak),
							*get_peak_area_stats(left_peak),
							*get_peak_rt_stats(right_peak),
							*get_peak_area_stats(right_peak),
							mf_mean, mf_stdev, a_value
							)
			
			# print('The following peaks were aligned by retention time but none of the "hits" matched:')
			for peak_pair in aligned_non_matching_peaks:
				# br()
				# print(f"Retention time: {peak_pair[0]['average_rt']}")
				# print(f"Left Peak: {pformat(peak_pair[0])}")
				left_peak = peak_pair[0]
				# print(f"Right Peak: {pformat(peak_pair[1])}")
				right_peak = peak_pair[1]
				
				# Write output data
				write_output_data(left_peak, outputCSV)
				write_output_data(right_peak, outputCSV)
			
			# print("Peaks in the left sample only:")
			
			for peak in left_peak_data:
				if peak not in left_aligned_peaks:
					# Write output data
					name = peak['hits'][0]['Name']
					CAS = peak['hits'][0]['CAS']
					
					left_rt_mean, left_rt_stdev, left_rt_n = get_peak_rt_stats(peak)
					left_area_mean, left_area_stdev, left_area_n = get_peak_area_stats(peak)
					
					write_peak(
							outputCSV, name, CAS,
							left_rt_mean=left_rt_mean,
							left_rt_stdev=left_rt_stdev,
							left_area_mean=left_area_mean,
							left_area_stdev=left_area_stdev,
							a_value=a_value
							)
			
			# outputCSV.write(f"{name};{CAS};;;")  # Name;CAS Number;;;
			#
			# outputCSV.write(
			# 	f"{left_rt_mean};{left_rt_stdev};{left_rt_stdev / left_rt_mean};;")  # Mean RT left;STDEV RT left;%RSD RT left;;
			# outputCSV.write(
			# 	f"{left_area_mean};{left_area_stdev};{left_area_stdev / left_area_mean};;")  # Mean Area left;STDEV Area left;%RSD Area left;;
			#
			# outputCSV.write(";;;;;;;;;;;;;;;;;;;;;;;;")
			# outputCSV.write("\n")
			
			# print("Peaks in the right sample only:")
			
			for peak in right_peak_data:
				if peak not in right_aligned_peaks:
					# Write output data
					name = peak['hits'][0]['Name']
					CAS = peak['hits'][0]['CAS']
					
					right_rt_mean, right_rt_stdev, right_rt_n = get_peak_rt_stats(peak)
					right_area_mean, right_area_stdev, right_area_n = get_peak_area_stats(peak)
					
					write_peak(
							outputCSV, name, CAS,
							right_rt_mean=right_rt_mean,
							right_rt_stdev=right_rt_stdev,
							right_area_mean=right_area_mean,
							right_area_stdev=right_area_stdev,
							a_value=a_value
							)
		
		# outputCSV.write(f"{name};{CAS};;;;;;;;;;;")  # Name;CAS Number
		#
		# outputCSV.write(f"{right_rt_mean};{right_rt_stdev};{right_rt_stdev / right_rt_mean};;")
		# 	Mean RT right;STDEV RT right;%RSD RT right;;
		# outputCSV.write(f"{right_area_mean};{right_area_stdev};{right_area_stdev / right_area_mean};;")
		# 	Mean Area right;STDEV Area right;%RSD Area right;;
		#
		# outputCSV.write(";;;;;;;;;;;;;;;;")
		# outputCSV.write("\n")
		
		else:
			print("No peaks were found in common")
		
		# TODO: peaks that only appear in left or right sample
		
		outputCSV.close()
		
		time.sleep(3)
		
		append_to_xlsx(
				output_filename + ".CSV",
				output_filename + ".xlsx",
				"Comparison",
				overwrite=True,
				separator=";",
				toFloats=True
				)
		
		comparison_format(output_filename + ".xlsx")
	
	def make_archive(self):
		"""
		Make Single Tarball Containing All Files
		"""
		
		import tarfile
		import datetime
		
		tar = tarfile.open(
				os.path.join(
						self.config.results_dir,
						f"{self.comparison_name}_COMPARISON_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.tar.gz"
						),
				mode='w:gz'
				)
		
		for sample in [self.left_sample, self.right_sample]:
			tar.add(
					os.path.join(self.config.results_dir, f"{sample}_FINAL.xlsx"),
					arcname=f"{sample}.xlsx")
			tar.add(
					os.path.join(self.config.charts_dir, sample),
					arcname=f'Charts/{sample}')
			tar.add(
					os.path.join(self.config.spectra_dir, sample),
					arcname=f'Spectra/{sample}')
			tar.add(
					os.path.join(self.config.csv_dir, f"{sample}_MERGED.csv"),
					arcname=f"{sample}_Merged.CSV")
			tar.add(
					os.path.join(self.config.csv_dir, f"{sample}_MATCHES.csv"),
					arcname=f"{sample}_Matches.CSV")
			tar.add(
					os.path.join(self.config.csv_dir, f"{sample}_STATISTICS_FULL.csv"),
					arcname=f'Statistics/{sample}_Statistics_Full.CSV')
			tar.add(
					os.path.join(self.config.csv_dir, f"{sample}_STATISTICS.csv"),
					arcname=f'Statistics/{sample}_Statistics.CSV')
			tar.add(
					os.path.join(self.config.csv_dir, f"{sample}_STATISTICS_LIT.csv"),
					arcname=f'Statistics/{sample}_Statistics_Lit.CSV')
			tar.add(
					os.path.join(self.config.csv_dir, f"{sample}_peak_data.json"),
					arcname=f"{sample}_Peak_Data.json")
			tar.add(
					os.path.join(self.config.results_dir, f"{sample}.info"),
					arcname=f"{sample}.info")
		
		for prefix in self.left_prefixList + self.right_prefixList:
			tar.add(
					os.path.join(self.config.expr_dir, f"{prefix}_tic.dat"),
					arcname=f"Experiments/{prefix}_tic.dat")
			tar.add(
					os.path.join(self.config.expr_dir, f"{prefix}_peaks.dat"),
					arcname=f"Experiments/{prefix}_peaks.dat")
			tar.add(
					os.path.join(self.config.expr_dir, f"{prefix}.expr"),
					arcname=f"Experiments/{prefix}_expr.dat")
			
			tar.add(
					os.path.join(self.config.csv_dir, f"{prefix}_COMBINED.csv"),
					arcname=f"Combined/{prefix}.CSV")
		
		tar.add(
				os.path.join(self.config.results_dir, f"{self.comparison_name}_COMPARISON.CSV"),
				arcname="Comparison.CSV")
		
		comparison_xlsx = f"{self.comparison_name}_COMPARISON.xlsx"
		
		tar.add(
				os.path.join(self.config.results_dir, comparison_xlsx),
				arcname="Comparison.xlsx")
		tar.add(
				os.path.join(self.config.charts_dir, "Comparison"), arcname='Charts/Comparison')
		
		print(f"""Saved as: {os.path.join(self.config.results_dir, comparison_xlsx)}""")


def write_peak(
		outputCSV, name, CAS,
		left_rt_mean='', left_rt_stdev='', left_rt_n='',
		left_area_mean='', left_area_stdev='', left_area_n='',
		right_rt_mean='', right_rt_stdev='', right_rt_n='',
		right_area_mean='', right_area_stdev='', right_area_n='',
		mf_mean='', mf_stdev='', a_value=0.05
		):
	"""
	
	:param outputCSV:
	:type outputCSV:
	:param name:
	:type name: str
	:param CAS:
	:type CAS: str
	:param left_rt_mean:
	:type left_rt_mean: str or numpy.array
	:param left_rt_stdev:
	:type left_rt_stdev: str or numpy.array
	:param left_rt_n:
	:type left_rt_n: str or numpy.array
	:param left_area_mean:
	:type left_area_mean: str or numpy.array
	:param left_area_stdev:
	:type left_area_stdev: str or numpy.array
	:param left_area_n:
	:type left_area_n: str or numpy.array
	:param right_rt_mean:
	:type right_rt_mean: str or numpy.array
	:param right_rt_stdev:
	:type right_rt_stdev: str or numpy.array
	:param right_rt_n:
	:type right_rt_n: str or numpy.array
	:param right_area_mean:
	:type right_area_mean: str or numpy.array
	:param right_area_stdev:
	:type right_area_stdev: str or numpy.array
	:param right_area_n:
	:type right_area_n: str or numpy.array
	:param mf_mean:
	:type mf_mean: str or numpy.array
	:param mf_stdev:
	:type mf_stdev: str or numpy.array
	:param a_value:
	:type a_value: float
	
	:return:
	:rtype:
	"""
	
	outputCSV.write(f"{name};{CAS};;;")  # Name;CAS Number;;;
	
	def parse_rsd(stdev, mean):
		if any([stdev == '', mean == '']):
			return ''
		else:
			return stdev / mean
	
	left_rt_rsd = parse_rsd(left_rt_stdev, left_rt_mean)
	left_area_rsd = parse_rsd(left_area_stdev, left_area_mean)
	
	right_rt_rsd = parse_rsd(right_rt_stdev, right_rt_mean)
	right_area_rsd = parse_rsd(right_area_stdev, right_area_mean)
	
	outputCSV.write(
			f"{left_rt_mean};{left_rt_stdev};{left_rt_rsd};;"
			)  # Mean RT left;STDEV RT left;%RSD RT left;;
	outputCSV.write(
			f"{left_area_mean};{left_area_stdev};{left_area_rsd};;"
			)  # Mean Area left;STDEV Area left;%RSD Area left;;
	
	outputCSV.write(
			f"{right_rt_mean};{right_rt_stdev};{right_rt_rsd};;"
			)  # Mean RT right;STDEV RT right;%RSD RT right;;
	outputCSV.write(
			f"{right_area_mean};{right_area_stdev};{right_area_rsd};;"
			)  # Mean Area right;STDEV Area right;%RSD Area right;;
	
	any_value_null = any([
			left_rt_mean == '', left_rt_n == '', left_rt_stdev == '',
			right_rt_mean == '', right_rt_n == '', right_rt_stdev == ''
			])
	
	if any_value_null:
		rt_t_stat, rt_p_val, rt_result = '', '', ''
		area_t_stat, area_p_val, area_result = '', '', ''
		area_t_stat_w, area_p_val_w, area_result_w = '', '', ''
		
	else:
		"""Retention Time t-statistics"""
		# Independent 2 Samples t-test
		rt_t_stat, rt_p_val, rt_result = t_test(
				left_rt_mean, left_rt_n, left_rt_stdev,
				right_rt_mean, right_rt_n, right_rt_stdev,
				a_value)
		
		"""Peak Area t-statistics"""
		# Independent 2 Samples t-test
		area_t_stat, area_p_val, area_result = t_test(
				left_area_mean, left_area_n, left_area_stdev,
				right_area_mean, right_area_n, right_area_stdev,
				a_value)
		
		# Welch's t-test
		area_t_stat_w, area_p_val_w, area_result_w = t_test(
				left_area_mean, left_area_n, left_area_stdev,
				right_area_mean, right_area_n, right_area_stdev,
				a_value, True)
		
	outputCSV.write(f"{rt_t_stat}; {rt_p_val}; {rt_result};;")
	outputCSV.write(f"{area_t_stat}; {area_p_val}; {area_result};;")
	outputCSV.write(f"{area_t_stat_w}; {area_p_val_w}; {area_result_w};;")
	
	"""MS Comparison"""
	if any([mf_stdev == '', mf_mean == '']):
		mf_rsd = ''
	else:
		mf_rsd = (mf_stdev / mf_mean)
		
	outputCSV.write(f"{mf_mean};{mf_stdev};{mf_rsd};;")
	
	outputCSV.write("\n")
	
	return


def comparison_format(input_file):  # Formatting for Final Output
	print('\nGenerating XLSX Output...')
	
	"""From http://openpyxl.readthedocs.io/en/default/"""
	wb = load_workbook(input_file)
	
	# grab the active worksheet
	ws = wb["Comparison"]
	
	# ws.cell(column=1, row=1).value = u"Threshold α={}".format(a_value)
	
	number_format_list = {
			'E': '0.000', 'F': '0.000000', 'G': '0.00%', 'I': '0.00', 'J': '0.00', 'K': '0.00%',  # left Sample
			'M': '0.000', 'N': '0.000000', 'O': '0.00%', 'Q': '0.00', 'R': '0.00', 'S': '0.00%',  # right sample
			'U': '0.000000', 'V': '0.000000',  # RT t-test
			'Y': '0.000000', 'Z': '0.000000',  # Area t-test
			'AC': '0.000000', 'AD': '0.000000',  # Area Welch's t-test
			'AG': '0.0', 'AH': '0.0000', 'AI': '0.00%',  # MS Comparison
			}
	spacer_width = 1
	
	width_list = {
			'B': 12, 'C': spacer_width * 2, 'D': 0,  # Name and CAS
			'E': 8, 'F': 11, 'G': 7, 'H': spacer_width, 'I': 15, 'J': 14, 'K': 8, 'L': spacer_width * 2,  # left Sample
			'M': 8, 'N': 11, 'O': 7, 'P': spacer_width, 'Q': 15, 'R': 14, 'S': 8, 'T': spacer_width * 2,  # right Sample
			'U': 11, 'V': 10, 'W': 7, 'X': spacer_width,  # RT t-test
			'Y': 11, 'Z': 10, 'AA': 7, 'AB': spacer_width,  # Area t-test
			'AC': 11, 'AD': 10, 'AE': 7, 'AF': spacer_width,  # Area Welch's t-test
			'AG': 8, 'AH': 10, 'AI': 8,  # MS Comparison
			}
	
	alignment_list = make_column_property_dict({'5': 'right'}, {'B': 'center'}, repeat=32, length=1)
	
	for column in ['W', 'AA', 'AE']:
		alignment_list[column] = 'center'
	
	format_sheet(ws, number_format_list, width_list, alignment_list)  #
	
	"""Header"""
	# Row 1
	ws.merge_cells("E1:K1")
	ws.merge_cells("M1:S1")
	ws.merge_cells('U1:AE1')
	
	# Row 2
	for offset in range(7):
		merge_string = get_column_letter(5 + (4 * offset)) + '2:' + get_column_letter(7 + (4 * offset)) + '2'
		# print(merge_string)
		ws.merge_cells(merge_string)
	
	# MS Comparisons
	ws.cell(column=33, row=1).value = "MS Comparisons"
	ws.merge_cells('AG1:AI2')
	
	# Centering text
	format_header(ws, make_column_property_dict(
			{"2": "center"}, repeat=35,
			length=1), 1, 2)
	
	# Save the file
	wb.save(input_file)
	return


def ms_comparisons(left_ms_data, right_ms_data):
	"""
	
	:param left_ms_data:
	:type left_ms_data:
	:param right_ms_data:
	:type right_ms_data:
	
	:return:
	:rtype:
	"""
	
	perms = []
	for perm in product(left_ms_data, right_ms_data):
		top_spec = numpy.column_stack((perm[0].mass_list, perm[0].mass_spec))
		bottom_spec = numpy.column_stack((perm[1].mass_list, perm[1].mass_spec))
		perms.append(
				spectrum_similarity.SpectrumSimilarity(
						top_spec, bottom_spec, t=0.25, b=1,
						xlim=(45, 500), x_threshold=0,
						print_graphic=False
						)[0] * 1000
				)
	
	# print(numpy.nanmean(perms))
	# print(numpy.nanstd(perms))
	return numpy.nanmean(perms), numpy.nanstd(perms)


def t_test(left_mean, left_n, left_stdev, right_mean, right_n, right_stdev, a_value=0.01, welch=False):
	"""
	
	:param left_mean:
	:type left_mean:
	:param left_n:
	:type left_n:
	:param left_stdev:
	:type left_stdev:
	:param right_mean:
	:type right_mean:
	:param right_n:
	:type right_n:
	:param right_stdev:
	:type right_stdev:
	:param a_value:
	:type a_value:
	:param welch:
	:type welch:
	
	:return:
	:rtype:
	"""
	
	from scipy import stats
	
	if not any(x == 0.0 for x in [left_mean, left_n, right_mean, right_n]):
		t_stat, p_val = stats.ttest_ind_from_stats(
				left_mean,
				left_stdev,
				left_n,
				right_mean,
				right_stdev,
				right_n,
				not welch
				)
		
		if str(t_stat) == "inf":
			t_stat = "#"
		
		result = "Diff" if p_val <= a_value else "Same"
	
	else:
		t_stat = ''
		p_val = ''
		result = "Diff"
	
	return t_stat, p_val, result


default_styles = [
		"D",  # diamond
		"s",  # square
		"X",  # bold cross
		"^",  # triangle
		"d",  # diamond
		"h",  # hexagon
		"o",  # dot
		"v",  # down triangle
		"<",  # left triangle
		">",  # right triangle
		]

default_bw_colours = [
		"DarkRed",
		"DeepSkyBlue",
		"Green",
		"Purple",
		"Black",
		"Sienna",
		"MediumTurquoise",
		"DarkOliveGreen",
		"m",
		"DarkSlateBlue",
		"OrangeRed",
		"CadetBlue",
		"Olive",
		"Red",
		"Blue",
		"PaleVioletRed",
		"Teal",
		"y",
		"RoyalBlue",
		]


def multiple_project_charts(
		projectList, show_outliers=True, show_raw_data=False, err_bar="range",
		outlier_mode="2stdev", leg_cols=1, column_width=4, filetypes=default_filetypes,
		styles=None, bw_colours=None, radar_colours=default_colours,
		radar_figsize=(4, 9), bw_figsize=None, mode="display", radar_use_log=False
		):
	"""All Samples Comparison (Charts Only)"""
	# TODO File names for charts
	
	print("This function needs reworking")
	return
	
	if styles is None:
		styles = default_styles[:]
	if bw_colours is None:
		bw_colours = default_bw_colours[:]
	
	"""Chart Data"""
	chart_data = pandas.concat(
			[
					pandas.read_csv(
							os.path.join(csv_dir, f"{sample}_CHART_DATA.csv"),
							sep=";",
							index_col=0
							) for sample in projectList
					],
			axis=1,
			sort=False
			)
	
	chart_data.drop("Compound Names", axis=1, inplace=True)
	
	# determine order of compounds on graph
	for compound in chart_data.index.values:
		chart_data["Count"] = chart_data.apply(
				df_count,
				args=(
						[f"{sample} Peak Area" for sample in projectList],
						), axis=1)
	
	chart_data['Compound Names'] = chart_data.index
	chart_data = chart_data.sort_values(['Count', 'Compound Names'])
	chart_data.fillna(0, inplace=True)
	
	radar_chart(chart_data, projectList, use_log=10)
	
	sample_list = []
	
	for project in projectList:
		with open(os.path.join(results_dir, f"{project}.info"), "r") as info_file:
			sample_list.append((project, [x.rstrip("\r\n") for x in info_file.readlines()]))
	
	print(sample_list)
	
	box_whisker(
			chart_data,
			sample_list,
			groupings=groupings
			)


def arguments():
	# Command line switches
	import argparse
	parser = argparse.ArgumentParser()
	
	parser.add_argument("-p", "--projects", help="List of projects to compare.", nargs=2, required=True)
	# parser.add_argument("-g", "--groups", help="List of groups for comparison chart.", nargs='+')
	parser.add_argument("--info", help="Show program info.", action='store_true')
	
	args = parser.parse_args()
	
	if len(sys.argv) == 1:
		print("No options supplied.")
		print('')
		parser.print_help(sys.stderr)
		sys.exit(1)
	
	return args


if __name__ == '__main__':
	import platform
	from domdf_python_tools.paths import maybe_make
	from GSMatch.GSMatch_Core.Config import GSMConfig
	
	clear()  # clear the display
	startup_string = f"""

{program_name} Version {__version__} running on {platform.system()}.
Copyright {__copyright__} Dominic Davis-Foster.
"""
	
	print(startup_string)
	
	# Load Configuration
	
	config = GSMConfig("config.ini")
	
	a_value = 0.01
	
	args = arguments()
	
	projectList = args.projects
	# projectList = ["Eley Contact", "Geco Rifle"]
	# projectList = ["Eley Contact", "Eley Contact Fired", "Geco Rifle", "Geco Rifle Fired"]
	#
	# if args.groups:
	# 	groupings = args.groups
	# else:
	# groupings = projectList[::2]
	
	br()
	
	if len(projectList) < 2:
		print("Error: Require Two or More Projects to Compare")
		sys.exit(1)
	
	elif len(projectList) == 2:
		"""Two Samples Comparison"""
		pass
		# comparison(*projectList, Dw=config.rt_modulation, Gw=config.gap_penalty, min_peaks=config.min_peaks)
		
		# else:
		"""All Samples Comparison (Charts Only)"""
	# multiple_project_charts(projectList)
	
	"""Charts"""
	"""print("\nGenerating Charts")
	maybe_make(os.path.join(charts_dir, "+".join(projectList)+"COMPARISON"))
	

	
	radar_chart(chart_data, [lot_name], use_log=10,
				mode=os.path.join(charts_dir, lot_name, "radar_log10_peak_area"))
	radar_chart(chart_data, [lot_name], use_log=False,
				mode=os.path.join(charts_dir, lot_name, "radar_peak_area"))
	mean_peak_area_multiple(chart_data, [lot_name],
							mode=os.path.join(charts_dir, lot_name, "mean_peak_area"))
	peak_area(chart_data, prefixList, use_log=10,
			  mode=os.path.join(charts_dir, lot_name, "log10_peak_area_percentage"))
	
	with open(os.path.join(results_dir, f"{lot_name}.info"), "w") as info_file:
		for prefix in prefixList:
			info_file.write(f"{prefix}\n")"""
	
	print("\nComplete.")


def find_aligned_peaks(sample, prefixList, peak_data, rt_alignment, ms_alignment):
	aligned_peaks = []
	# print(f"{sample} Peaks")
	for index, aligned_peak in enumerate(rt_alignment[sample]):
		found_peak = False
		for peak in peak_data:
			# print(aligned_peak, peak["average_rt"])
			if peak["average_rt"] == aligned_peak:
				# print(peak)
				# for key in peak:
				# 	print(f"{key}: {peak[key]}")
				peak["ms_list"] = [ms_alignment.iloc[index][prefix] for prefix in prefixList]
				aligned_peaks.append(peak)
				found_peak = True
				break
		if not found_peak:
			aligned_peaks.append(None)
	
	return aligned_peaks


def get_peak_rt_stats(peak):
	rt_mean = peak['average_rt']
	rt_stdev = numpy.nanstd(peak['rt_data'])
	rt_n = len(peak['rt_data'])
	
	return rt_mean, rt_stdev, rt_n


def get_peak_area_stats(peak):
	area_mean = peak['average_peak_area']
	area_stdev = numpy.nanstd(peak['area_data'])
	area_n = len(peak['area_data'])
	
	return area_mean, area_stdev, area_n


def write_output_data(peak, outputCSV, right_peak=False):
	name = peak['hits'][0]['Name']
	CAS = peak['hits'][0]['CAS']
	
	rt_mean, rt_stdev, rt_n = get_peak_rt_stats(peak)
	area_mean, area_stdev, area_n = get_peak_area_stats(peak)
	
	outputCSV.write(f"{name};{CAS};;;")  # Name;CAS Number;;;
	if right_peak:
		outputCSV.write(";;;;;;;;")
	
	outputCSV.write(
			f"{rt_mean};{rt_stdev};{rt_stdev / rt_mean};;"
			)  # Mean RT;STDEV RT;%RSD RT;;
	outputCSV.write(
			f"{area_mean};{area_stdev};{area_stdev / area_mean};;"
			)  # Mean Area;STDEV Area;%RSD Area;;
	
	outputCSV.write(";;;;;;;;;;;;;;;;")
	if not right_peak:
		outputCSV.write(";;;;;;;;")
	outputCSV.write("\n")

#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
#  GSMatch_Rework.py
#  
#  Copyright 2017-2019 Dominic Davis-Foster <dominic@davis-foster.co.uk>
#  
#  This program is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 3 of the License, or
#  (at your option) any later version.
#  
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#  
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#   

"""GunShotMatch Rework

Program for the analysis of OGSR samples to identify matching compounds 
between samples. 
	
Samples to be analysed can be passed following the -f/--samples argument, 
or listed in a file called "list.txt" in the base directory. Each sample 
name must be on a new line. For more information read the documentation.
	
The files must be in the subdirectory CSV and called 
		e.g. GECO_Case2_MS.CSV and GECO_Case2_GC.CSV
"""

program_name ="GunShotMatch"
__version__ = "1.0.0 rework"
copyright = "2017-2019"


# Imports
from utils import timing # Times the program
import sys
sys.path.append("..")

import os
import re
import csv
import json
import math
import time
import numpy
import pandas
import locale
import shutil
import atexit
import tarfile
import logging
import datetime
import platform
import warnings
import operator
import traceback
import itertools

from collections import Counter
from multiprocessing import Pool
from itertools import chain, permutations

from gsm_core import GSMConfig, get_ms_alignment, get_peak_alignment
from utils.charts import PlotSpectrum, box_whisker_wrapper, radar_chart_wrapper, mean_peak_area_wrapper, peak_area_wrapper
from utils.pynist import *
from utils import DirectoryHash, pynist, SpectrumSimilarity
from utils.mathematical import rounders
from utils.terminal import clear, br
from utils.paths import maybe_make
from utils.helper import as_text
from utils.spreadsheet import format_header, format_sheet, make_column_property_list, append_to_xlsx

from pyms.GCMS.Class import IonChromatogram
from pyms.GCMS.IO.JCAMP.Function import JCAMP_reader
from pyms.GCMS.Function import build_intensity_matrix, build_intensity_matrix_i
from pyms.Noise.Window import window_smooth, window_smooth_im
from pyms.Noise.Analysis import window_analyzer
from pyms.Noise.SavitzkyGolay import savitzky_golay, savitzky_golay_im
from pyms.Utils.IO import dump_object
from pyms.Baseline.TopHat import tophat, tophat_im
from pyms.Deconvolution.BillerBiemann.Function import BillerBiemann, rel_threshold, num_ions_threshold
from pyms.Peak.IO import store_peaks, load_peaks
from pyms.Peak.Function import peak_sum_area, peak_pt_bounds
from pyms.Peak.List.DPA.Class import PairwiseAlignment
from pyms.Peak.List.DPA.Function import align_with_tree, exprl2alignment
from pyms.Experiment.IO import store_expr, load_expr
from pyms.Experiment.Class import Experiment

from openpyxl import Workbook, worksheet, load_workbook		# https://openpyxl.readthedocs.io/en/default/
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string


# Setup for reading strings with thousand seperators as floats
# From https://stackoverflow.com/a/31074271
locale.setlocale(locale.LC_ALL, "")

verbose = False

"""From https://stackoverflow.com/questions/14463277/how-to-disable-python-warnings"""
#with warnings.catch_warnings():
#	warnings.simplefilter("ignore")	

"""Import JCAMP-DX Files"""
def quantitative_processing(jcamp_file, log_stdout = True):

	# Parameters
	sample_name = os.path.splitext(os.path.basename(jcamp_file))[0]
	
	"""Log Stdout to File"""
	if log_stdout:
		sys.stdout = open(os.path.join(LOG_DIRECTORY,sample_name + ".log"), "w")
	
	
	data = JCAMP_reader(jcamp_file)
	
	# list of all retention times, in seconds
	times = data.get_time_list()
	# get Total Ion Chromatogram
	tic = data.get_tic()
	# RT Range, time step, no. scans, min, max, mean and median m/z
	data.info()
	
	
	# Mass Binning
	im = build_intensity_matrix_i(data)  # convert to intensity matrix
	
	print(" Minimum m/z bin: {}".format(im.get_min_mass()))
	print(" Maximum m/z bin: {}".format(im.get_max_mass()))
	
	# Crop masses
	min_mass, max_mass, *_ = mass_range
	if min_mass < im.get_min_mass():
		min_mass = im.get_min_mass()
	if max_mass > im.get_max_mass():
		max_mass = im.get_max_mass()
	im.crop_mass(min_mass, max_mass)
	
	## Data filtering
	
	# Note that Turbomass does not use smoothing for qualitative method.
	
	# Top-hat baseline Correction seems to bring down noise,
	#  retaning shapes, but keeps points on actual peaks
	
	n_scan, n_mz = im.get_size()
	for ii in range(n_mz):
		# print("\rWorking on IC#", ii+1, '  ',end='')
		ic = im.get_ic_at_index(ii)
		ic_smooth = savitzky_golay(ic)
		ic_bc = tophat(ic_smooth, struct=tophat_struct)
		im.set_ic_at_index(ii, ic_bc)
		
	# Peak Detection based on Biller and Biemann, 1974, with a window
	#  of n points, and combining y scans if they apex next to each other
	peak_list = BillerBiemann(im, points=bb_points, scans=bb_scans)
	
	print(" Number of peaks identified before filtering: {}".format(len(peak_list)))
	
	# Filtering peak lists with automatic noise filtering
	noise_level = window_analyzer(tic)
	peak_list = num_ions_threshold(peak_list, noise_thresh, noise_level)
	print(" Number of peaks identified: {}".format(len(peak_list)))
	
	filtered_peak_list = []
	
	for peak in peak_list:
		apex_mass_list = peak.get_mass_spectrum().mass_list
		apex_mass_spec = peak.get_mass_spectrum().mass_spec
		base_peak_intensity = max(apex_mass_spec)
		base_peak_index = [index for index, intensity in enumerate(apex_mass_spec) if intensity == base_peak_intensity][
			0]
		base_peak_mass = apex_mass_list[base_peak_index]
		# print(base_peak_mass)
		if base_peak_mass in base_peak_filter:
			continue  # skip the peak if the base peak is at e.g. m/z 73, i.e. septum bleed
		
		area = peak_sum_area(im, peak)
		peak.set_area(area)
		filtered_peak_list.append(peak)
	
	# Save the TIC and Peak List
	tic.write(os.path.join(EXPERIMENTS_DIRECTORY, "{}_tic.dat".format(sample_name)), formatting=False)
	store_peaks(filtered_peak_list, os.path.join(EXPERIMENTS_DIRECTORY, "{}_peaks.dat".format(sample_name)))
	
	# Create an experiment
	expr = Experiment(sample_name, filtered_peak_list)
	expr.sele_rt_range(["{}m".format(target_range[0]), "{}m".format(target_range[1])])
	store_expr(os.path.join(EXPERIMENTS_DIRECTORY, "{}.expr".format(sample_name)), expr)


"""Qualitative Processing"""
def qualitative_processing(sample_name, rt_list):
	combined_csv_file = os.path.join(CSV_DIRECTORY, "{}_COMBINED.csv".format(sample_name))
	
	number_of_peaks = 80
	
	# Load saved TIC
	time_list = []
	intensity_list = []
	
	with open(os.path.join(EXPERIMENTS_DIRECTORY, "{}_tic.dat".format(sample_name))) as tic_file:
		ticreader = csv.reader(tic_file, delimiter=" ")
		for row in ticreader:
			row = list(filter(None, row))
			intensity_list.append(float(row[1]))
			time_list.append(float(row[0]))
	
	intensity_array = numpy.array(intensity_list)
	tic = IonChromatogram(intensity_array, time_list)
	
	peak_list = load_peaks(os.path.join(EXPERIMENTS_DIRECTORY, "{}_peaks.dat".format(sample_name)))
	
	# Peak Areas
	peak_area_list = []
	for peak in peak_list:
		area = peak.get_area()
		peak_area_list.append(area)
	
	# Write to GunShotMatch Combine-like CSV file
	combine_csv = open(combined_csv_file, "w")
	
	combine_csv.write(sample_name)
	combine_csv.write("\n")
	
	combine_csv.write("Retention Time;Peak Area;;Lib;Match;R Match;Name;CAS Number;Notes\n")
	
	report_buffer = []
	
	# Filter to those peaks present in all samples, by UID
	for peak in peak_list:
		# if str(rounders(peak.get_rt()/60,"0.000")) in rt_list:
		# print(peak.get_rt()/60.0)
		if peak.get_rt() / 60.0 in rt_list:
			report_buffer.append(['',
								  # rounders(peak.get_rt()/60,"0.000"),
								  (peak.get_rt() / 60),
								  '',
								  peak.get_mass_spectrum(),
								  # '{:,}'.format(rounders(peak.get_area()/60,"0.0"))
								  '{:,}'.format(peak.get_area() / 60)
								  ])
	
	report_buffer = report_buffer[::-1]  # Reverse list order
	
	report_buffer = report_buffer[:number_of_peaks]
	
	report_buffer.sort(key=operator.itemgetter(1))
	
	for row in report_buffer:
		index = report_buffer.index(row)
		
		# if index == 19:
		
		ms = row[3]
		
		create_msp("{}_{}".format(sample_name, row[1]), ms.mass_list, ms.mass_spec)
		n_hits = 10
		matches_dict = nist_ms_comparison("{}_{}".format(sample_name, row[1]), ms.mass_list, ms.mass_spec, n_hits)
		
		combine_csv.write("{};{};Page {} of 80;;;;;;{}\n".format(row[1], row[4], index + 1, row[2]))
		
		for hit in range(1, n_hits + 1):
			combine_csv.write(';;{};{};{};{};{};{};\n'.format(hit,
															  matches_dict["Hit{}".format(hit)]["Lib"],
															  matches_dict["Hit{}".format(hit)]["MF"],
															  matches_dict["Hit{}".format(hit)]["RMF"],
															  matches_dict["Hit{}".format(hit)]["Name"].replace(";",
																												":"),
															  matches_dict["Hit{}".format(hit)]["CAS"],
															  ))
		
		time.sleep(2)
	
	combine_csv.close()
	
	return 0

def create_msp(sample_name, mass_list, mass_spec):
	"""Generate .MSP files for NIST MS Search"""
	
	msp_file = open(os.path.join(MSP_DIRECTORY, sample_name + ".MSP"), "w")
	msp_file.write("Name: {}\n".format(sample_name))
	msp_file.write("Num Peaks: {}\n".format(len(mass_list)))
	for mass, intensity in zip(mass_list, mass_spec):
		msp_file.write("{} {},\n".format(rounders(mass, "0.0"), intensity))
	msp_file.close()

def nist_ms_comparison(sample_name, mass_list, mass_spec, n_hits=5):
	data_dict = {}
	
	try:
		pynist.generate_ini(nist_path, "mainlib", n_hits)
		
		def remove_chars(input_string):
			for i in range(n_hits + 1):
				input_string = input_string.replace("Hit {}  : ", "")
			
			return input_string.replace("MF:", "") \
				.replace(":", "").replace("<", "").replace(">", "") \
				.replace(" ", "").replace(sample_name, "")
		
		raw_output = pynist.nist_db_connector(nist_path, os.path.join(MSP_DIRECTORY, "{}.MSP".format(sample_name)))
		
		# Process output
		for i in range(n_hits + 1):
			# print(raw_output)
			raw_output = raw_output.replace("Hit {}  : ".format(i), "Hit{};".format(i)).replace("Hit {} : ".format(i),
																								"Hit{};".format(
																									i)).replace(
				"Hit {}: ".format(i), "Hit{};".format(i))
		raw_output = raw_output.replace("<<", '"').replace(">>", '"').split("\n")
		# print(raw_output)
		
		# for row in raw_output:
		#	print(row)
		
		matches_dict = {}
		
		for i in range(1, n_hits + 1):
			row = list(csv.reader([raw_output[i]], delimiter=";", quotechar='"'))[0]
			# print(row)
			matches_dict[row[0]] = {"Name": row[1], "MF": (row[3].replace("MF:", '').replace(" ", '')),
									"RMF": (row[4].replace("RMF:", '').replace(" ", '')),
									"CAS": (row[6].replace("CAS:", '').replace(" ", '')),
									"Lib": (row[8].replace("Lib:", '').replace(" ", ''))}
	# print(matches_dict[row[0]])
	
	# for match in matches_dict:
	#	print(match)
	#	print(matches_dict[match])
	
	except:
		traceback.print_exc()  # print the error
		pynist.reload_ini(nist_path)
		sys.exit(1)
	
	print("\r\033[KSearch Complete")
	pynist.reload_ini(nist_path)
	return matches_dict


"""Merge Function"""
def Merge():
	if PL_len <= 1:
		print("Require two or more samples to combine. Check ./list.txt or --samples and try again.")
		print("")
		sys.exit(print(__doc__))
	
	merge_list = ()
	
	for prefix in prefixList:
		# read combined file
		with open(os.path.join(CSV_DIRECTORY, "{}_COMBINED.csv".format(prefix)), "r") as combined_file:
			combined_file_reader = csv.reader(combined_file, delimiter=";", quotechar='"')
			merge_list = merge_list + ([row for row in combined_file_reader],)
	
	merge_list = merge_list + (["\n"] * len(merge_list[0]),)
	
	merged_csv = []
	for prefix in prefixList:
		merged_csv.append([prefix, '', '', '', '', '', '', '', ''])
	
	# from https://stackoverflow.com/questions/7946798/interleave-multiple-lists-of-the-same-length-in-python
	merged_csv += [val for tup in zip(*merge_list) for val in tup][PL_len:]
	
	# Write output
	with open(os.path.join(CSV_DIRECTORY, "{}_MERGED.csv".format(lot_name)), "w") as f:
		f.write(';'.join(list(itertools.chain(*merged_csv))).replace("\n;", "\n"))
# f.write(str(list(itertools.chain(*merged_csv))).replace("[","").replace("]","").replace(",",";").replace("'\n'","\n"))
# Semicolon to match JIGSAW output


"""Counter"""
def Match_Counter(ms_comp_list,seperator=";"):	
	# Also generates final output

	if PL_len == 1:
		print("Require two or more samples to process. Check ./list.txt or --samples and try again.")
		print("")
		#sys.exit(print(__doc__))
	
	f = open(os.path.join(CSV_DIRECTORY,"{}_MERGED.csv".format(lot_name)),"r")		#Open merged output of lot of samples

	csv_f = csv.reader(f, delimiter=seperator)
	csv_input = []
	
	for row in csv_f:
		csv_input.append(row)
	peak_data = []
	
	n_hits = 10
	
	for peak in csv_input[2::11]:
		rt_data = [locale.atof(x) for x in peak[0::9] if x not in [None,'']]
		area_data = [locale.atof(x) for x in peak[1::9] if x is not None]
		index = (csv_input.index(peak))
		hits = []
		names = []
		output_data = {"average_rt":numpy.mean(rt_data), "average_peak_area":numpy.mean(area_data), "rt_data":rt_data, "area_data":area_data}
		for hit in range(n_hits):
			for prefix in prefixList:
				hits.append([prefix, {"hit":csv_input[index+hit+1][2],
									  "MF":csv_input[index+hit+1][4],
									  "RMF":csv_input[index+hit+1][5],
									  "Name":csv_input[index+hit+1][6],
									  "CAS":csv_input[index+hit+1][7]
									 }])		
				names.append(csv_input[index+hit+1][6])
				csv_input[index+hit+1] = csv_input[index+hit+1][9:]

		names.sort()
		"""From https://stackoverflow.com/questions/2600191/how-to-count-the-occurrences-of-a-list-item"""
		names_count = Counter(names)
		
		hits_data = []
		
		# get average match factor &c. for each compound
		for compound in names_count:
			mf_data = []
			rmf_data = []
			hit_num_data = []

			for prefix in prefixList:
				length_before = len(mf_data)
				for hit in hits:

					if (hit[1]["Name"] == compound) and (hit[0]==prefix):

						mf_data.append(float(hit[1]["MF"]))
						rmf_data.append(float(hit[1]["RMF"]))
						hit_num_data.append(int(hit[1]["hit"]))
						CAS = hit[1]["CAS"]
				if len(mf_data) == length_before:
					mf_data.append(numpy.nan)
					rmf_data.append(numpy.nan)
					hit_num_data.append(numpy.nan)
				
			hits_data.append({"Name":compound,
							  "Count":names_count[compound],
							  "average_MF": numpy.nanmean(mf_data),
							  "average_RMF":numpy.nanmean(rmf_data),
							  "average_hit":numpy.nanmean(hit_num_data),
							  "CAS":CAS,
							  "mf_data":mf_data,
							  "rmf_data":rmf_data,
							  "hit_num_data":hit_num_data
							})
		
		hits_data = sorted(hits_data, key=lambda k: (k["Count"], k["average_MF"], k["average_hit"]), reverse=True)
		output_data["hits"] = hits_data[:5]
		peak_data.append(output_data)
	
	
	"""Matches Sheet"""
	with open(os.path.join(CSV_DIRECTORY, lot_name + "_MATCHES.csv"),"w") as matches_csv_output:
		matches_csv_output.write("{};;;".format(lot_name))
	
		for prefix in prefixList:
			matches_csv_output.write("{};;;".format(prefix))
		matches_csv_output.write("\n")
	
		matches_csv_output.write("Retention Time;Peak Area;;" + "Page;RT;Area;"*PL_len +";Match Factor;;;;Reverse Match Factor;;;;Hit Number;;;;Retention Time;;;;Peak Area;;\n")
		matches_csv_output.write("Name;CAS Number;Freq.;" + "Hit No.;Match;R Match;"*PL_len + ";Mean;STDEV;%RSD;;Mean;STDEV;%RSD;;Mean;STDEV;%RSD;;Mean;STDEV;%RSD;;Mean;STDEV;%RSD\n")
		
		for peak in peak_data:
			matches_csv_output.write("{average_rt};{average_peak_area};;".format(**peak) 
					+ ";".join([str(val) for pair in zip(['']*PL_len,peak["rt_data"], peak["area_data"]) for val in pair]) 
					+ ";;;;;;;;;;;;;;{average_rt};".format(**peak)
					+ "{};{};;".format(numpy.nanstd(peak["rt_data"]),numpy.nanstd(peak["rt_data"])/peak["average_rt"])
					+ "{average_peak_area};".format(**peak)
					+ "{};{};;\n".format(numpy.nanstd(peak["area_data"]),numpy.nanstd(peak["area_data"])/peak["average_peak_area"])
				)
			
			for hit in peak["hits"]:
				matches_csv_output.write("{Name};{CAS};{Count};".format(**hit) 
						+ ";".join([str(val) for pair in zip(hit["hit_num_data"],hit["mf_data"],hit["rmf_data"]) for val in pair]) 
						+ ";;{average_MF};".format(**hit)
						+ "{};{};".format((numpy.nanstd(hit["mf_data"])),numpy.nanstd(hit["mf_data"])/hit["average_MF"])
						+ ";{average_RMF};".format(**hit)
						+ "{};{};".format((numpy.nanstd(hit["rmf_data"])),numpy.nanstd(hit["rmf_data"])/hit["average_RMF"])
						+ ";{average_hit};".format(**hit)
						+ "{};{};\n".format((numpy.nanstd(hit["hit_num_data"])),numpy.nanstd(hit["hit_num_data"])/hit["average_hit"])
					 )


	"""Get list of CAS Numbers for compounds reported in literature"""
	with open("./lib/CAS.txt","r") as f:
		CAS_list = f.readlines()
	for index, CAS in enumerate(CAS_list):
		CAS_list[index] = CAS.rstrip("\r\n")

		
	"""Statistics Sheet"""
	statistics_full_output = open(os.path.join(CSV_DIRECTORY, "{}_STATISTICS_FULL.csv".format(lot_name)),"w")
	statistics_output = open(os.path.join(CSV_DIRECTORY, "{}_STATISTICS.csv".format(lot_name)),"w")
	statistics_lit_output = open(os.path.join(CSV_DIRECTORY, "{}_STATISTICS_LIT.csv".format(lot_name)),"w")
	
	statistics_header = "{};;;;Retention Time;;;;Peak Area;;;;Match Factor\
		;;;;Reverse Match Factor;;;;Hit Number;;;;MS Comparison;;;\n\
		Name;CAS Number;;;Mean;STDEV;%RSD;;Mean;STDEV;%RSD;;Mean;STDEV;\
		%RSD;;Mean;STDEV;%RSD;;Mean;STDEV;%RSD;;Mean;STDEV;%RSD\n"\
		.format(lot_name)
	
	statistics_full_output.write(statistics_header)
	statistics_output.write(statistics_header)
	statistics_lit_output.write(statistics_header)
	
	# Initialise dictionary for chart data
	chart_data = {"Compound":[], "{} Peak Area".format(lot_name):[], "{} Standard Deviation".format(lot_name):[]}
	for prefix in prefixList:
		chart_data[prefix] = []
	
	for peak, ms in zip(peak_data, ms_comp_list):
		write_peak(statistics_full_output,peak, ms)
		if peak["hits"][0]["Count"] > (PL_len/2): # Write to Statistics; also need similarity > 800
			write_peak(statistics_output,peak, ms)
			if peak["hits"][0]["CAS"].replace("-","") in CAS_list: # Write to Statistics_Lit
				write_peak(statistics_lit_output,peak, ms)
				# Create Chart Data
				chart_data["Compound"].append(peak["hits"][0]["Name"])
				for prefix, area in zip(prefixList, peak["area_data"]):
					chart_data[prefix].append(area)
				chart_data["{} Peak Area".format(lot_name)].append(numpy.mean(peak["area_data"]))
				chart_data["{} Standard Deviation".format(lot_name)].append(numpy.mean(peak["area_data"]))
				

	statistics_full_output.close()
	statistics_output.close()
	statistics_lit_output.close()
	
	with open(os.path.join(CSV_DIRECTORY, "{}_peak_data.json".format(lot_name)), "w") as jsonfile:
		for dict in peak_data:
			jsonfile.write(json.dumps(dict))
			jsonfile.write("\n")
		
	time.sleep(2)
	
	"""Convert to XLSX and format"""
	# GC-MS
	append_to_xlsx(os.path.join(CSV_DIRECTORY,"{}_MERGED.csv".format(lot_name)), os.path.join(RESULTS_DIRECTORY,lot_name + "_FINAL.xlsx"), "GC-MS", overwrite=True, seperator=";", toFloats=True)
	# Counter
	append_to_xlsx(os.path.join(CSV_DIRECTORY, "{}_MATCHES.csv".format(lot_name)), os.path.join(RESULTS_DIRECTORY,lot_name + "_FINAL.xlsx"), "Matches", seperator=";", toFloats=True)
	# Statistics_Full
	append_to_xlsx(os.path.join(CSV_DIRECTORY, "{}_STATISTICS_FULL.csv".format(lot_name)), os.path.join(RESULTS_DIRECTORY,lot_name + "_FINAL.xlsx"), "Statistics_Full", seperator=";", toFloats=True)
	# Statistics
	append_to_xlsx(os.path.join(CSV_DIRECTORY, "{}_STATISTICS.csv".format(lot_name)), os.path.join(RESULTS_DIRECTORY,lot_name + "_FINAL.xlsx"), "Statistics", seperator=";", toFloats=True)
	# Statistics_Lit
	append_to_xlsx(os.path.join(CSV_DIRECTORY, "{}_STATISTICS_LIT.csv".format(lot_name)), os.path.join(RESULTS_DIRECTORY,lot_name + "_FINAL.xlsx"), "Statistics - Lit Only", seperator=";", toFloats=True)
	
	formatXLSX(os.path.join(RESULTS_DIRECTORY,lot_name + "_FINAL.xlsx"))
	
	"""Charts"""
	chart_data = pandas.DataFrame(chart_data)
	print(chart_data)
	return chart_data

def write_peak(file_object, peak, ms):
	file_object.write("{};{};{};;".format(peak["hits"][0]["Name"],peak["hits"][0]["CAS"],peak["hits"][0]["Count"]))
	file_object.write("{};{};{};;".format(peak["average_rt"],numpy.nanstd(peak["rt_data"]),numpy.nanstd(peak["rt_data"])/peak["average_rt"]))
	file_object.write("{};{};{};;".format(peak["average_peak_area"],numpy.nanstd(peak["area_data"]),numpy.nanstd(peak["area_data"])/peak["average_peak_area"]))
	file_object.write("{};{};{};;".format(peak["hits"][0]["average_MF"],numpy.nanstd(peak["hits"][0]["mf_data"]),numpy.nanstd(peak["hits"][0]["mf_data"])/peak["hits"][0]["average_MF"]))
	file_object.write("{};{};{};;".format(peak["hits"][0]["average_RMF"],numpy.nanstd(peak["hits"][0]["rmf_data"]),numpy.nanstd(peak["hits"][0]["rmf_data"])/peak["hits"][0]["average_RMF"]))
	file_object.write("{};{};{};;".format(peak["hits"][0]["average_hit"],numpy.nanstd(peak["hits"][0]["hit_num_data"]),numpy.nanstd(peak["hits"][0]["hit_num_data"])/peak["hits"][0]["average_hit"]))
	file_object.write("{};{};{};;".format(numpy.mean(ms),numpy.std(ms),numpy.std(ms)/numpy.mean(ms)))
	file_object.write("\n")


"""Between Samples Spectra Comparison"""
def ms_comparisons(ms_data):
	perms = []
	for i in permutations(prefixList, 2):
		"""from https://stackoverflow.com/questions/10201977/how-to-reverse-tuples-in-python"""
		if i[::-1] not in perms:
			perms.append(i)
	
	ms_comparison = []
	
	rows_list = []
	
	for row_idx in range(len(ms_data)):
		similarity_list = []
		rows_list.append((ms_data.iloc[row_idx], perms))
	
	with Pool(len(ms_data)) as p:
		ms_comparison = p.map(single_ms_comparison, rows_list)
	
	# for row_idx in range(len(ms_data)):
	# similarity_list = []
	# row = ms_data.iloc[row_idx]
	# for perm in perms:
	
	# top_spec = numpy.column_stack((row.loc[perm[0]].mass_list, row.loc[perm[0]].mass_spec))
	# bottom_spec = numpy.column_stack((row.loc[perm[1]].mass_list, row.loc[perm[1]].mass_spec))
	# similarity_list.append(SpectrumSimilarity.SpectrumSimilarity(top_spec, bottom_spec, t = 0.25, b = 1,
	# xlim = (45, 500), x_threshold = 0, print_graphic = False)[0]*1000)
	
	# #ms_comparison.append(rounders((numpy.mean(similarity_list)*100),"0"))
	# ms_comparison.append(similarity_list)
	
	return ms_comparison

def single_ms_comparison(args):
	row, perms = args
	
	similarity_list = []
	
	for perm in perms:
		top_spec = numpy.column_stack((row.loc[perm[0]].mass_list, row.loc[perm[0]].mass_spec))
		bottom_spec = numpy.column_stack((row.loc[perm[1]].mass_list, row.loc[perm[1]].mass_spec))
		similarity_list.append(SpectrumSimilarity.SpectrumSimilarity(top_spec, bottom_spec, t=0.25, b=1,
																	 xlim=(45, 500), x_threshold=0,
																	 print_graphic=False)[0] * 1000)
	
	# ms_comparison.append(rounders((numpy.mean(similarity_list)*100),"0"))
	return similarity_list


"""Formating Functions"""
def formatXLSX(inputFile):	# Formatting for Combined Output
	print('\nGenerating XLSX Output...')

	"""From http://openpyxl.readthedocs.io/en/default/"""
	wb = load_workbook(inputFile)
	
	"""GC-MS"""
	ws = wb["GC-MS"]
	number_format_list = make_column_property_list({'1':'0.000','2':'0.0','3':'0','5':'0','6':'0'}, repeat=PL_len, length=9)	
	width_list = make_column_property_list({"1":14,"2":12,"4":0,"5":9,"6":9,"8":15,"9":10}, repeat=PL_len, length=9)
	alignment_list = make_column_property_list({"5":"center","6":"center","8":"center"})
	format_sheet(ws, number_format_list, width_list, alignment_list)
		
	for offset in range(PL_len):
		merge_string = get_column_letter(1+(9*offset)) + '1:' + get_column_letter(9+(9*offset))  + '1'
		#print(merge_string)
		ws.merge_cells(merge_string)
	
	format_header(ws, make_column_property_list({"1":"center","2":"center","5":"center","6":"center","8":"center","9":"center"},repeat=PL_len, length=9),1,2)

	
	"""Matches"""
	ws = wb["Matches"]
	header_number_format_list = make_column_property_list({'5':'0.000','6':'0.00'}, {"A":"0.000", "B":"0.00"}, {"17":"0.000", "18":'0.000000', "19":'0.00%', "21":"0.00", "22":"0.00", "23":'0.00%'},repeat=PL_len, length=3)
	header_h_alignment_list = make_column_property_list({'5':"right",'6':"right",}, {"A":"right", "B":"right"}, {"17":"right", "18":"right", "19":"right", "21":"right", "22":"right", "23":"right"},repeat=PL_len, length=3)
	#header_v_alignment_list = make_column_property_list({'5':"center",'6':"center",}, {"A": "center", "B":"center", "AF":"center", "AG":"center", "AH":"center", "AJ":"center", "AK":"center", "AL":"center"})
	hits_number_format_list = make_column_property_list({'4':'0','5':'0','6':'0'}, {"C":"0"}, {"5":"0.0","6":"0.0000","7":"0.00%","9":"0.0","10":"0.0000","11":"0.00%","13":"0.0","14":"0.0000","15":"0.00%"},repeat=PL_len, length=3)
	hits_alignment_list = make_column_property_list({'4':'center','5':'center','6':'center'}, {"B":"center","C":"center"}, {"13":"center"},repeat=PL_len, length=3)
	width_list = make_column_property_list({"4":8,"5":8,"6":11}, {"B":12}, {"5":6,"6":9,"7":9,"9":6,"10":9,"11":9,"13":5,"14":7,"15":9,"17":8,"18":9,"19":9,"21":14,"22":12,"23":9,"4":1,"8":1,"12":1,"16":1,"20":1},repeat=PL_len, length=3)

	format_matches(ws, header_number_format_list, header_h_alignment_list, hits_number_format_list,hits_alignment_list,width_list)
	
	format_header(ws, make_column_property_list({"4":"center"}, repeat=PL_len, length=3)) # First Row
	format_header(ws, make_column_property_list({"2":"center"}, {"A":"right"},repeat=((PL_len*3)+25), length=1),2,2) # Second Row
	format_header(ws, make_column_property_list({"2":"center"}, repeat=((PL_len*3)+25), length=1),3,3) # Third Row
	
	for offset in range(PL_len):
		merge_string = get_column_letter(4+(3*offset)) + '1:' + get_column_letter(6+(3*offset))  + '1'
		#print(merge_string)
		ws.merge_cells(merge_string)	
	
	offset = PL_len*3
	for index in [5,9,13,17,21]:
		merge_string = get_column_letter(index+offset) + '2:' + get_column_letter(index+2+offset)  + '2'
		ws.merge_cells(merge_string)
	
	
	"""Statistics"""
	number_format_list = {'C':'0','E':'0.000','F':'0.000000','G':'0.00%','I':'0.00','J':'0.00','K':'0.00%','M':'0.0','N':'0.0000','O':'0.00%','Q':'0.0','R':'0.0000','S':'0.00%','U':'0.0','V':'0.0000','W':'0.00%','Y':'0.0','Z':'0.000','AA':'0.000%'}	
	width_list = {"B":12,"C":2,"D":1,"E":8,"F":11,"G":8,"H":1,"I":13,"J":12,"K":10,"L":1,"M":8,"N":10,"O":10,"P":1,"Q":8,"R":10,"S":10,"T":1,"U":6,"V":8,"W":10,"X":1, "Y":8, "Z":9, "AA":9}
	alignment_list = {"B":"center","C":"center"}

	for sheet in ["Statistics_Full", "Statistics", "Statistics - Lit Only"]:
		ws = wb[sheet]
		format_sheet(ws, number_format_list, width_list, alignment_list)
		for offset in range(6):
			ws.merge_cells(get_column_letter(5+(4*offset)) + '1:' + get_column_letter(7+(4*offset))  + '1')	
	
		format_header(ws, make_column_property_list({"2":"center"}, repeat=27, length=1),1,2) 
	
	
	"""Contents Page"""
	ws = wb.create_sheet("Index",0)
	contents = [["GC-MS","Combined GC-MS data aligned by retention time.",True],
		["Matches", "List of possible matching compounds for each retention time, based on all samples.",True],
		["Statistics_Full", "Statistics for the top hit for each retention time.",True],
		["Statistics", "Statistics for the top hit for each retention time, filtered by the occurance of the compound in the hits, and the mass spectral similarity between samples.",True],
		["Statistics - Lit Only", "As above, but only for compounds reported in literature as being present in propellent or GSR.",True],
	]
				
	print("\nThe worksheets in the output xlsx file are as follows:")
	ws.append(["GunShotMatch Version {} Output".format(__version__)])
	for row in contents: 
		print(row[0] + " "*(24-len(row[0])) + row[1])
		if row[2]:
			ws.append(["",'=HYPERLINK("#\'{0}\'!A1","{0}")'.format(row[0]),row[1]])
		else:
			ws.append(["","{0}".format(row[0]),row[1]])	
	
	br()
	
	ws.column_dimensions["B"].width = 50.0
	for row in ws.iter_rows("B2:B{}".format(len(contents)+1)):
		for cell in row:
			if cell.value:
				if cell.value.startswith("=HYPERLINK"):
					cell.font = Font(color="0000EE",
								underline="single")

#	ws.cell(column=4, row=2).value = "Sample list:"
#	for index, prefix in enumerate(prefixList):
#		ws.cell(row=2, column=(5+index)).value = prefix
	
	# Save the file
	wb.save(inputFile)
	return
	
def format_matches(ws, header_number_format_list=None, header_alignment_list=None, hits_number_format_list=None, hits_alignment_list=None, width_list=None, ):
	
	for row in ws.iter_rows("A1:{}{}".format(get_column_letter(ws.max_column),ws.max_row)):
		for cell in row:
			cell.alignment = Alignment(vertical="center", wrap_text=False)
	
	
	headers = [x+1 for x in range(ws.max_row)][3::6]

	for column_cells in ws.columns:
		length = max(len(as_text(cell.value)) for cell in column_cells)
		#print(column_cells[0].column, length)
		ws.column_dimensions[column_cells[0].column].width = length
		#ws.column_dimensions[column_cells[0].column].bestFit = True


	for row in ws.iter_rows('A4:{}{}'.format(get_column_letter(ws.max_column),ws.max_row)):
		for cell in row:
			if cell.row in headers:
				if header_number_format_list and (cell.column in header_number_format_list):
					cell.number_format = header_number_format_list[cell.column]
				if header_alignment_list and (cell.column in header_alignment_list):
					cell.alignment = Alignment(horizontal=header_alignment_list[cell.column],
										vertical="bottom",
										wrap_text=False)
					
					
			else: # hits
				if hits_number_format_list and (cell.column in hits_number_format_list):
					cell.number_format = hits_number_format_list[cell.column]
				if hits_alignment_list and (cell.column in hits_alignment_list):
					cell.alignment = Alignment(horizontal=hits_alignment_list[cell.column],
										vertical="center",
										wrap_text=False)

	if width_list:
		for column in width_list:
			if width_list[column] == 0:
				ws.column_dimensions[column].hidden = True
			else:
				ws.column_dimensions[column].width = width_list[column]
			

"""Mass Spectra Images"""
def GenerateSpectraFromAlignment(rt_data, ms_data):
	path = os.path.join(SPECTRA_DIRECTORY,lot_name)
	maybe_make(path)

	print("\nGenerating mass spectra images. Please wait.")
	#bar = progressbar.ProgressBar(max_value=sum(rt_data.count()))
	bar_counter = 1

	# Delete Existing Files
	for filename in os.listdir(path):
		os.unlink(os.path.join(path, filename))

	use_mp = True

	if use_mp:
		with Pool(len(rt_data)) as p:		
			args = [(sample, rt_data, ms_data, path) for sample in rt_data.columns.values]
		
			p.map(SpectrumImageWrapper, args)
	else:
		for sample in rt_data.columns.values:
			GenerateSpectrumImage(sample, rt_data, ms_data, path)
	#for sample in rt_data.columns.values:
	#	
	#		bar.update(bar_counter); bar_counter += 1
	
	#bar.finish()

def SpectrumImageWrapper(args):
	return GenerateSpectrumImage(*args)
	
def GenerateSpectrumImage(sample, rt_data, ms_data, path):
	for row_idx in range(len(rt_data)):
		rt = rt_data.iloc[row_idx].loc[sample]
		ms = ms_data.iloc[row_idx].loc[sample]

		PlotSpectrum(numpy.column_stack((ms.mass_list, ms.mass_spec)),
					label = "{}: {}".format(sample, rounders(rt,"0.000")),
					xlim = (45,500),
					mode = path)
	
	return


"""Make Single Tarball Containing All Files"""
def make_archive():
	# Creates single tarball containing main results
	tar = tarfile.open(os.path.join(RESULTS_DIRECTORY, (
				lot_name + "_" + datetime.datetime.now().strftime("%Y%m%d%H%M%S") + '.tar.gz')),mode='w:gz')
	
	tar.add(os.path.join(RESULTS_DIRECTORY, lot_name + "_FINAL.xlsx"), arcname=(lot_name + ".xlsx"))
	tar.add(os.path.join(CHARTS_DIRECTORY, lot_name), arcname='Charts')
	tar.add(os.path.join(SPECTRA_DIRECTORY, lot_name), arcname='Spectra')
	
	tar.add(os.path.join(CSV_DIRECTORY, lot_name + "_MERGED.csv"), arcname='Merged.CSV')
	tar.add(os.path.join(CSV_DIRECTORY, lot_name + "_MATCHES.csv"), arcname='Matches.CSV')
	tar.add(os.path.join(CSV_DIRECTORY, "{}_STATISTICS_FULL.csv".format(lot_name)),
			arcname='Statistics/Statistics_Full.CSV')
	tar.add(os.path.join(CSV_DIRECTORY, "{}_STATISTICS.csv".format(lot_name)), arcname='Statistics/Statistics.CSV')
	tar.add(os.path.join(CSV_DIRECTORY, "{}_STATISTICS_LIT.csv".format(lot_name)),
			arcname='Statistics/Statistics_Lit.CSV')
	tar.add(os.path.join(CSV_DIRECTORY, "{}_peak_data.json".format(lot_name)), arcname="Peak_Data.json")
	tar.add(os.path.join(RESULTS_DIRECTORY, f"{lot_name}.info"), arcname="sample_list.info")
	
	
	for prefix in prefixList:
		tar.add(os.path.join(EXPERIMENTS_DIRECTORY, "{}_tic.dat".format(prefix)),
				arcname="Experiments/{}_tic.dat".format(prefix))
		tar.add(os.path.join(EXPERIMENTS_DIRECTORY, "{}_peaks.dat".format(prefix)),
				arcname="Experiments/{}_peaks.dat".format(prefix))
		tar.add(os.path.join(EXPERIMENTS_DIRECTORY, "{}.expr".format(prefix)),
				arcname="Experiments/{}_expr.dat".format(prefix))
		
		tar.add(os.path.join(CSV_DIRECTORY, "{}_COMBINED.csv".format(prefix)),
				arcname="Combined/{}.CSV".format(prefix))

	print(f"\nSaved as: {os.path.join(RESULTS_DIRECTORY, lot_name + '_FINAL.xlsx')}")


def arguments():
	# Command line switches
	import argparse
	parser = argparse.ArgumentParser()
	
	parser.add_argument("--samples", help="List of samples to run.", nargs='+')
	parser.add_argument("--name", help="Human-Readable Name for the Project.")
	parser.add_argument("--info", help="Show program info.", action='store_true')
	parser.add_argument("--default", help="Reload Default configuration.", action='store_true')
	
	args = parser.parse_args()

	if len(sys.argv) == 1:
		#logger.warning("No options supplied.")
		print("Error: No options supplied.")
		print('')
		parser.print_help(sys.stderr)
		sys.exit(1)

	if args.default:
		print("\nReloading Default Configuration")
		shutil.copyfile("./lib/default.ini", "./config.ini")
		sys.exit(0)
	
	return args


if __name__ == '__main__':
	from utils import timing
	clear()  # clear the display
	startup_string = f"\n\n{program_name} Version {__version__} running on {platform.system()}.\nCopyright {copyright} Dominic Davis-Foster."
	print(startup_string)
	
	# Load Configuration
	config = GSMConfig("config.ini")
	nist_path = config.nist_path
	msp_directory = config.MSP_DIRECTORY
	MSP_DIRECTORY = config.MSP_DIRECTORY
	EXPERIMENTS_DIRECTORY = config.EXPERIMENTS_DIRECTORY
	RESULTS_DIRECTORY = config.RESULTS_DIRECTORY
	CSV_DIRECTORY = config.CSV_DIRECTORY
	SPECTRA_DIRECTORY = config.SPECTRA_DIRECTORY
	CHARTS_DIRECTORY = config.CHARTS_DIRECTORY
	RAW_DIRECTORY = config.RAW_DIRECTORY
	LOG_DIRECTORY = os.path.join(RESULTS_DIRECTORY, "logs")
	maybe_make(LOG_DIRECTORY)
	
	bb_points = config.bb_points
	bb_scans = config.bb_scans
	noise_thresh = config.noise_thresh
	target_range = config.target_range
	base_peak_filter = config.base_peak_filter
	tophat_struct = config.tophat_struct
	prefixList = config.prefixList
	mass_range = config.mass_range
	
	# within replicates alignment parameters
	Dw = config.rt_modulation  # [s]
	Gw = config.gap_penalty
	min_peaks = config.min_peaks
	
	do_quantitative = config.do_quantitative
	do_qualitative = config.do_qualitative
	do_merge = config.do_merge
	do_counter = config.do_counter
	do_spectra = config.do_spectra
	do_charts = config.do_charts
	
	base_peak_filter = [int(x) for x in base_peak_filter]
	target_range = tuple(target_range)


	args = arguments()
	
	if args.samples:
		prefixList = args.samples
		# overrides whatever was set from the config file

	"""Define Exit Functions"""
	if "-h" not in str(sys.argv):
		atexit.register(pynist.reload_ini, nist_path)

	

	if args.name:
		lot_name = args.name
	else:
		lot_name = re.sub(r'\d+', '', str(prefixList[0].rstrip("\n\r "))).replace("__","_")
	PL_len = len(prefixList)

	# define the input experiments list
	expr_list = []

	

	if do_quantitative:
		print("Quantitative Processing in Progress...")
		with Pool(PL_len) as p:
			p.map(quantitative_processing, [os.path.join(RAW_DIRECTORY,"{}.JDX".format(prefix)) for prefix in prefixList])
		for prefix in prefixList:
		#	quantitative_processing("/media/VIDEO/ownCloud/GSR/GunShotMatch/GSMatch/Results/RAW/{}.JDX".format(prefix))
			"""Read and Print Log"""
			with open(os.path.join(LOG_DIRECTORY, prefix + ".log"), "r") as f:
				print(f.read())
			
	for prefix in prefixList:
		file_name = os.path.join(EXPERIMENTS_DIRECTORY, prefix + ".expr")
		expr_list.append(load_expr(file_name))
	
	print("\nAligning\n")
	F1 = exprl2alignment(expr_list)
	T1 = PairwiseAlignment(F1, Dw, Gw)
	A1 = align_with_tree(T1, min_peaks=min_peaks)

	A1.write_csv(os.path.join(EXPERIMENTS_DIRECTORY,'{}_rt.csv'.format(lot_name)),
				 os.path.join(EXPERIMENTS_DIRECTORY,'{}_area.csv'.format(lot_name)))
	rt_alignment = get_peak_alignment(A1)
	ms_alignment = get_ms_alignment(A1)

	#print(rt_alignment)
	
	if do_qualitative:
		print("Qualitative Processing in Progress...")
		for prefix in prefixList:
			#print(list(rt_alignment[prefix]))
			qualitative_processing(prefix, list(rt_alignment[prefix]))
	
	if do_merge:
		Merge()
	
	if do_counter:
		chart_data = Match_Counter(ms_comparisons(ms_alignment))
		chart_data = chart_data.set_index("Compound", drop=True)
		
		# remove duplicate compounds:
		#chart_data_count = Counter(chart_data["Compound"])
		chart_data_count = Counter(chart_data.index)
		replacement_data = {"Compound":[], f"{lot_name} Peak Area":[], f"{lot_name} Standard Deviation":[]}
		
		for prefix in prefixList:
			replacement_data[prefix] = []
		
		for compound in chart_data_count:
			if chart_data_count[compound] > 1:
				replacement_data["Compound"].append(compound)
				replacement_data[f"{lot_name} Peak Area"].append( sum(chart_data.loc[compound,f"{lot_name} Peak Area"]) )
				
				peak_data = []
				for prefix in prefixList:
					replacement_data[prefix].append(sum(chart_data.loc[compound, prefix]))
					peak_data.append(sum(chart_data.loc[compound, prefix]))
				
				replacement_data[f"{lot_name} Standard Deviation"].append(numpy.std(peak_data))
				
				chart_data = chart_data.drop(compound, axis=0)
		
		replacement_data = pandas.DataFrame(replacement_data)
		replacement_data = replacement_data.set_index("Compound", drop=False)
		chart_data = chart_data.append(replacement_data, sort=False)
		chart_data.sort_index(inplace=True)
		chart_data = chart_data.drop("Compound", axis=1)
		chart_data['Compound Names'] = chart_data.index
		
		chart_data.to_csv(os.path.join(CSV_DIRECTORY,"{}_CHART_DATA.csv".format(lot_name)), sep=";")
	else:
		chart_data = pandas.read_csv(os.path.join(CSV_DIRECTORY,"{}_CHART_DATA.csv".format(lot_name)), sep=";", index_col=0)
		#chart_data = chart_data.set_index("Compound", drop=True)

	if do_spectra:
		GenerateSpectraFromAlignment(rt_alignment, ms_alignment)
	
	if do_charts:
		print("\nGenerating Charts")
		maybe_make(os.path.join(CHARTS_DIRECTORY, lot_name))

		chart_data.to_csv(os.path.join(CSV_DIRECTORY,"{}_CHART_DATA.csv".format(lot_name)), sep=";")
		
		from lib.charts import peak_area, radar_chart
		
		radar_chart(chart_data, [lot_name], use_log = 10, legend=False,
					mode=os.path.join(CHARTS_DIRECTORY, lot_name, "radar_log10_peak_area"))
		radar_chart(chart_data, [lot_name], use_log = False, legend=False,
					mode=os.path.join(CHARTS_DIRECTORY, lot_name, "radar_peak_area"))
		mean_peak_area_wrapper(chart_data, [lot_name],
					mode=os.path.join(CHARTS_DIRECTORY, lot_name, "mean_peak_area"))
		peak_area_wrapper(chart_data, lot_name, prefixList,
					mode=os.path.join(CHARTS_DIRECTORY, lot_name, "peak_area_percentage"))
		peak_area_wrapper(chart_data, lot_name, prefixList, percentage = False,
					mode=os.path.join(CHARTS_DIRECTORY, lot_name, "peak_area"))
		peak_area_wrapper(chart_data, lot_name, prefixList, use_log=10, mode=os.path.join(CHARTS_DIRECTORY, lot_name, "log10_peak_area_percentage"))
		
		samples_to_compare = [
			(lot_name, prefixList),
			#	("ELEY_CASE_SUBTRACT","Eley Case"),
			#	("WINCHESTER_SUBTRACT","Winchester Pistol"),
			#	("WINCHESTER_CASE_SUBTRACT","Winchester Case"),
			#	("GECO_SUBTRACT","Geco"),
			#	("GECO_CASE_SUBTRACT","Geco Case"),
			#	("ELEY_SHOTGUN_SUBTRACT", "Eley Hawk"),
		]  # must be in order you want them on the graph
		
		box_whisker_wrapper(chart_data, samples_to_compare, mode=os.path.join(CHARTS_DIRECTORY, lot_name, "box_whisker"))
		
	
	with open(os.path.join(RESULTS_DIRECTORY, f"{lot_name}.info"),"w") as info_file:
		for prefix in prefixList:
			info_file.write(f"{prefix}\n")
	
	make_archive()
	
	pynist.reload_ini(nist_path)
	
	print("\nComplete.")





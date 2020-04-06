#!/usr/bin/env python3
# -*- coding: utf-8 -*-

#  !/usr/bin/env python
#   -*- coding: utf-8 -*-
#
#  NewProject.py
#
#  This file is part of GunShotMatch
#
#  Copyright © 2019 Dominic Davis-Foster <dominic@davis-foster.co.uk>
#
#  GunShotMatch is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 3 of the License, or
#  (at your option) any later version.
#
#  GunShotMatch is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#

#
"""
GunShotMatch

Program for the analysis of OGSR samples to identify matching compounds 
between samples. 

"""

# stdlib
import os
import re
import sys
import csv
import json
import time
import locale
import atexit
import tarfile
import operator
import datetime
import platform
import traceback
import itertools

from collections import Counter
from multiprocessing import Pool

# 3rd party
import numpy
import pandas

from mathematical.utils import rounders
from domdf_python_tools.terminal import br
from domdf_python_tools.utils import as_text
from domdf_python_tools.paths import maybe_make
from domdf_spreadsheet_tools import format_header, format_sheet, make_column_property_dict, append_to_xlsx

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from pyms.GCMS.Class import IonChromatogram
from pyms.Peak.List.IO import load_peaks
from pyms.DPA.PairwiseAlignment import PairwiseAlignment, align_with_tree
from pyms.DPA.Alignment import exprl2alignment
from pyms.Experiment import load_expr
from pyms.GCMS.IO.JCAMP import JCAMP_reader
from pyms.IntensityMatrix import build_intensity_matrix_i
from pyms.Noise.Analysis import window_analyzer
from pyms.Noise.SavitzkyGolay import savitzky_golay
from pyms.TopHat import tophat
from pyms.BillerBiemann import BillerBiemann, num_ions_threshold
from pyms.Peak.Function import peak_sum_area
from pyms.Peak.List.IO import store_peaks
from pyms.Experiment import store_expr
from pyms.Experiment import Experiment

from chemistry_tools import spectrum_similarity

# this package
from GSMatch.utils.pynist import *
from GSMatch.utils import DirectoryHash, pynist
from GSMatch.GSMatch_Core.PeakAlignment import get_ms_alignment, get_peak_alignment

sys.path.append("..")

__author__ = "Dominic Davis-Foster"
__copyright__ = "Copyright 2017-2019 Dominic Davis-Foster"

__license__ = "GPLv3"
__version__ = "1.0.0 Rework"
__email__ = "dominic@davis-foster.co.uk"

program_name = "GunShotMatch"
copyright = __copyright__

# Setup for reading strings with thousand separators as floats
# From https://stackoverflow.com/a/31074271
locale.setlocale(locale.LC_ALL, "")

verbose = False

"""From https://stackoverflow.com/questions/14463277/how-to-disable-python-warnings"""


# with warnings.catch_warnings():
# 	warnings.simplefilter("ignore")


def single_ms_comparison(arguments):
	"""
	Performs a single Mass Spectrum similarity calculation for the same peak in
		two different samples.
	
	:param arguments: tuple containing mass spectrum data and the samples to
		which these spectra belong
	:type arguments: tuple
	:return:
	:rtype:
	"""
	
	ms_data, perms = arguments
	
	similarity_list = []
	
	for perm in perms:
		top_spec = numpy.column_stack((ms_data.loc[perm[0]].mass_list, ms_data.loc[perm[0]].mass_spec))
		bottom_spec = numpy.column_stack((ms_data.loc[perm[1]].mass_list, ms_data.loc[perm[1]].mass_spec))
		similarity_list.append(spectrum_similarity.SpectrumSimilarity(
			top_spec, bottom_spec, t=0.25, b=1, xlim=(45, 500),
			x_threshold=0, print_graphic=False
		)[0] * 1000)
	
	# ms_comparison.append(rounders((numpy.mean(similarity_list)*100),"0"))
	return similarity_list


def format_matches(
		ws, header_number_format_dict=None, header_alignment_dict=None,
		hits_number_format_dict=None, hits_alignment_dict=None, width_dict=None
):
	"""
	Format xlsx spreadsheet containing matches data
	
	:param ws: The worksheet for format
	:type ws: :class:`openpyxl.worksheet.worksheet.Worksheet`
	:param header_number_format_dict: Header number format for each column
	:type header_number_format_dict: dict
	:param header_alignment_dict: Header alignment for each column
	:type header_alignment_dict: dict
	:param hits_number_format_dict: Hits number format for each column
	:type hits_number_format_dict: dict
	:param hits_alignment_dict: Hits alignment for each column
	:type hits_alignment_dict: dict
	:param width_dict: Width for each column
	:type width_dict: dict
	"""
	
	for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
		for cell in row:
			cell.alignment = Alignment(vertical="center", wrap_text=False)
	
	headers = [x + 1 for x in range(ws.max_row)][3::6]
	
	for column_cells in ws.columns:
		length = max(len(as_text(cell.value)) for cell in column_cells)
		ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length
	# ws.column_dimensions[column_cells[0].column].bestFit = True
	
	for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
		for cell in row:
			# Column Headers
			if cell.row in headers:
				if header_number_format_dict and (cell.column in header_number_format_dict):
					cell.number_format = header_number_format_dict[cell.column]
				if header_alignment_dict and (cell.column in header_alignment_dict):
					cell.alignment = Alignment(
						horizontal=header_alignment_dict[cell.column],
						vertical="bottom",
						wrap_text=False
					)
			# Hit Data
			else:
				if hits_number_format_dict and (cell.column in hits_number_format_dict):
					cell.number_format = hits_number_format_dict[cell.column]
				if hits_alignment_dict and (cell.column in hits_alignment_dict):
					cell.alignment = Alignment(
						horizontal=hits_alignment_dict[cell.column],
						vertical="center",
						wrap_text=False
					)
	if width_dict:
		for column in width_dict:
			if width_dict[column] == 0:
				ws.column_dimensions[column].hidden = True
			else:
				ws.column_dimensions[column].width = width_dict[column]


def write_peak(file_object, peak, ms):
	"""
	Write the given peak and mass spectrum to the file
	
	:param file_object: file object to write data to
	:type file_object: :class:`io.TextIOWrapper
	:param peak:
	:type peak:
	:param ms:
	:type ms:
	"""
	
	file_object.write("{};{};{};;".format(
		peak["hits"][0]["Name"], peak["hits"][0]["CAS"],
		peak["hits"][0]["Count"]
	))
	file_object.write("{};{};{};;".format(
		peak["average_rt"], numpy.nanstd(peak["rt_data"]),
		numpy.nanstd(peak["rt_data"]) / peak["average_rt"]
	))
	file_object.write("{};{};{};;".format(
		peak["average_peak_area"], numpy.nanstd(peak["area_data"]),
		numpy.nanstd(peak["area_data"]) / peak["average_peak_area"]
	))
	file_object.write("{};{};{};;".format(
		peak["hits"][0]["average_MF"], numpy.nanstd(peak["hits"][0]["mf_data"]),
		numpy.nanstd(peak["hits"][0]["mf_data"]) / peak["hits"][0]["average_MF"]
	))
	file_object.write("{};{};{};;".format(
		peak["hits"][0]["average_RMF"], numpy.nanstd(peak["hits"][0]["rmf_data"]),
		numpy.nanstd(peak["hits"][0]["rmf_data"]) / peak["hits"][0]["average_RMF"]
	))
	file_object.write("{};{};{};;".format(
		peak["hits"][0]["average_hit"], numpy.nanstd(peak["hits"][0]["hit_num_data"]),
		numpy.nanstd(peak["hits"][0]["hit_num_data"]) / peak["hits"][0]["average_hit"]
	))
	file_object.write("{};{};{};;".format(
		numpy.mean(ms), numpy.std(ms), numpy.std(ms) / numpy.mean(ms)
	))
	file_object.write("\n")


def parse_arguments():
	"""
	Parse command line switches
	
	:return:
	"""
	
	import shutil
	import argparse
	parser = argparse.ArgumentParser()
	
	parser.add_argument("--samples", help="List of samples to run.", nargs='+')
	parser.add_argument("--name", help="Human-Readable Name for the Project.")
	parser.add_argument("--info", help="Show program info.", action='store_true')
	parser.add_argument("--default", help="Reload Default configuration.", action='store_true')
	
	args = parser.parse_args()
	
	if len(sys.argv) == 1:
		# logger.warning("No options supplied.")
		print("Error: No options supplied.")
		print('')
		parser.print_help(sys.stderr)
		sys.exit(1)
	
	if args.default:
		print("\nReloading Default Configuration")
		shutil.copyfile("./lib/default.ini", "./config.ini")
		sys.exit(0)
	
	return args


class NewProject:
	"""
	
	
	:param config: Configuration to use
	:type config: :class:`GSMatch.GSMatch_Core.Config.GSMConfig`, optional
	"""
	
	def __init__(self, config=None):
		"""
		Initialise class NewProject
		
		:param config: Configuration to use
		:type config: :class:`GSMatch.GSMatch_Core.Config.GSMConfig`, optional
		"""
		
		if config is None:
			from GSMatch.GSMatch_Core.Config import GSMConfig
			self.config = GSMConfig("config.ini")
		else:
			self.config = config
		
		# Define Exit Functions
		atexit.register(pynist.reload_ini, self.config.nist_path)
		
		# Determine length of prefixList and name of lot of samples
		self.PL_len = len(self.config.prefixList)
		self.lot_name = re.sub(r'\d+', '', str(self.config.prefixList[0].rstrip("\n\r "))).replace("__", "_")
		
		# Define the input experiments list
		self.expr_list = []
	
	def run(self):
		# Indicate which steps to perform
		print(f"do_quantitative: {self.config.do_quantitative}")
		print(f"do_qualitative: {self.config.do_qualitative}")
		print(f"do_merge: {self.config.do_merge}")
		print(f"do_counter: {self.config.do_counter}")
		print(f"do_spectra: {self.config.do_spectra}")
		print(f"do_charts: {self.config.do_charts}")
		
		if self.config.do_quantitative:
			print("Quantitative Processing in Progress...")
			
			# Number of workers for performing Quantitative Processing in parallel
			# If 0 processing will be performed sequentially
			n_quant_workers = self.PL_len
			
			if n_quant_workers:
				# Perform Quantitative Processing in parallel
				with Pool(n_quant_workers) as p:
					p.map(
						self.quantitative_processing, [os.path.join(
							self.config.raw_dir,
							"{}.JDX".format(prefix)
						) for prefix in self.config.prefixList])
					
			for prefix in self.config.prefixList:
				if not n_quant_workers:
					# Perform Quantitative Processing sequentially
					self.quantitative_processing(os.path.join(self.config.raw_dir, "{}.JDX".format(prefix)), False)
				
				# Read the log file and print the contents
				with open(os.path.join(self.config.log_dir, prefix + ".log"), "r") as f:
					print(f.read())
		
		# Loads the experiment file created during Quantitative Processing
		for prefix in self.config.prefixList:
			file_name = os.path.join(self.config.expr_dir, prefix + ".expr")
			self.expr_list.append(load_expr(file_name))
		
		if any([self.config.do_qualitative, self.config.do_merge, self.config.do_spectra, self.config.do_counter]):
			# Perform dynamic peak alignment
			print("\nAligning\n")
			print(self.expr_list)
			
			F1 = exprl2alignment(self.expr_list)
			print(F1)
			
			T1 = PairwiseAlignment(F1, self.config.rt_modulation, self.config.gap_penalty)
			A1 = align_with_tree(T1, min_peaks=self.config.min_peaks)
			
			A1.write_csv(
				os.path.join(self.config.expr_dir, '{}_rt.csv'.format(self.lot_name)),
				os.path.join(self.config.expr_dir, '{}_area.csv'.format(self.lot_name)))
			rt_alignment = get_peak_alignment(A1)
			ms_alignment = get_ms_alignment(A1)
			
			# print(rt_alignment)
			# print(ms_alignment)
			
		if self.config.do_qualitative:
			print("Qualitative Processing in Progress...")
			for prefix in self.config.prefixList:
				# print(list(rt_alignment[prefix]))
				self.qualitative_processing(prefix, list(rt_alignment[prefix]))
		
		if self.config.do_merge:
			self.merge()
		
		if self.config.do_counter:
			chart_data = self.match_counter(self.ms_comparisons(ms_alignment))
			chart_data = chart_data.set_index("Compound", drop=True)
			
			# remove duplicate compounds:
			# chart_data_count = Counter(chart_data["Compound"])
			chart_data_count = Counter(chart_data.index)
			replacement_data = {"Compound": [], f"{self.lot_name} Peak Area": [],
								f"{self.lot_name} Standard Deviation": []}
			
			for prefix in self.config.prefixList:
				replacement_data[prefix] = []
			
			for compound in chart_data_count:
				if chart_data_count[compound] > 1:
					replacement_data["Compound"].append(compound)
					replacement_data[f"{self.lot_name} Peak Area"].append(
						sum(chart_data.loc[compound, f"{self.lot_name} Peak Area"]))
					
					peak_data = []
					for prefix in self.config.prefixList:
						replacement_data[prefix].append(sum(chart_data.loc[compound, prefix]))
						peak_data.append(sum(chart_data.loc[compound, prefix]))
					
					replacement_data[f"{self.lot_name} Standard Deviation"].append(numpy.std(peak_data))
					
					chart_data = chart_data.drop(compound, axis=0)
			
			replacement_data = pandas.DataFrame(replacement_data)
			replacement_data = replacement_data.set_index("Compound", drop=False)
			chart_data = chart_data.append(replacement_data, sort=False)
			chart_data.sort_index(inplace=True)
			chart_data = chart_data.drop("Compound", axis=1)
			chart_data['Compound Names'] = chart_data.index
			
			chart_data.to_csv(os.path.join(self.config.csv_dir, "{}_CHART_DATA.csv".format(self.lot_name)), sep=";")
		else:
			chart_data = pandas.read_csv(
				os.path.join(self.config.csv_dir, "{}_CHART_DATA.csv".format(self.lot_name)),
				sep=";", index_col=0)
		
		# chart_data = chart_data.set_index("Compound", drop=True)
		
		if self.config.do_spectra:
			self.generate_spectra_from_alignment(rt_alignment, ms_alignment)
			
			# Write Mass Spectra to OpenChrom-like CSV files
			
			def generate_spectra_csv(rt_data, ms_data, name):
				# Write Mass Spectra to OpenChrom-like CSV files
				
				ms = ms_data[0]  # first mass spectrum
				
				spectrum_csv_file = os.path.join(self.config.spectra_dir, self.lot_name, f"{name}_data.csv")
				spectrum_csv = open(spectrum_csv_file, 'w')
				spectrum_csv.write('RT(milliseconds);RT(minutes) - NOT USED BY IMPORT;RI;')
				spectrum_csv.write(';'.join(str(mz) for mz in ms.mass_list))
				spectrum_csv.write("\n")
				
				for rt, ms in zip(rt_data, ms_data):
					spectrum_csv.write(f"{int(rt * 60000)};{rounders(rt, '0.0000000000')};0;")
					spectrum_csv.write(';'.join(str(intensity) for intensity in ms.mass_spec))
					spectrum_csv.write('\n')
				spectrum_csv.close()
			
			for prefix in self.config.prefixList:
				print(prefix)
				# print(rt_alignment[prefix])
				# print(ms_alignment[prefix])
				generate_spectra_csv(rt_alignment[prefix], ms_alignment[prefix], prefix)
		
		if self.config.do_charts:
			print("\nGenerating Charts")
			
			chart_data.to_csv(os.path.join(self.config.csv_dir, "{}_CHART_DATA.csv".format(self.lot_name)), sep=";")
			
			maybe_make(os.path.join(self.config.charts_dir, self.lot_name))
			
			if chart_data.empty:
				print("ALERT: No peaks were found for compounds that have")
				print("       previously been reported in literature.")
				print("       Check the results for more information\n")
			
			else:
				from GSMatch.GSMatch_Core.charts import box_whisker_wrapper, radar_chart_wrapper, \
					mean_peak_area_wrapper, \
					peak_area_wrapper
				
				# from GSMatch.GSMatch_Core.charts import peak_area_wrapper, radar_chart_wrapper
				
				radar_chart_wrapper(
					chart_data, [self.lot_name], use_log=10, legend=False,
					mode=os.path.join(self.config.charts_dir, self.lot_name, "radar_log10_peak_area"))
				radar_chart_wrapper(
					chart_data, [self.lot_name], use_log=False, legend=False,
					mode=os.path.join(self.config.charts_dir, self.lot_name, "radar_peak_area"))
				mean_peak_area_wrapper(
					chart_data, [self.lot_name],
					mode=os.path.join(self.config.charts_dir, self.lot_name, "mean_peak_area"))
				peak_area_wrapper(
					chart_data, self.lot_name, self.config.prefixList,
					mode=os.path.join(self.config.charts_dir, self.lot_name, "peak_area_percentage"))
				peak_area_wrapper(
					chart_data, self.lot_name, self.config.prefixList, percentage=False,
					mode=os.path.join(self.config.charts_dir, self.lot_name, "peak_area"))
				peak_area_wrapper(
					chart_data, self.lot_name, self.config.prefixList, use_log=10,
					mode=os.path.join(self.config.charts_dir, self.lot_name, "log10_peak_area_percentage"))
				
				samples_to_compare = [(self.lot_name, self.config.prefixList)]
				
				box_whisker_wrapper(
					chart_data, samples_to_compare,
					mode=os.path.join(self.config.charts_dir, self.lot_name, "box_whisker"))
		
		with open(os.path.join(self.config.results_dir, f"{self.lot_name}.info"), "w") as info_file:
			for prefix in self.config.prefixList:
				info_file.write(f"{prefix}\n")
		
		# TODO: self.make_archive()
		
		pynist.reload_ini(self.config.nist_path)
		
		print("\nComplete.")
	
	def quantitative_processing(self, jcamp_file, log_stdout=True):
		"""
		Import JCAMP-DX Files

		:param jcamp_file:
		:type jcamp_file:
		:param log_stdout:
		:type log_stdout:
		
		:return:
		:rtype:
		"""
		
		# Determine the name of the sample from the filename
		sample_name = os.path.splitext(os.path.basename(jcamp_file))[0]
		
		# Log Stdout to File
		if log_stdout:
			sys.stdout = open(os.path.join(self.config.log_dir, sample_name + ".log"), "w")
		
		# Load data using JCAMP_reader
		data = JCAMP_reader(jcamp_file)
		
		# list of all retention times, in seconds
		times = data.get_time_list()
		# get Total Ion Chromatogram
		tic = data.get_tic()
		# RT Range, time step, no. scans, min, max, mean and median m/z
		data.info()
		
		# Build "intensity matrix" by binning data with integer bins and a
		# 	window of -0.3 to +0.7, the same as NIST uses
		im = build_intensity_matrix_i(data)
		
		# Show the m/z of the maximum and minimum bins
		print(" Minimum m/z bin: {}".format(im.get_min_mass()))
		print(" Maximum m/z bin: {}".format(im.get_max_mass()))
		
		# Crop masses
		min_mass, max_mass, *_ = self.config.mass_range
		
		if min_mass < im.get_min_mass():
			min_mass = im.get_min_mass()
		if max_mass > im.get_max_mass():
			max_mass = im.get_max_mass()
		im.crop_mass(min_mass, max_mass)
		
		# Perform Data filtering
		n_scan, n_mz = im.get_size()
		
		# Iterate over each IC in the intensity matrix
		for ii in range(n_mz):
			# print("\rWorking on IC#", ii+1, '  ',end='')
			ic = im.get_ic_at_index(ii)
			
			# Perform Savitzky-Golay smoothing.
			# Note that Turbomass does not use smoothing for qualitative method.
			ic_smooth = savitzky_golay(ic)
			
			# Perform Tophat baseline correction
			# Top-hat baseline Correction seems to bring down noise,
			#  		retaining shapes, but keeps points on actual peaks
			ic_bc = tophat(ic_smooth, struct=self.config.tophat_struct)
			
			# Set the IC in the intensity matrix to the filtered one
			im.set_ic_at_index(ii, ic_bc)
		
		# Peak Detection based on Biller and Biemann (1974), with a window
		# 	of <points>, and combining <scans> if they apex next to each other
		peak_list = BillerBiemann(im, points=self.config.bb_points, scans=self.config.bb_scans)
		
		print(" Number of peaks identified before filtering: {}".format(len(peak_list)))
		
		# Filtering peak lists with automatic noise filtering
		noise_level = window_analyzer(tic)
		# should we also do rel_threshold() here?
		# https://pymassspec.readthedocs.io/en/master/pyms/BillerBiemann.html#pyms.BillerBiemann.rel_threshold
		peak_list = num_ions_threshold(peak_list, self.config.noise_thresh, noise_level)
		
		filtered_peak_list = []
		
		for peak in peak_list:
			# Get mass and intensity lists for the mass spectrum at the apex of the peak
			apex_mass_list = peak.mass_spectrum.mass_list
			apex_mass_spec = peak.mass_spectrum.mass_spec
			
			# Determine the intensity of the base peak in the mass spectrum
			base_peak_intensity = max(apex_mass_spec)
			
			# Determine the index of the base peak in the mass spectrum
			base_peak_index = [
				index for index, intensity in enumerate(apex_mass_spec)
				if intensity == base_peak_intensity][0]
			
			# Finally, determine the mass of the base peak
			base_peak_mass = apex_mass_list[base_peak_index]
			
			# skip the peak if the base peak is at e.g. m/z 73, i.e. septum bleed
			if base_peak_mass in self.config.base_peak_filter:
				continue
			
			area = peak_sum_area(im, peak)
			peak.set_area(area)
			filtered_peak_list.append(peak)
			
		print(" Number of peaks identified: {}".format(len(filtered_peak_list)))
		
		# Save the TIC and Peak List
		tic.write(os.path.join(self.config.expr_dir, "{}_tic.dat".format(sample_name)), formatting=False)
		store_peaks(filtered_peak_list, os.path.join(self.config.expr_dir, "{}_peaks.dat".format(sample_name)))
		
		# Create an experiment
		expr = Experiment(sample_name, filtered_peak_list)
		expr.sele_rt_range(["{}m".format(self.config.target_range[0]), "{}m".format(self.config.target_range[1])])
		store_expr(os.path.join(self.config.expr_dir, "{}.expr".format(sample_name)), expr)
	
	def qualitative_processing(self, sample_name, rt_list, n_peaks=80):
		"""
		
		
		:param sample_name:
		:type sample_name:
		:param rt_list:
		:type rt_list:
		:param n_peaks: Number of peaks to include in the report
		:type n_peaks: int
		
		:return:
		:rtype:
		"""
		
		# Initialise variables
		time_list = []
		intensity_list = []
		
		# Load saved TIC
		with open(os.path.join(self.config.expr_dir, "{}_tic.dat".format(sample_name))) as tic_file:
			ticreader = csv.reader(tic_file, delimiter=" ")
			for row in ticreader:
				row = list(filter(None, row))
				intensity_list.append(float(row[1]))
				time_list.append(float(row[0]))
		
		# tic = IonChromatogram(numpy.array(intensity_list), time_list)
		
		# Load the peak list from file
		peak_list = load_peaks(os.path.join(self.config.expr_dir, "{}_peaks.dat".format(sample_name)))
		
		# Obtain area for each peak
		peak_area_list = []
		for peak in peak_list:
			area = peak.get_area()
			peak_area_list.append(area)
		
		# Write output to CSV file
		combined_csv_file = os.path.join(self.config.csv_dir, "{}_COMBINED.csv".format(sample_name))
		with open(combined_csv_file, "w") as combine_csv:
			
			# Sample name and header row
			combine_csv.write(f"""{sample_name}
Retention Time;Peak Area;;Lib;Match;R Match;Name;CAS Number;Notes
""")
		
			report_buffer = []
			# Filter to those peaks present in all samples, by UID
			for peak in peak_list:
				# if str(rounders(peak.get_rt()/60,"0.000")) in rt_list:
				# print(peak.get_rt()/60.0)
				# TODO: there is a simpler way to do this as part of the DPA functions
				# 	DDF 20/11/19
				if peak.get_rt() / 60.0 in rt_list:
					report_buffer.append([
						'',
						# rounders(peak.get_rt()/60,"0.000"),
						(peak.get_rt() / 60),
						'',
						peak.get_mass_spectrum(),
						# '{:,}'.format(rounders(peak.get_area()/60,"0.0"))
						'{:,}'.format(peak.get_area() / 60)
					])
			
			# TODO: I thought this was supposed to filter to show the 80 largest peaks,
			# 	but I'm not sure it actually does that
			# 	DDF 20/11/19
			
			# Reverse list order
			report_buffer = report_buffer[::-1]
			# Get last 80 peaks
			report_buffer = report_buffer[:n_peaks]
			# Sort by retention time
			report_buffer.sort(key=operator.itemgetter(1))
			
			# Iterate over peaks
			for row_idx, row in enumerate(report_buffer):
				# TODO: some tidying up here; is writing to disk the most efficient?
				# 	DDF 20/11/19
				
				# if row_idx == 19:
				
				# Get mass spectrum
				ms = row[3]
				
				# Number of hits to get from NIST MS Search
				n_hits = 10
				
				# Create MSP file for the peak
				self.create_msp("{}_{}".format(sample_name, row[1]), ms.mass_list, ms.mass_spec)
				
				matches_dict = self.nist_ms_comparison(
					"{}_{}".format(sample_name, row[1]),
					ms.mass_list, ms.mass_spec, n_hits
				)
				
				combine_csv.write("{};{};Page {} of 80;;;;;;{}\n".format(row[1], row[4], row_idx + 1, row[2]))
				
				for hit in range(1, n_hits + 1):
					combine_csv.write(';;{};{};{};{};{};{};\n'.format(
						hit,
						matches_dict["Hit{}".format(hit)]["Lib"],
						matches_dict["Hit{}".format(hit)]["MF"],
						matches_dict["Hit{}".format(hit)]["RMF"],
						matches_dict["Hit{}".format(hit)]["Name"].replace(";", ":"),
						matches_dict["Hit{}".format(hit)]["CAS"],
					))
				
				time.sleep(2)
				
		return 0
	
	def create_msp(self, sample_name, mass_list, mass_spec):
		"""
		Generate .MSP files for NIST MS Search
		
		:param sample_name: The name of the sample
		:type sample_name: str
		:param mass_list:
		:type mass_list:
		:param mass_spec:
		:type mass_spec:
		"""
		
		msp_file = open(os.path.join(self.config.msp_dir, sample_name + ".MSP"), "w")
		msp_file.write("Name: {}\n".format(sample_name))
		msp_file.write("Num Peaks: {}\n".format(len(mass_list)))
		for mass, intensity in zip(mass_list, mass_spec):
			msp_file.write("{} {},\n".format(rounders(mass, "0.0"), intensity))
		msp_file.close()
	
	def nist_ms_comparison(self, sample_name, mass_list, mass_spec, n_hits=5):
		"""
		
		
		:param sample_name:
		:type sample_name:
		:param mass_list:
		:type mass_list:
		:param mass_spec:
		:type mass_spec:
		:param n_hits:
		:type n_hits:
		
		:return:
		:rtype:
		"""
		
		data_dict = {}
		
		try:
			pynist.generate_ini(self.config.nist_path, "mainlib", n_hits)
			
			def remove_chars(input_string):
				for i in range(n_hits + 1):
					input_string = input_string.replace("Hit {}  : ", "")
				
				return input_string.replace("MF:", "") \
					.replace(":", "").replace("<", "").replace(">", "") \
					.replace(" ", "").replace(sample_name, "")
			
			raw_output = pynist.nist_db_connector(self.config.nist_path,
												  os.path.join(self.config.msp_dir, "{}.MSP".format(sample_name)))
			
			# Process output
			for i in range(n_hits + 1):
				raw_output = raw_output.replace("Hit {}  : ".format(i), "Hit{};".format(i)) \
					.replace("Hit {} : ".format(i), "Hit{};".format(i)) \
					.replace("Hit {}: ".format(i), "Hit{};".format(i))
			
			raw_output = raw_output.replace("<<", '"').replace(">>", '"').split("\n")
			
			matches_dict = {}
			
			for i in range(1, n_hits + 1):
				row = list(csv.reader([raw_output[i]], delimiter=";", quotechar='"'))[0]
				
				matches_dict[row[0]] = {
					"Name": row[1], "MF": (row[3].replace("MF:", '').replace(" ", '')),
					"RMF": (row[4].replace("RMF:", '').replace(" ", '')),
					"CAS": (row[6].replace("CAS:", '').replace(" ", '')),
					"Lib": (row[8].replace("Lib:", '').replace(" ", ''))
				}
		
		except:
			traceback.print_exc()  # print the error
			pynist.reload_ini(self.config.nist_path)
			sys.exit(1)
		
		print("\r\033[KSearch Complete", end='')
		pynist.reload_ini(self.config.nist_path)
		return matches_dict
	
	def merge(self):
		"""
		Merge Function
		"""
		
		if self.PL_len <= 1:
			print("Require two or more samples to merge. Check ./list.txt or --samples and try again.")
			print("")
			sys.exit(print(__doc__))
		
		merge_list = ()
		
		for prefix in self.config.prefixList:
			# read combined file
			with open(os.path.join(self.config.csv_dir, "{}_COMBINED.csv".format(prefix)), "r") as combined_file:
				combined_file_reader = csv.reader(combined_file, delimiter=";", quotechar='"')
				merge_list = merge_list + ([row for row in combined_file_reader],)
		
		merge_list = merge_list + (["\n"] * len(merge_list[0]),)
		
		merged_csv = []
		for prefix in self.config.prefixList:
			merged_csv.append([prefix, '', '', '', '', '', '', '', ''])
		
		# from https://stackoverflow.com/questions/7946798/interleave-multiple-lists-of-the-same-length-in-python
		merged_csv += [val for tup in zip(*merge_list) for val in tup][self.PL_len:]
		
		# Write output
		with open(os.path.join(self.config.csv_dir, "{}_MERGED.csv".format(self.lot_name)), "w") as f:
			f.write(';'.join(list(itertools.chain(*merged_csv))).replace("\n;", "\n"))
	
	def match_counter(self, ms_comp_list, separator=";"):
		"""
		Counter
		Also generates final output

		:param ms_comp_list:
		:type ms_comp_list:
		:param separator:
		:type separator: str
		"""
		
		if self.PL_len == 1:
			print("Require two or more samples to process. Check ./list.txt or --samples and try again.")
			print("")
		
		f = open(os.path.join(
			self.config.csv_dir,
			"{}_MERGED.csv".format(self.lot_name)
		), "r")  # Open merged output of lot of samples
		
		csv_f = csv.reader(f, delimiter=separator)
		csv_input = []
		
		for row in csv_f:
			csv_input.append(row)
		peak_data = []
		
		n_hits = 10
		
		for peak in csv_input[2::11]:
			rt_data = [locale.atof(x) for x in peak[0::9] if x not in [None, '']]
			area_data = [locale.atof(x) for x in peak[1::9] if x is not None]
			index = (csv_input.index(peak))
			hits = []
			names = []
			output_data = {
				"average_rt": numpy.mean(rt_data),
				"average_peak_area": numpy.mean(area_data),
				"rt_data": rt_data, "area_data": area_data,
			}
			for hit in range(n_hits):
				for prefix in self.config.prefixList:
					hits.append(
						[prefix, {
							"hit": csv_input[index + hit + 1][2],
							"MF": csv_input[index + hit + 1][4],
							"RMF": csv_input[index + hit + 1][5],
							"Name": csv_input[index + hit + 1][6],
							"CAS": csv_input[index + hit + 1][7]
						}]
					)
					names.append(csv_input[index + hit + 1][6])
					csv_input[index + hit + 1] = csv_input[index + hit + 1][9:]
			
			names.sort()
			names_count = Counter(names)
			
			hits_data = []
			
			# get average match factor &c. for each compound
			for compound in names_count:
				mf_data = []
				rmf_data = []
				hit_num_data = []
				
				for prefix in self.config.prefixList:
					length_before = len(mf_data)
					for hit in hits:
						
						if (hit[1]["Name"] == compound) and (hit[0] == prefix):
							mf_data.append(float(hit[1]["MF"]))
							rmf_data.append(float(hit[1]["RMF"]))
							hit_num_data.append(int(hit[1]["hit"]))
							CAS = hit[1]["CAS"]
					if len(mf_data) == length_before:
						mf_data.append(numpy.nan)
						rmf_data.append(numpy.nan)
						hit_num_data.append(numpy.nan)
				
				hits_data.append({
					"Name": compound,
					"Count": names_count[compound],
					"average_MF": numpy.nanmean(mf_data),
					"average_RMF": numpy.nanmean(rmf_data),
					"average_hit": numpy.nanmean(hit_num_data),
					"CAS": CAS,
					"mf_data": mf_data,
					"rmf_data": rmf_data,
					"hit_num_data": hit_num_data
				})
			
			hits_data = sorted(hits_data, key=lambda k: (k["Count"], k["average_MF"], k["average_hit"]), reverse=True)
			output_data["hits"] = hits_data[:5]
			peak_data.append(output_data)
		
		# Matches Sheet
		with open(os.path.join(self.config.csv_dir, self.lot_name + "_MATCHES.csv"), "w") as matches_csv_output:
			matches_csv_output.write("{};;;".format(self.lot_name))
			
			for prefix in self.config.prefixList:
				matches_csv_output.write("{};;;".format(prefix))
			matches_csv_output.write("\n")
			
			matches_csv_output.write(
				"Retention Time;Peak Area;;" +
				"Page;RT;Area;" * self.PL_len +
				";Match Factor;;;;Reverse Match Factor;;;;Hit Number;;;;Retention Time;;;;Peak Area;;\n")
			matches_csv_output.write(
				"Name;CAS Number;Freq.;" +
				"Hit No.;Match;R Match;" * self.PL_len +
				";Mean;STDEV;%RSD;;Mean;STDEV;%RSD;;Mean;STDEV;%RSD;;Mean;STDEV;%RSD;;Mean;STDEV;%RSD\n")
			
			for peak in peak_data:
				matches_csv_output.write(
					"{average_rt};{average_peak_area};;".format(**peak) + ";".join(
						[str(val) for pair in zip([''] * self.PL_len, peak["rt_data"], peak["area_data"]) for val in
						 pair])
					+ ";;;;;;;;;;;;;;{average_rt};".format(**peak)
					+ "{};{};;".format(
						numpy.nanstd(peak["rt_data"]),
						numpy.nanstd(peak["rt_data"]) / peak["average_rt"])
					+ "{average_peak_area};".format(**peak)
					+ "{};{};;\n".format(
						numpy.nanstd(peak["area_data"]),
						numpy.nanstd(peak["area_data"]) / peak["average_peak_area"])
				)
				
				for hit in peak["hits"]:
					matches_csv_output.write(
						"{Name};{CAS};{Count};".format(**hit)
						+ ";".join(
							[str(val) for pair in zip(hit["hit_num_data"], hit["mf_data"], hit["rmf_data"]) for val in
							 pair])
						+ ";;{average_MF};".format(**hit)
						+ "{};{};".format((
							numpy.nanstd(hit["mf_data"])),
							numpy.nanstd(hit["mf_data"]) / hit["average_MF"])
						+ ";{average_RMF};".format(**hit)
						+ "{};{};".format((
							numpy.nanstd(hit["rmf_data"])),
							numpy.nanstd(hit["rmf_data"]) / hit["average_RMF"])
						+ ";{average_hit};".format(**hit)
						+ "{};{};\n".format((
							numpy.nanstd(hit["hit_num_data"])),
							numpy.nanstd(hit["hit_num_data"]) / hit["average_hit"])
					)
		
		"""Get list of CAS Numbers for compounds reported in literature"""
		with open("./lib/CAS.txt", "r") as f:
			CAS_list = f.readlines()
		for index, CAS in enumerate(CAS_list):
			CAS_list[index] = CAS.rstrip("\r\n")
		
		# Statistics Sheet
		statistics_full_output = open(os.path.join(
			self.config.csv_dir,
			"{}_STATISTICS_FULL.csv".format(self.lot_name)
		), "w")
		
		statistics_output = open(os.path.join(
			self.config.csv_dir,
			"{}_STATISTICS.csv".format(self.lot_name)
		), "w")
		
		statistics_lit_output = open(os.path.join(
			self.config.csv_dir,
			"{}_STATISTICS_LIT.csv".format(self.lot_name)
		), "w")
		
		statistics_header = ";".join([
										self.lot_name, '', '', '',
										"Retention Time", '', '', '',
										"Peak Area", '', '', '',
										"Match Factor", '', '', '',
										"Reverse Match Factor", '', '', '',
										"Hit Number", '', '', '',
										"MS Comparison", '', '', '',
										"\nName", "CAS Number", '', ''
									] + ["Mean", "STDEV", "%RSD", ''] * 6) + "\n"
		
		statistics_full_output.write(statistics_header)
		statistics_output.write(statistics_header)
		statistics_lit_output.write(statistics_header)
		
		# Initialise dictionary for chart data
		chart_data = {
			"Compound": [],
			"{} Peak Area".format(self.lot_name): [],
			"{} Standard Deviation".format(self.lot_name): []
		}
		
		for prefix in self.config.prefixList:
			chart_data[prefix] = []
		
		for peak, ms in zip(peak_data, ms_comp_list):
			peak["ms_comparison"] = ms
			
			write_peak(statistics_full_output, peak, ms)
			if peak["hits"][0]["Count"] > (self.PL_len / 2):  # Write to Statistics; TODO: also need similarity > 800
				write_peak(statistics_output, peak, ms)
				if peak["hits"][0]["CAS"].replace("-", "") in CAS_list:  # Write to Statistics_Lit
					write_peak(statistics_lit_output, peak, ms)
					# Create Chart Data
					chart_data["Compound"].append(peak["hits"][0]["Name"])
					for prefix, area in zip(self.config.prefixList, peak["area_data"]):
						chart_data[prefix].append(area)
					chart_data["{} Peak Area".format(self.lot_name)].append(numpy.mean(peak["area_data"]))
					chart_data["{} Standard Deviation".format(self.lot_name)].append(numpy.mean(peak["area_data"]))
		
		statistics_full_output.close()
		statistics_output.close()
		statistics_lit_output.close()
		
		with open(os.path.join(self.config.csv_dir, "{}_peak_data.json".format(self.lot_name)), "w") as jsonfile:
			for dictionary in peak_data:
				jsonfile.write(json.dumps(dictionary))
				jsonfile.write("\n")
		
		time.sleep(2)
		
		"""Convert to XLSX and format"""
		# GC-MS
		append_to_xlsx(
			os.path.join(self.config.csv_dir, "{}_MERGED.csv".format(self.lot_name)),
			os.path.join(self.config.results_dir, self.lot_name + "_FINAL.xlsx"),
			"GC-MS", overwrite=True, separator=";",
			toFloats=True
		)
		# Counter
		append_to_xlsx(
			os.path.join(self.config.csv_dir, "{}_MATCHES.csv".format(self.lot_name)),
			os.path.join(self.config.results_dir, self.lot_name + "_FINAL.xlsx"),
			"Matches",
			separator=";",
			toFloats=True
		)
		# Statistics_Full
		append_to_xlsx(
			os.path.join(self.config.csv_dir, "{}_STATISTICS_FULL.csv".format(self.lot_name)),
			os.path.join(self.config.results_dir, self.lot_name + "_FINAL.xlsx"),
			"Statistics_Full",
			separator=";",
			toFloats=True
		)
		# Statistics
		append_to_xlsx(
			os.path.join(self.config.csv_dir, "{}_STATISTICS.csv".format(self.lot_name)),
			os.path.join(self.config.results_dir, self.lot_name + "_FINAL.xlsx"),
			"Statistics",
			separator=";",
			toFloats=True
		)
		# Statistics_Lit
		append_to_xlsx(
			os.path.join(self.config.csv_dir, "{}_STATISTICS_LIT.csv".format(self.lot_name)),
			os.path.join(self.config.results_dir, self.lot_name + "_FINAL.xlsx"),
			"Statistics - Lit Only",
			separator=";",
			toFloats=True
		)
		
		self.format_xlsx(os.path.join(self.config.results_dir, self.lot_name + "_FINAL.xlsx"))
		
		"""Charts"""
		chart_data = pandas.DataFrame(chart_data)
		print(chart_data)
		return chart_data
	
	def ms_comparisons(self, ms_data):
		"""
		Between Samples Spectra Comparison

		:param ms_data:
		:type ms_data:

		:return:
		:rtype:
		"""
		
		perms = []
		for i in itertools.permutations(self.config.prefixList, 2):
			# from https://stackoverflow.com/questions/10201977/how-to-reverse-tuples-in-python
			if i[::-1] not in perms:
				perms.append(i)
		
		ms_comparison = []
		
		rows_list = []
		
		for row_idx in range(len(ms_data)):
			similarity_list = []
			rows_list.append((ms_data.iloc[row_idx], perms))
		
		with Pool(len(ms_data)) as p:
			ms_comparison = p.map(single_ms_comparison, rows_list)
		
		# TODO: linear mode
		
		# for row_idx in range(len(ms_data)):
		# similarity_list = []
		# row = ms_data.iloc[row_idx]
		# for perm in perms:
		
		# top_spec = numpy.column_stack((row.loc[perm[0]].mass_list, row.loc[perm[0]].mass_spec))
		# bottom_spec = numpy.column_stack((row.loc[perm[1]].mass_list, row.loc[perm[1]].mass_spec))
		# similarity_list.append(SpectrumSimilarity.SpectrumSimilarity(top_spec, bottom_spec, t = 0.25, b = 1,
		# xlim = (45, 500), x_threshold = 0, print_graphic = False)[0]*1000)
		
		# #ms_comparison.append(rounders((numpy.mean(similarity_list)*100),"0"))
		# ms_comparison.append(similarity_list)
		
		return ms_comparison
	
	def format_xlsx(self, input_file):
		"""
		Formatting for Combined Output

		:param input_file: xlsx file to format
		:type input_file: str
		"""
		
		print('\nGenerating XLSX Output...')
		
		# From http://openpyxl.readthedocs.io/en/default/
		wb = load_workbook(input_file)
		
		"""GC-MS"""
		ws = wb["GC-MS"]
		number_format_list = make_column_property_dict(
			{'1': '0.000', '2': '0.0', '3': '0', '5': '0', '6': '0'},
			repeat=self.PL_len,
			length=9
		)
		width_dict = make_column_property_dict(
			{"1": 14, "2": 12, "4": 0, "5": 9, "6": 9, "8": 15, "9": 10},
			repeat=self.PL_len,
			length=9
		)
		alignment_list = make_column_property_dict({"5": "center", "6": "center", "8": "center"})
		format_sheet(ws, number_format_list, width_dict, alignment_list)
		
		for offset in range(self.PL_len):
			merge_string = get_column_letter(1 + (9 * offset)) + '1:' + get_column_letter(9 + (9 * offset)) + '1'
			ws.merge_cells(merge_string)
		
		format_header(ws, make_column_property_dict(
			{"1": "center", "2": "center", "5": "center", "6": "center", "8": "center", "9": "center"},
			repeat=self.PL_len,
			length=9
		), 1, 2)
		
		"""Matches"""
		ws = wb["Matches"]
		header_number_format_dict = make_column_property_dict(
			{'5': '0.000', '6': '0.00'}, {"A": "0.000", "B": "0.00"},
			{"17": "0.000", "18": '0.000000', "19": '0.00%', "21": "0.00", "22": "0.00", "23": '0.00%'},
			repeat=self.PL_len,
			length=3
		)
		header_h_alignment_list = make_column_property_dict(
			{'5': "right", '6': "right", }, {"A": "right", "B": "right"},
			{"17": "right", "18": "right", "19": "right", "21": "right", "22": "right", "23": "right"},
			repeat=self.PL_len,
			length=3
		)
		# header_v_alignment_list = make_column_property_dict(
		# 	{'5': "center", '6': "center"},
		# 	{"A": "center", "B": "center", "AF": "center", "AG": "center", "AH": "center", "AJ": "center",
		# 	 "AK": "center", "AL": "center"})
		hits_number_format_dict = make_column_property_dict(
			{'4': '0', '5': '0', '6': '0'}, {"C": "0"},
			{"5": "0.0", "6": "0.0000", "7": "0.00%", "9": "0.0", "10": "0.0000", "11": "0.00%", "13": "0.0",
			 "14": "0.0000", "15": "0.00%"},
			repeat=self.PL_len,
			length=3
		)
		hits_alignment_dict = make_column_property_dict(
			{'4': 'center', '5': 'center', '6': 'center'},
			{"B": "center", "C": "center"},
			{"13": "center"},
			repeat=self.PL_len,
			length=3
		)
		width_dict = make_column_property_dict(
			{"4": 8, "5": 8, "6": 11},
			{"B": 12},
			{"5": 6, "6": 9, "7": 9, "9": 6, "10": 9, "11": 9, "13": 5, "14": 7,
			 "15": 9, "17": 8, "18": 9, "19": 9, "21": 14, "22": 12, "23": 9,
			 "4": 1, "8": 1, "12": 1, "16": 1, "20": 1},
			repeat=self.PL_len,
			length=3
		)
		
		format_matches(
			ws, header_number_format_dict, header_h_alignment_list,
			hits_number_format_dict, hits_alignment_dict, width_dict
		)
		
		format_header(ws, make_column_property_dict(
			{"4": "center"},
			repeat=self.PL_len,
			length=3
		))  # First Row
		format_header(ws, make_column_property_dict(
			{"2": "center"},
			{"A": "right"},
			repeat=((self.PL_len * 3) + 25),
			length=1
		), 2, 2)  # Second Row
		format_header(ws, make_column_property_dict(
			{"2": "center"},
			repeat=((self.PL_len * 3) + 25),
			length=1
		), 3, 3)  # Third Row
		
		for offset in range(self.PL_len):
			merge_string = get_column_letter(4 + (3 * offset)) + '1:' + get_column_letter(6 + (3 * offset)) + '1'
			ws.merge_cells(merge_string)
		
		offset = self.PL_len * 3
		for index in [5, 9, 13, 17, 21]:
			merge_string = get_column_letter(index + offset) + '2:' + get_column_letter(index + 2 + offset) + '2'
			ws.merge_cells(merge_string)
		
		"""Statistics"""
		number_format_list = {
			'C': '0', 'E': '0.000', 'F': '0.000000', 'G': '0.00%', 'I': '0.00', 'J': '0.00', 'K': '0.00%',
			'M': '0.0', 'N': '0.0000', 'O': '0.00%', 'Q': '0.0', 'R': '0.0000', 'S': '0.00%', 'U': '0.0',
			'V': '0.0000', 'W': '0.00%', 'Y': '0.0', 'Z': '0.000', 'AA': '0.000%'
		}
		width_dict = {
			"B": 12, "C": 2, "D": 1, "E": 8, "F": 11, "G": 8, "H": 1, "I": 13, "J": 12, "K": 10, "L": 1, "M": 8,
			"N": 10, "O": 10, "P": 1, "Q": 8, "R": 10, "S": 10, "T": 1, "U": 6, "V": 8, "W": 10, "X": 1, "Y": 8,
			"Z": 9, "AA": 9
		}
		alignment_list = {"B": "center", "C": "center"}
		
		for sheet in ["Statistics_Full", "Statistics", "Statistics - Lit Only"]:
			ws = wb[sheet]
			format_sheet(ws, number_format_list, width_dict, alignment_list)
			for offset in range(6):
				ws.merge_cells(get_column_letter(5 + (4 * offset)) + '1:' + get_column_letter(7 + (4 * offset)) + '1')
			
			format_header(ws, make_column_property_dict({"2": "center"}, repeat=27, length=1), 1, 2)
		
		"""Contents Page"""
		ws = wb.create_sheet("Index", 0)
		contents = [
			["GC-MS", "Combined GC-MS data aligned by retention time.", True],
			["Matches", "List of possible matching compounds for each retention time, based on all samples.", True],
			["Statistics_Full", "Statistics for the top hit for each retention time.", True],
			[
				"Statistics",
				"Statistics for the top hit for each retention time, filtered by the occurrence of the compound\
in the hits, and the mass spectral similarity between samples.",
				True
			],
			[
				"Statistics - Lit Only",
				"As above, but only for compounds reported in literature as being present in propellant or GSR.",
				True
			],
		]
		
		print("\nThe worksheets in the output xlsx file are as follows:")
		ws.append(["GunShotMatch Version {} Output".format(__version__)])
		for row in contents:
			print(row[0] + " " * (24 - len(row[0])) + row[1])
			if row[2]:
				ws.append(["", '=HYPERLINK("#\'{0}\'!A1","{0}")'.format(row[0]), row[1]])
			else:
				ws.append(["", "{0}".format(row[0]), row[1]])
		
		br()
		
		ws.column_dimensions["B"].width = 50.0
		for row in ws.iter_rows(min_row=2, max_row=len(contents) + 1, min_col=2, max_col=2):
			for cell in row:
				if cell.value:
					if cell.value.startswith("=HYPERLINK"):
						cell.font = Font(
							color="0000EE",
							underline="single"
						)
		
		# ws.cell(column=4, row=2).value = "Sample list:"
		# for index, prefix in enumerate(prefixList):
		# 	ws.cell(row=2, column=(5+index)).value = prefix
		
		# Save the file
		wb.save(input_file)
		return
	
	def generate_spectra_from_alignment(self, rt_data, ms_data):
		"""
		Mass Spectra Images

		:param rt_data:
		:type rt_data:
		:param ms_data:
		:type ms_data:

		:return:
		:rtype:
		"""
		
		path = os.path.join(self.config.spectra_dir, self.lot_name)
		maybe_make(path)
		
		print("\nGenerating mass spectra images. Please wait.")
		
		# Delete Existing Files
		for filename in os.listdir(path):
			os.unlink(os.path.join(path, filename))
		
		if len(rt_data) > 20:
			arguments = [(sample, rt_data, ms_data, path) for sample in rt_data.columns.values]
			with Pool(len(arguments)) as p:
				
				p.map(self.spectrum_image_wrapper, arguments)
		else:
			for sample in rt_data.columns.values:
				self.generate_spectrum_image(sample, rt_data, ms_data, path)
	
	def spectrum_image_wrapper(self, arguments):
		return self.generate_spectrum_image(*arguments)
	
	def generate_spectrum_image(self, sample, rt_data, ms_data, path):
		"""
		
		:param sample:
		:type sample:
		:param rt_data:
		:type rt_data:
		:param ms_data:
		:type ms_data:
		:param path:
		:type path:
		
		:return:
		:rtype:
		"""
		
		# return sample
		from GSMatch.GSMatch_Core.charts import PlotSpectrum
		
		for row_idx in range(len(rt_data)):
			rt = rt_data.iloc[row_idx].loc[sample]
			ms = ms_data.iloc[row_idx].loc[sample]
			
			# TODO: Use mass range given in settings
			
			PlotSpectrum(
				numpy.column_stack((ms.mass_list, ms.mass_spec)),
				label="{} {}".format(sample, rounders(rt, "0.000")),
				xlim=(45, 500),
				mode=path
			)
		
		return
	
	def make_archive(self):
		"""
		Creates single tarball containing main results
		
		:return:
		:rtype:
		"""
		
		tar = tarfile.open(os.path.join(self.config.results_dir, (
				self.lot_name + "_" + datetime.datetime.now().strftime("%Y%m%d%H%M%S") + '.tar.gz')), mode='w:gz')
		
		tar.add(
			os.path.join(self.config.results_dir, self.lot_name + "_FINAL.xlsx"),
			arcname=(self.lot_name + ".xlsx"))
		tar.add(
			os.path.join(self.config.charts_dir, self.lot_name),
			arcname='Charts')
		tar.add(
			os.path.join(self.config.spectra_dir, self.lot_name),
			arcname='Spectra')
		tar.add(
			os.path.join(self.config.csv_dir, self.lot_name + "_MERGED.csv"),
			arcname='Merged.CSV')
		tar.add(
			os.path.join(self.config.csv_dir, self.lot_name + "_MATCHES.csv"),
			arcname='Matches.CSV')
		tar.add(
			os.path.join(self.config.csv_dir, "{}_STATISTICS_FULL.csv".format(self.lot_name)),
			arcname='Statistics/Statistics_Full.CSV')
		tar.add(
			os.path.join(self.config.csv_dir, "{}_STATISTICS.csv".format(self.lot_name)),
			arcname='Statistics/Statistics.CSV')
		tar.add(
			os.path.join(self.config.csv_dir, "{}_STATISTICS_LIT.csv".format(self.lot_name)),
			arcname='Statistics/Statistics_Lit.CSV')
		tar.add(
			os.path.join(self.config.csv_dir, "{}_peak_data.json".format(self.lot_name)),
			arcname="Peak_Data.json")
		tar.add(
			os.path.join(self.config.results_dir, f"{self.lot_name}.info"),
			arcname="sample_list.info")
		
		for prefix in self.config.prefixList:
			tar.add(
				os.path.join(self.config.expr_dir, "{}_tic.dat".format(prefix)),
				arcname="Experiments/{}_tic.dat".format(prefix))
			tar.add(
				os.path.join(self.config.expr_dir, "{}_peaks.dat".format(prefix)),
				arcname="Experiments/{}_peaks.dat".format(prefix))
			tar.add(
				os.path.join(self.config.expr_dir, "{}.expr".format(prefix)),
				arcname="Experiments/{}_expr.dat".format(prefix))
			tar.add(
				os.path.join(self.config.csv_dir, "{}_COMBINED.csv".format(prefix)),
				arcname="Combined/{}.CSV".format(prefix))
		
		print(f"\nSaved as: {os.path.join(self.config.results_dir, self.lot_name + '_FINAL.xlsx')}")


if __name__ == '__main__':
	from utils import timing
	from domdf_python_tools.terminal import clear
	
	clear()  # clear the display
	startup_string = f"""

{program_name} Version {__version__} running on {platform.system()}.
{copyright} Dominic Davis-Foster.
"""
	print(startup_string)
	
	gsm = NewProject()
	
	args = parse_arguments()
	
	if args.samples:
		gsm.config.prefixList = args.samples
	# overrides whatever was set from the config file
	
	# """Define Exit Functions"""
	# if "-h" not in str(sys.argv):
	# 	atexit.register(pynist.reload_ini, nist_path)
	
	if args.name:
		gsm.lot_name = args.name
	
	gsm.run()

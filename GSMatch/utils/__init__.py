#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
#  GunShotMatch Utilities Module
#
#  Shared Functions for GunShotMatch  
#
#  Copyright 2018, 2019 Dominic Davis-Foster <dominic@davis-foster.co.uk>
#  
#  This package is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 2 of the License, or
#  (at your option) any later version.
#  
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#  
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#  
# 
#  Includes code by the following developers:
#		TehTechGuy <https://stackoverflow.com/users/1117216/tehtechguy>
#		Triptych <https://stackoverflow.com/users/43089/triptych>
#		PaulMcG <https://stackoverflow.com/users/165216/paulmcg>
#
#  Includes a modified version of MassSpectraPlot,
#   originally by Martin N. Copyright 2015
#	Licenced under The MIT Licence
#
#  Includes sha-1 Directory Hash from ActiveState Code
#	Copyright Stephen Akiki 2009
#	Licenced under The MIT Licence

__all__ = ["utils", "helper", "MassSpectraPlot", "pynist", "timing", "DirectoryHash", "notification"]

import sys, csv, os

col_list = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ','BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ','CA','CB','CC','CD','CE','CF','CG','CH','CI','CJ','CK','CL','CM','CN','CO','CP','CQ','CR','CS','CT','CU','CV','CW','CX','CY','CZ','DA','DB','DC','DE','DF'] #Only Supports 12 samples

def append_to_xlsx(csv_input_file, xlsx_output_file, sheet_title, seperator=",", overwrite = False, use_io = False):
	from openpyxl import  Workbook, worksheet, load_workbook	# https://openpyxl.readthedocs.io/en/default/
	
	if overwrite:
		wb = Workbook()
		ws = wb.active
		wb.remove_sheet(ws)
	else:
		wb = load_workbook(xlsx_output_file)
	ws = wb.create_sheet(sheet_title)
	ws = wb[sheet_title]
	#print(sheet_title)
	
	if use_io:
		import io
		f = io.open(csv_input_file, encoding='latin-1')
	else:
		f = open(csv_input_file)
	reader = csv.reader(f, delimiter=seperator)
	
	import traceback
	for row in reader:
		#print(row)
		try:
			ws.append(row)
		except:
			traceback.print_exc()	#print the error
			print(row)
	f.close()

	wb.save(xlsx_output_file)

def within6sec(values):
	results = []
	for index, value1 in enumerate(values):
		try:
			value2 = values[index+1]
			if value1 not in [0, None, ''] and value2 not in [0, None, '']:
				results.append((float(value1)-0.1) < (float(value2)) < (float(value1)+0.1))
			else:
				results.append(False)
		except:
			pass
	if any(x == False for x in results):
		return False
	else:
		return True

def within6secV2(values):
	values = list(set(filter(None,values)))
	values.sort()
#	print(values[0])
#	print(values[-1])
#	print(float(values[-1])-float(values[0]))
	return float(values[-1])-float(values[0]) <= 0.2

def within18sec(values):
	values = list(set(filter(None,values)))
	values.sort()
#	print(values[0])
#	print(values[-1])
#	print(float(values[-1])-float(values[0]))
	return float(values[-1])-float(values[0]) <= 0.6

def within1min(value1, value2):
	if value1 not in [0, None, ''] and value2 not in [0, None, '']:
		return (float(value1)-1) < (float(value2)) < (float(value1)+1)
	else:
		return False

def stat_format(wsX, sheet_type="Statistics"):
			from openpyxl import  Workbook, worksheet, load_workbook	# https://openpyxl.readthedocs.io/en/default/
			from openpyxl.styles import Font, Fill, Alignment
			from openpyxl.utils import get_column_letter, column_index_from_string
			width_list2 = [32,14,3] + [1.5,7.5,6.5,9] + [1.5,12,12,9] 
			if sheet_type == "Statistics":
				width_list2 = width_list2 + [1.5,6.5,8,9]*2 + [1.5,6,6.5,9]	+ [1.5,6,8,9]#+ [1.5,9,9]*2
			elif sheet_type == "Comparison":
				width_list2 = width_list2 + [1.5,7.5,6.5,9] + [1.5,12,12,9] + [1.5,10,10,6]*3 + [1.5,6,8,9]
			
			for column_index, width in enumerate(width_list2):
				column_name = get_column_letter(column_index+1)
				col_width = width_list2[column_index]
				wsX.column_dimensions[column_name].width = float(col_width)
				
				"""From https://stackoverflow.com/questions/38619471/iterate-through-all-rows-in-specific-column-openpyxl"""
				for row in wsX.iter_rows('{0}{1}:{0}{2}'.format(column_name,wsX.min_row,wsX.max_row)):
					for index, cell in enumerate(row):
						if column_index > 0:
							cell.alignment = Alignment(horizontal='center',
													vertical='center',
													wrap_text=True)	
						elif column_index == 1:
							cell.alignment = Alignment(horizontal='center')
						else:
							cell.alignment = Alignment(wrap_text=True)
						try:
							float(cell.internal_value)
							cell.value = float(cell.internal_value)
							if column_index > 3:
								if column_index in [5,8,9,12,16,20,24]:
			#						if column_index == 5:
			#							cell.value = (cell.value)*100000
									cell.number_format = "0.0"
								elif column_index in [4,13,17,21,25,28,29,31,32]:
									cell.number_format = "0.000"
								elif column_index in [6,10,14,18,22,26]:
									#cell.style = "Percent"
									cell.number_format = "0.000%"
								
								if sheet_type == "Comparison":
									if column_index in [20,24,28,12,33]:
										cell.number_format = "0.000"
									if column_index in [21,25,29]:
										cell.number_format = "0.00000"
									if column_index in [32,17]:
										cell.number_format = "0.0"
									if column_index in [34]:
										cell.number_format = "0.000%"
									
						except:
							cell.value = cell.internal_value
			
			#wsX.cell("F2").value = "STDEV E5"
		#   ^ Uncomment these lines to enable STDEV*100000
			if sheet_type == "Statistics":
				wsX.merge_cells("E1:G1")
				wsX.merge_cells("I1:K1")
				wsX.merge_cells("M1:O1")
				wsX.merge_cells("Q1:S1")
				wsX.merge_cells("U1:W1")
				wsX.merge_cells("Y1:AA1")
		#		wsX.merge_cells("AC1:AD1")
		#		wsX.merge_cells("AF1:AG1")
			elif sheet_type == "Comparison":
				wsX.merge_cells("E1:K1")
				wsX.merge_cells("M1:S1")
				wsX.merge_cells("U1:AE1")
				
				wsX.merge_cells("E2:G2")
				wsX.merge_cells("I2:K2")
				wsX.merge_cells("M2:O2")
				wsX.merge_cells("Q2:S2")
				wsX.merge_cells("U2:W2")
				wsX.merge_cells("Y2:AA2")
				wsX.merge_cells("AC2:AE2")
				wsX.merge_cells("AG2:AI2")

def maybe_make(directory):
	# makes a directory only if it doesn't already exist
	if not os.path.exists(directory):
		os.makedirs(directory)

def load_config(configfile):
	import ConfigParser
	import platform
	"""Configuration -----"""
	Config = ConfigParser.ConfigParser()
	Config.read(configfile)
	
	if platform.system() == "Linux":
		lib2nist_path = Config.get("main", "LinuxLibNistPath")
		if len(lib2nist_path) == 0:
			print("Error: Lib2Nist path not found in configuration file.")
			sys.exit(1)

		nist_path = Config.get("main", "LinuxNistPath")
		if len(nist_path) == 0:
			print("Error: NIST MS Search path not found in configuration file.")
			sys.exit(1)
	else:
		lib2nist_path = Config.get("main", "LibNistPath")
		if len(lib2nist_path) == 0:
			print("Error: Lib2Nist path not found in configuration file.")
			sys.exit(1)

		nist_path = Config.get("main", "NistPath")
		if len(nist_path) == 0:
			print("Error: NIST MS Search path not found in configuration file.")
			sys.exit(1)

	if platform.system() == "Windows":
		LO_Path = os.path.join(Config.get("main","WinLOPath"),"scalc.exe")
	else:
		LO_Path = "libreoffice"
		
	CSV_DIRECTORY = os.path.abspath(Config.get("main","CSVPath"))
	OUTPUT_DIRECTORY = os.path.abspath(Config.get("main","OutputPath"))		#Gets created if not present
	SPECTRA_DIRECTORY = os.path.abspath(Config.get("main","SpectraPath"))
	XY_DIRECTORY = os.path.abspath(Config.get("main","xyPath"))				#Gets created if not present
	MSP_DIRECTORY = os.path.abspath(Config.get("main","MSPPath"))			#Gets created if not present
	COMPARISON_DIRECTORY = os.path.abspath(Config.get("main","ImagesPath"))	#Gets created if not present
	RESULTS_DIRECTORY = os.path.abspath(Config.get("main","ResultsPath"))	#Gets created if not present
	
	"""Create the obove folders if they don't exist"""
	# if not os.path.exists(OUTPUT_DIRECTORY):
		# os.makedirs(OUTPUT_DIRECTORY)
	# if not os.path.exists(XY_DIRECTORY):
		# os.makedirs(XY_DIRECTORY)
	# if not os.path.exists(MSP_DIRECTORY):
		# os.makedirs(MSP_DIRECTORY)
	# if not os.path.exists(COMPARISON_DIRECTORY):
		# os.makedirs(COMPARISON_DIRECTORY)
	# if not os.path.exists(RESULTS_DIRECTORY):
		# os.makedirs(RESULTS_DIRECTORY)
	maybe_make(OUTPUT_DIRECTORY)
	maybe_make(XY_DIRECTORY)
	maybe_make(MSP_DIRECTORY)
	maybe_make(COMPARISON_DIRECTORY)
	maybe_make(RESULTS_DIRECTORY)
	
	
	"""Autoprocess1"""
	auto1,auto2,auto3 = [],[],[]
	for step in ["combine","merge","open_lo","counter","jigsaw"]:
		if Config.getboolean("Auto1", step):
			auto1.append(step)

	"""Autoprocess2"""
	for step in ["combine","merge","open_lo","counter","jigsaw"]:
		if Config.getboolean("Auto2", step):
			auto2.append(step)

	"""Autoprocess3"""
	for step in ["combine","merge","open_lo","counter","jigsaw"]:
		if Config.getboolean("Auto3", step):
			auto3.append(step)
	
	"""Sample Lists"""
	prefixList = Config.get("samples","samples").replace(";",",").replace("\t",",").replace(" ","").split(",")
	
	
		
	return lib2nist_path, nist_path, LO_Path, CSV_DIRECTORY, OUTPUT_DIRECTORY, SPECTRA_DIRECTORY, XY_DIRECTORY, MSP_DIRECTORY, COMPARISON_DIRECTORY, RESULTS_DIRECTORY, auto1, auto2, auto3, prefixList		

class GSMConfig(object):
	def __init__(self, configfile):
		self.configfile = configfile
		print("\nUsing configuration file {}".format(configfile))
		self.get_config(self.configfile)
		
	def get_config(self, configfile, parent=None):
		# Returns configuration to parent if set, if not to self
		
		if parent is None:
			parent=self
		
		import platform
		import sys
		
		if sys.version_info[0] == 3:
			import configparser as ConfigParser
			from itertools import chain, permutations
		else:
			import ConfigParser
			from itertools import chain, izip, permutations
		
		"""Configuration -----"""
		Config = ConfigParser.ConfigParser()
		Config.read(configfile)
	
		if platform.system() == "Linux":
			lib2nist_path = Config.get("main", "LinuxLibNistPath")
			if len(lib2nist_path) == 0:
				print("Error: Lib2Nist path not found in configuration file.")
				sys.exit(1)

			nist_path = Config.get("main", "LinuxNistPath")
			if len(nist_path) == 0:
				print("Error: NIST MS Search path not found in configuration file.")
				sys.exit(1)
		else:
			lib2nist_path = Config.get("main", "LibNistPath")
			if len(lib2nist_path) == 0:
				print("Error: Lib2Nist path not found in configuration file.")
				sys.exit(1)

			nist_path = Config.get("main", "NistPath")
			if len(nist_path) == 0:
				print("Error: NIST MS Search path not found in configuration file.")
				sys.exit(1)

		if platform.system() == "Windows":
			LO_Path = os.path.join(Config.get("main","WinLOPath"),"scalc.exe")
		else:
			LO_Path = "libreoffice"
			
		CSV_DIRECTORY = os.path.abspath(Config.get("main","CSVPath"))
		OUTPUT_DIRECTORY = os.path.abspath(Config.get("main","OutputPath"))		#Gets created if not present
		SPECTRA_DIRECTORY = os.path.abspath(Config.get("main","SpectraPath"))
		XY_DIRECTORY = os.path.abspath(Config.get("main","xyPath"))				#Gets created if not present
		MSP_DIRECTORY = os.path.abspath(Config.get("main","MSPPath"))			#Gets created if not present
		COMPARISON_DIRECTORY = os.path.abspath(Config.get("main","ImagesPath"))	#Gets created if not present
		RESULTS_DIRECTORY = os.path.abspath(Config.get("main","ResultsPath"))	#Gets created if not present
		
		"""Create the obove folders if they don't exist"""
		maybe_make(OUTPUT_DIRECTORY)
		maybe_make(XY_DIRECTORY)
		maybe_make(MSP_DIRECTORY)
		maybe_make(COMPARISON_DIRECTORY)
		maybe_make(RESULTS_DIRECTORY)
		
		"""Autoprocess1"""
		auto1,auto2,auto3 = [],[],[]
		for step in ["combine","merge","open_lo","counter","jigsaw"]:
			if Config.getboolean("Auto1", step):
				auto1.append(step)

		"""Autoprocess2"""
		for step in ["combine","merge","open_lo","counter","jigsaw"]:
			if Config.getboolean("Auto2", step):
				auto2.append(step)

		"""Autoprocess3"""
		for step in ["combine","merge","open_lo","counter","jigsaw"]:
			if Config.getboolean("Auto3", step):
				auto3.append(step)
		
		"""Sample Lists"""
		prefixList = Config.get("samples","samples").replace(";",",").replace("\t",",").replace(" ","").split(",")
		
		parent.lib2nist_path = lib2nist_path
		parent.nist_path = nist_path
		parent.LO_Path = LO_Path
		parent.CSV_DIRECTORY = CSV_DIRECTORY
		parent.OUTPUT_DIRECTORY = OUTPUT_DIRECTORY
		parent.SPECTRA_DIRECTORY = SPECTRA_DIRECTORY
		parent.XY_DIRECTORY = XY_DIRECTORY
		parent.MSP_DIRECTORY = MSP_DIRECTORY
		parent.COMPARISON_DIRECTORY = COMPARISON_DIRECTORY
		parent.RESULTS_DIRECTORY = RESULTS_DIRECTORY
		parent.auto1 = auto1
		parent.auto2 = auto2
		parent.auto3 = auto3
		parent.prefixList = prefixList
	
	def get_all(self):
		self.get_config(self.configfile)
		return self.lib2nist_path, self.nist_path, self.LO_Path, \
			self.CSV_DIRECTORY, self.OUTPUT_DIRECTORY, \
			self.SPECTRA_DIRECTORY, self.XY_DIRECTORY, \
			self.MSP_DIRECTORY, self.COMPARISON_DIRECTORY, \
			self.RESULTS_DIRECTORY, self.auto1, self.auto2, self.auto3, \
			self.prefixList		


if __name__ == '__main__':
    sys.exit(1)

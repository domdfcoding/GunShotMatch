#!/usr/bin/env python
# -*- coding: utf-8 -*-
# encoding=utf8
#
#  GSMatch.py
#  
#  Copyright 2017-2019 Dominic Davis-Foster <dominic@davis-foster.co.uk>
#  
#  This program is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 3 of the License, or
#  (at your option) any later version.
#  
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#  
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#   

"""GunShotMatch 0.9.0

Program for the analysis of OGSR samples to identify matching compounds 
between samples. Can currently process PerkinElmer GC-MS .CSV files only.
	
Samples to be analysed can be passed following the -f/--samples argument, 
or listed in a file called "list.txt" in the base directory. Each sample 
name must be on a new line. For more information read the documentation.
	
The files must be in the subdirectory CSV and called 
		e.g. GECO_Case2_MS.CSV and GECO_Case2_GC.CSV

PerkinElmer is a trademark of PerkinElmer Inc.

Tested with Python 2 ONLY.
"""

from __future__ import print_function #for py2

program_name ="GunShotMatch"
__version__ = '0.9.0'
copyright = "2017-2019"

verbose = True

"""Imports"""
from utils import timing # Times the program
import sys
sys.path.append("..")

if sys.version_info[0] == 3:
	import configparser as ConfigParser
	from itertools import chain, permutations
else:
	import ConfigParser
	from itertools import chain, izip, permutations

import  logging
import os
import time
import argparse
import csv
import datetime
import warnings
import re
import math
import shutil
import atexit
import platform
import numpy
import progressbar

from inspect import currentframe, getframeinfo
from subprocess import Popen, PIPE, STDOUT
from decimal import Decimal, ROUND_HALF_UP
from collections import Counter
from openpyxl import  Workbook, worksheet, load_workbook	# https://openpyxl.readthedocs.io/en/default/
from openpyxl.styles import Font, Fill, Alignment
from openpyxl.chart import BarChart, RadarChart, Series, Reference, LineChart
from openpyxl.utils import get_column_letter, column_index_from_string
#from shutil import copyfile
from utils.helper import entry,  parent_path, RepresentsInt, rounders
from utils import append_to_xlsx, load_config, stat_format, within6sec 
from utils.pynist import *
from utils import DirectoryHash
from utils.helper import clear
from utils import GSMConfig

"""From https://stackoverflow.com/questions/21129020/how-to-fix-unicodedecodeerror-ascii-codec-cant-decode-byte
Big shout out to fisherman https://stackoverflow.com/users/2309581/fisherman"""
try: # python2
	reload(sys) 
	sys.setdefaultencoding('utf8')
except: # python3
	pass # no longer necessary for py3

def verbosePrint(text):	#print only if the --verbose flag is set
	#if verbose:
	#	print(str(text))
	logger.debug(str(text))

"""Combine functions"""
def GCMScombine(inputFile, OUTPUT_DIRECTORY, CSV_DIRECTORY):
	"""Define variables"""
	GCpeaks=[] #Stores relevant data for GC peaks, in this case the RT and peak area
	stripCSVms=[] #Stores the MS data after removing redundant lines
	
	"""Open the output file"""
	while True:
		try:
			outputCSV = open(os.path.join(OUTPUT_DIRECTORY, inputFile+".CSV"),"w")
			break
		except IOError:
			print("The file \"" + inputFile + ".CSV\" is locked for editing in another program. Press [Return] to try again.")
			raw_input(">") #fix
	
	"""Read GC Data"""
	with open(os.path.join(CSV_DIRECTORY,inputFile+"_GC_80.CSV"),"r") as f:
		inputCSVgc = f.readlines()
	
	"""Pull relevant GC Data"""
	for lineNo in range(2,len(inputCSVgc)): 
		line = inputCSVgc[lineNo]
		for x in range(0,81):
			if line[:2]== str((["1,","2,","3,","4,","5,","6,","7,","8,","9,"] + range(10,81)+['',''])[x]):
				GCpeaks.append(line.split(",")[1] + "," + line.split(",")[4] + ",")
		#print(str(GCpeaks) + "\n")
		
	"""Remove redundant rows from GC-MS"""
	with open(os.path.join(CSV_DIRECTORY,inputFile+"_MS_80.CSV"),"r") as f:	#open MS data
		inputCSVms = f.readlines()						#read lines to a list
	
	reader = csv.reader(open(os.path.join(CSV_DIRECTORY,inputFile+"_MS_80.CSV")),delimiter=",", quotechar = '"',doublequote = True)
	inputCSVms = [row for row in reader]

	
	for lineNo in range(0,len(inputCSVms)):				
		line = inputCSVms[lineNo]

		#Include only rows that relate to compounds
		for page in range(1,81):	#Replace rows of "Page 1 of 80" with a single cell
			#if '"Page '+ str(page) +' of 80","Page '+ str(page) +' of 80"' in line:
			#print(line)
			
			if "Page "+ str(page) +" of 80" in line:
				stripCSVms.append("Page "+ str(page) +" of 80,,,,,\n")
		if "mainlib" in line:

			if line[5] == "" and len(line) == 7:

				stripCSVms.append('"{}","{}","{}","{}","{}","{}",""\n'.format(line[0],line[1],line[2],line[3],line[4],line[6]))
			else:
				stripCSVms.append('"{}","{}","{}","{}","{}","{}",""\n'.format(line[0],line[1],line[2],line[3],line[4],line[5]))


	"""Add spaces to MS Data"""	
	#for lineNo in range(0,len(stripCSVms)+80):
	for lineNo in range(1000):
		try:
			line = stripCSVms[lineNo]
		except IndexError:
			next
		
		try:	#add additional lines to ensure regular spacing
			if line[:4] == "Page" and (stripCSVms[lineNo-1])[:4] == "Page":
				for i in range(5):
					stripCSVms.insert(lineNo,",,,,,,\n")
				verbosePrint(str(lineNo) + "Insert 5")
			if line[:4] == "Page" and (stripCSVms[lineNo-1])[:3] == '"1"':
				for i  in range(4):
					stripCSVms.insert(lineNo,",,,,,,\n")
				verbosePrint(str(lineNo) + "Insert 4")
			if line[:4] == "Page" and (stripCSVms[lineNo-1])[:3] == '"2"':
				for i  in range(3):
					stripCSVms.insert(lineNo,",,,,,,\n")
				verbosePrint(str(lineNo) + "Insert 3")
			if line[:4] == "Page" and (stripCSVms[lineNo-1])[:3] == '"3"':
				stripCSVms.insert(lineNo,",,,,,,\n")
				stripCSVms.insert(lineNo,",,,,,,\n")
				verbosePrint(str(lineNo) + "Insert 2")
			if line[:4] == "Page" and (stripCSVms[lineNo-1])[:3] == '"4"':
				for i  in range(1):
					stripCSVms.insert(lineNo,",,,,,,\n")
				verbosePrint(str(lineNo) + "Insert 1")
		except IndexError:
			traceback.print_exc()
			
	"""Combine GC and MS Data"""
	#for lineNo in range(0,len(stripCSVms)+20):
	for lineNo in range(0,1500):
		try:
			line = stripCSVms[lineNo]
			#Add extra columns in front of lines for matches to ensure 
			if line[:4] != "Page":	# alignment after merging GC data
				stripCSVms[stripCSVms.index(line)] = ",,"+line
		except IndexError:
			continue
			
	for GCpeakNO in range(0,81):
		lineNo = 0							#Insert the GC data into the
		while lineNo < len(stripCSVms):		# applicable rows of MS data
			line = stripCSVms[lineNo]
			if line[:7] == "Page "+str((["1 ","2 ","3 ","4 ","5 ","6 ","7 ","8 ","9 "] + range(10,81) + ["",""])[GCpeakNO]):
				#print(GCpeaks[GCpeakNO])
				stripCSVms.insert(lineNo,GCpeaks[GCpeakNO])
				lineNo = lineNo+1
			lineNo = lineNo+1		
	
	"""Add column heading"""
	stripCSVms.insert(0,inputFile + "\nRetention Time,Peak Area,,Lib,Match,R Match,Name,CAS Number,Notes\n")
	
	"""Save	"""
	outputCSV.write(''.join(stripCSVms))
	outputCSV.close()
	print(inputFile + " Done!")

# Not used?
def getRTlist(sample):
	tempRTlist = []	#list of retention times for the sample passed as an argument
	
	f = open("./"+sample+".CSV", "r")		#Open GCMS merged output of sample
	csv_f = csv.reader(f)

	for row in csv_f:
		try:
			verbosePrint(str(row[0]))
			if row[0] not in ["Retention Time",sample]:
				if len(row[0]) > 0:					#Gathers retention
					tempRTlist.append(row[0])		# time data
		except IndexError:
			traceback.print_exc()		#Print error but continue

	f.close()
	return tempRTlist


"""Spacer Functions"""
def single_spacer(prefix, GCpeaks80, OUTPUT_DIRECTORY):
	"""Space a single sample"""
	spacedOutput = []
	csv_input = []
	f = open(os.path.join(OUTPUT_DIRECTORY,(prefix+".CSV")))
	csv_f = csv.reader(f)
	
	for row in csv_f:
		csv_input.append(row)

	for row in csv_input:
		if len(row)>7:
			if row[7] == "":
				del row[7]
		while len(row)<9:
			row.append("")
	
	spacedOutput.append(str(csv_input[0]))
	spacedOutput.append(str(csv_input[1]))
	
	done_times = []
	
	for time in GCpeaks80:
		for row_num in range(len(csv_input)):
			if row_num not in [0,1]:
				if len(csv_input[row_num][0]) > 0:
					if csv_input[row_num][0] not in done_times:
						try:
							if csv_input[row_num][0] == time:
								spacedOutput.append(str(csv_input[row_num]))
								spacedOutput.append(str(csv_input[row_num+1]))
								spacedOutput.append(str(csv_input[row_num+2]))
								spacedOutput.append(str(csv_input[row_num+3]))
								spacedOutput.append(str(csv_input[row_num+4]))
								spacedOutput.append(str(csv_input[row_num+5]))
								done_times.append(time)
								break
							elif float(time) < float(csv_input[row_num][0]):
								for i in range(6):
									spacedOutput.append(str(['', '', '', '', '', '', '', '', '']))
								break
						except:
							#traceback.print_exc()
							#print(row_num)
							#print(time)
							break
		done_times.append(time)
			
	with open(os.path.join(OUTPUT_DIRECTORY,prefix + "_SPACED.CSV"),"w") as f:
		joinedOutput = '\n'.join(spacedOutput)
		f.write(joinedOutput.replace("[",'').replace("]",'').replace("\"",'\''))	
	
	f.close()
	
	return spacedOutput
		
def batch_spacer(prefixList,OUTPUT_DIRECTORY,CSV_DIRECTORY):
	"""Space batch of samples"""
	GCpeaks80 = get_top_80(prefixList,CSV_DIRECTORY)
	for prefix in prefixList:
		single_spacer(prefix, GCpeaks80,OUTPUT_DIRECTORY)

def get_top_80(prefixList, CSV_DIRECTORY):	
	"""Get list of top 80 peaks in all samples"""
	GCpeaks80 = []
	for prefix in prefixList:
		with open(os.path.join(CSV_DIRECTORY,prefix.rstrip("\n\r ")+"_GC_80.CSV"),"r") as f:
			inputCSVgc80 = f.readlines()
	
		for lineNo in range(2,len(inputCSVgc80)): 
			line = inputCSVgc80[lineNo]
			
			for x in range(0,81):
				if line.split(",")[0] == str((range(1,81) + ["",""])[x]):
					GCpeaks80.append(line.split(",")[1])
					#add extra '+ line.split(",")[x] + ",""' to include additional columns, and change headings later on
			verbosePrint(GCpeaks80)
	
	GCpeaks80 = list(set(GCpeaks80)) #Remove duplicates
	GCpeaks80.sort(key=float) #Sort list of retention times low to high

	return GCpeaks80



#@atexit.register
def final_cleanup(prefixList, OUTPUT_DIRECTORY,nist_path,lot_name):
	#if sys.argv[1] not in ["--help","-h"]:
		
		for prefix in prefixList:
			if os.path.isfile(os.path.join(OUTPUT_DIRECTORY, prefix + "_SPECTRA.CSV")):
				os.unlink(os.path.join(OUTPUT_DIRECTORY, prefix + "_SPECTRA.CSV"))
			if os.path.isfile(prefix + "_SPECTRA.CSV"):
				shutil.move(prefix + "_SPECTRA.CSV",os.path.join(OUTPUT_DIRECTORY, prefix + "_SPECTRA.CSV"))
		
		if os.path.isfile(os.path.join(OUTPUT_DIRECTORY, lot_name + "_COUNTER.CSV")):
			os.unlink(os.path.join(OUTPUT_DIRECTORY, lot_name + "_COUNTER.CSV"))
		if os.path.isfile(lot_name + "_COUNTER.CSV"):
			shutil.move(lot_name + "_COUNTER.CSV",os.path.join(OUTPUT_DIRECTORY, lot_name + "_COUNTER.CSV"))
		
		nist_cleanup(nist_path)
		
		import glob
		for filename in glob.glob(r'./ms_comparisons_*.txt'):
			if os.path.exists(os.path.join(OUTPUT_DIRECTORY,filename)):
				os.unlink(os.path.join(OUTPUT_DIRECTORY,filename))
			shutil.move(filename, OUTPUT_DIRECTORY)


#@atexit.register
def nist_cleanup(nist_path):
#	if sys.argv[1] not in ["--help","-h"]:
		reload_ini(nist_path)

def hasher(string):
	import hashlib
	return hashlib.sha1(str(string)).hexdigest()


class GunShotMatch(GSMConfig):
	def __init__(self, config_file):
		
		# set config file
		self.config_file = config_file
		print("\nUsing configuration file {}".format(self.config_file))
		# get configuration
		self.get_config(self.config_file, self)
		
		# Samples
		if len(self.prefixList) == 0:
			print("No samples specified. Please enter samples in list.txt or following the --samples argument\n.")
			sys.exit(1)
		
		for prefix in self.prefixList:
			if prefix == "":
				self.prefixList.remove("")
		
		self.PL_len = len(self.prefixList)
		self.lot_name = re.sub(r'\d+', '', str(self.prefixList[0].rstrip("\n\r "))).replace("__","_")
		
		print("\nReady")
		print("Samples to process: " + str(self.prefixList))

	
	
	
	"""Merge Functions"""
	def Merge(self):
		GCpeaks80 = get_top_80(self.prefixList,self.CSV_DIRECTORY)
		
		if self.PL_len == 1:
			print("Require two or more samples to combine. Check ./list.txt or --samples and try again.")
			print("")
			sys.exit(print(__doc__))

		if self.PL_len > 1:
			sample_space_list = {"sample_1_spaced": single_spacer(self.prefixList[0], GCpeaks80,self.OUTPUT_DIRECTORY), "sample_2_spaced": single_spacer(self.prefixList[1], GCpeaks80,self.OUTPUT_DIRECTORY)}

		for i in range(2,12):
			if self.PL_len > i:
				sample_space_list["sample_" + str(i+1) + "_spaced"] = single_spacer(self.prefixList[i], GCpeaks80,self.OUTPUT_DIRECTORY)
			else: sample_space_list["sample_" + str(i+1) + "_spaced"] = ['']*len(sample_space_list["sample_1_spaced"])
		
		for i in range(1,12):
			"""From https://stackoverflow.com/questions/10895567/find-longest-string-key-in-dictionary"""
			while len(sample_space_list["sample_" + str(i+1) + "_spaced"]) < max((len(v), k) for k,v in sample_space_list.iteritems())[0]:
				"""From https://stackoverflow.com/questions/26367812/appending-to-list-in-python-dictionary"""
				sample_space_list.setdefault("sample_" + str(i+1) + "_spaced", []).append(str(['', '', '', '', '', '', '', '', '']))
			verbosePrint(len(sample_space_list["sample_" + str(i+1) + "_spaced"]))
		
		newline_List = ["\n"]*len(sample_space_list["sample_1_spaced"])
		
		""" from https://stackoverflow.com/questions/11125212/interleaving-lists-in-python"""	
		single_CSV = list(chain.from_iterable(izip(sample_space_list["sample_1_spaced"],sample_space_list["sample_2_spaced"],sample_space_list["sample_3_spaced"],sample_space_list["sample_4_spaced"],
				sample_space_list["sample_5_spaced"],sample_space_list["sample_6_spaced"],sample_space_list["sample_7_spaced"],sample_space_list["sample_8_spaced"],sample_space_list["sample_9_spaced"],
				sample_space_list["sample_10_spaced"],sample_space_list["sample_11_spaced"],sample_space_list["sample_12_spaced"],newline_List)))

		output_file_name = self.lot_name + "_COMBINED.CSV"
		
		"""From https://stackoverflow.com/questions/2400504/easiest-way-to-replace-a-string-using-a-dictionary-of-replacements"""
		replace_dictionary = {"']['":'","', "['":'"', "']":'"', '"':"'", "', '":'","', "','":'","', "'\n'":'"\n"'}
		pattern = re.compile('|'.join(re.escape(key) for key in replace_dictionary.keys()))
		
		with open(output_file_name,"w") as f:
			joinedOutput = ''.join(single_CSV)
			f.write(('\"' + pattern.sub(lambda x: replace_dictionary[x.group()], joinedOutput)[1:]).replace("', '",'","'))		

		with open(os.path.join(self.OUTPUT_DIRECTORY,"FULL_" + output_file_name),"w") as f:
			joinedOutput = ''.join(single_CSV)
			f.write(('\"' + pattern.sub(lambda x: replace_dictionary[x.group()], joinedOutput)[1:]).replace("', '",'","'))		
		
		return output_file_name

	def jigsaw(self): 
		interactive = False
		
		if interactive: clear()
		
		from utils import terminalsize as tp
		
		if self.PL_len == 1:
			print("Require two or more samples to combine. Check ./list.txt or --samples and try again.")
			print("")
			sys.exit(print(__doc__))
		
		#Read CSV Data
		file_path = self.lot_name + "_COMBINED.CSV"
		
		with open(file_path, "r") as f:
			raw_input_csv = csv.reader(f)
			input_csv = []
			for row in raw_input_csv:
				input_csv.append(row)
		
		#Initialise output
		with open(file_path, 'w') as f:
			f.write(";".join(input_csv[0]))
			f.write("\n")
			f.write(";".join(input_csv[1]))
			f.write("\n")
			
		output_file = open(file_path, 'a')
		
		#Get cluster data, ignoring bleed
		index_list =[]
		cluster_dict = {}
		blanks_dict = {}
		peaks_dict = {}

		skip_list = []

		for index, row in enumerate(input_csv):
			bleed = False
			while True:
				if any(" of " in x for x in row):
					try:
						cluster = [[row[0+(i*9)] for i in range(self.PL_len)],[input_csv[index+1][6+(i*9)] for i in range(self.PL_len)], [input_csv[index+2][6+(i*9)] for i in range(self.PL_len)], [input_csv[index+3][6+(i*9)] for i in range(self.PL_len)], [input_csv[index+4][6+(i*9)] for i in range(self.PL_len)], [input_csv[index+5][6+(i*9)] for i in range(self.PL_len)]]
					except IndexError:
						break
					
					peaks = []
					for x in range(self.PL_len):
						if len(cluster[0][x]) != 0 and len(cluster[1][x]) != 0:
							peaks.append(x)
					
					for row2 in cluster:
						if any("silox" in x.lower() for x in row2):
							bleed = True
							break
						elif any("silane" in x.lower() for x in row2):
							bleed = True
							break
						elif any("TMS" in x.upper() for x in row2):
							bleed = True
							break
						elif any("Ethyl [5-hydroxy-1-(6-methoxy-4-methyl-3-quinolinyl)-3-methyl-1H-pyrazol-4-yl]acetate".lower() in x.lower() for x in row2):
							if len(peaks) == 1:
								bleed = True
								break
					
					if bleed:
						break
					if len(peaks) == 0:
						break
					
					index_list.append(index)
					cluster_dict[index] = cluster				
					peaks_dict[index] = peaks
					
				break
		
		def write_output():
			for row in range(0,6):
				output_row = []
				for cell_index in range(len(input_csv[index])):
					cell_temp = []
					if _next_2:
						cell_temp.append(input_csv[_next_2+row][cell_index])
						skip_list.append(i+3)
					if _next_1:
						cell_temp.append(input_csv[_next_1+row][cell_index])
						skip_list.append(i+2)
					if _next:
						cell_temp.append(input_csv[_next+row][cell_index])
						skip_list.append(i+1)		
					
					cell_temp.append(input_csv[index+row][cell_index])
					skip_list.append(i)

					cell_temp = filter(None, cell_temp) # from https://stackoverflow.com/questions/3845423/remove-empty-strings-from-a-list-of-strings				
					
					if len(cell_temp) == 0:
						cell_temp = ['']
					
					output_row.append(cell_temp[0])

				output_file.write(";".join(output_row))
				output_file.write("\n")
		# end of function	
		
		for i, index in enumerate(index_list):
			if interactive: clear()
			
			cell_width = int((tp.get_terminal_size()[0]-self.PL_len-2)/self.PL_len)
			
			if i in skip_list:
				continue

			if len(peaks_dict[index]) <= self.PL_len:
				
				peaks = peaks_dict[index]
				
				if interactive:							
					for row in cluster_dict[index]:
						for cell in row:
							sys.stdout.write("|" + cell[:cell_width] + " "*(cell_width-len(cell)))
						sys.stdout.write("|\n")
					print("#"*(tp.get_terminal_size()[0]-1))
							
				try:
					_next = index_list[i+1]
					if list(numpy.setdiff1d(peaks_dict[_next],peaks)) == peaks_dict[_next]:	
						if interactive:
							for row in cluster_dict[_next]:
								for cell in row:
									sys.stdout.write("|" + cell[:cell_width] + " "*(cell_width-len(cell)))
								sys.stdout.write("|\n")
							print("#"*(tp.get_terminal_size()[0]-1))
					
						peaks = list(set(peaks+peaks_dict[_next])) #Remove duplicates
					else:
						_next = False
				except IndexError:
					_next = False
				
				if _next:
					try:
						_next_1 = index_list[i+2]
						if list(numpy.setdiff1d(peaks_dict[_next_1],peaks)) == peaks_dict[_next_1]:	
							if interactive:
								for row in cluster_dict[_next_1]:
									for cell in row:
										sys.stdout.write("|" + cell[:cell_width] + " "*(cell_width-len(cell)))
									sys.stdout.write("|\n")
								print("#"*(tp.get_terminal_size()[0]-1))
								
							peaks = list(set(peaks+peaks_dict[_next_1])) #Remove duplicates
						else:
							_next_1 = False
					except IndexError:
						_next_1 = False
				else:
					_next_1 = False
				
				if _next_1:
					try:					
						_next_2 = index_list[i+3]
						if list(numpy.setdiff1d(peaks_dict[_next_2],peaks)) == peaks_dict[_next_2]:	
							if interactive:
								for row in cluster_dict[_next_2]:
									for cell in row:
										sys.stdout.write("|" + cell[:cell_width] + " "*(cell_width-len(cell)))
									sys.stdout.write("|\n")
								print("#"*(tp.get_terminal_size()[0]-1))
								
							peaks = list(set(peaks+peaks_dict[_next_2])) #Remove duplicates
						else:
							_next_2 = False
					except IndexError:
						_next_2 = False
				else:
					_next_2 = False
				
				if len(peaks) == self.PL_len:
					
					tmp = cluster_dict[index][0]
								
					if _next:
						tmp += cluster_dict[_next][0]
					if _next_1:
						tmp += cluster_dict[_next_1][0]
					if _next_2:
						tmp += cluster_dict[_next_2][0]
					
					tmp = list(set(filter(None, tmp)))
					tmp.sort()
						
					if within6sec(tmp):
						if interactive:
							while True:	
								inp_resp = raw_input("options >")
								if inp_resp.lower().startswith("del"):
									print("delete")
									break
								elif inp_resp.lower().startswith("mer"):
									write_output()
									break
						
						elif not interactive:
							write_output()		

	def open_lo(self,input_file):
		print("In LibreOffice, align the rows that correspond to the same peak, and delete any rows corresponding to bleed or which do not have enough peaks.")
		os.system('""' + self.LO_Path + '"" -o ' + input_file)
		while True:
			n = raw_input("Once complete, close LibreOffice and type DONE. >")
			if n == "DONE":
				break		
		while True:
			n = raw_input("Did you leave any rows with missing peaks? >")	
			if n.lower() == "no":
				break
			elif n.lower() == "yes":
				print("Go back and correct it then!") #Sassy
				os.system('""' + LO_Path + '"" -o ' + input_file)
				while True:
					m = raw_input("Once complete, close LibreOffice and type DONE. >")
					if m == "DONE":
						break	
				

	
	
	
	"""Counter and Formating Functions"""	
	def Match_Counter(self, seperator=";"):	
		# Also generates final output

		if self.PL_len == 1:
			print("Require two or more samples to combine. Check ./list.txt or --samples and try again.")
			print("")
			sys.exit(print(__doc__))
		
		input_file_name = self.lot_name + "_COMBINED.CSV"
		
		"""write column headings to final output"""
		with open(os.path.join(self.OUTPUT_DIRECTORY, self.lot_name + "_FINAL.csv"),"w") as final_csv_output:
			final_csv_output.write(self.lot_name+";;;")
			for num in range(self.PL_len):
				final_csv_output.write(self.prefixList[num].rstrip("\n\r ") + ";;;")
			final_csv_output.write("\nRetention Time;Peak Area;;" + "Page;RT;Area;"*self.PL_len + ";Match Factor;;;;Reverse Match Factor;;;;Hit Number;;;;Retention Time;;;;Peak Area;;;"+"\n") 
			final_csv_output.write("Name;CAS Number;Freq.;" + "Hit No.;Match;R Match;"*self.PL_len + ";" + "Mean;STDEV;%RSD;;"*5 + "\n")
			final_csv_output.close()	
		
		f = open("./"+input_file_name)		#Open merged output of lot of samples
#		if openLO:
#			seperator = ","
#		elif args.jigsaw:
#			seperator = ";"
#		seperator = ";" # fix
		csv_f = csv.reader(f, delimiter=seperator)

		csv_input = []
		
		for row in csv_f:
			csv_input.append(row)

		for i in range(len(csv_input)):
			rows_appended = []
			match_list = []
			try:
				"""From https://stackoverflow.com/questions/4843158/check-if-a-python-list-item-contains-a-string-inside-another-string"""
				if any("Page" in s for s in csv_input[i]): #csv_input[i][][:4] == "Page" or :
					gc_data = ['','','']
					RT_list = []
					area_list = []
					try:
						for j in range(self.PL_len):
							gc_data.append(csv_input[i][2+(j*9)][5:])
							gc_data.append(csv_input[i][0+(j*9)]) #get gc peak areas and retention times for final output
							gc_data.append(csv_input[i][1+(j*9)])
							if str(csv_input[i][0+(j*9)]).lower() != "no peak":
								RT_list.append(float(csv_input[i][0+(j*9)]))
								area_list.append(float(csv_input[i][1+(j*9)]))

					except IndexError:
						print("GC")
					
					gc_data.append(";;;;;;;;;;;;")
					gc_data.append(str(numpy.mean(RT_list))) # Retention Time Average
					gc_data.append(str(numpy.std(RT_list, ddof=1))) # Retention Time STDEV	
					gc_data.append(str(numpy.std(RT_list, ddof=1)/numpy.mean(RT_list))) # Retention Time %RSD
					gc_data.append("")
					gc_data.append(str(numpy.mean(area_list))) # Peak Area Average
					gc_data.append(str(numpy.std(area_list, ddof=1))) # Peak Area STDEV
					gc_data.append(str(numpy.std(area_list, ddof=1)/numpy.mean(area_list))) # Peak Area %RSD	
					
					gc_data[0] = str(rounders(numpy.mean(RT_list),'0.001'))
					gc_data[1] = str(rounders(numpy.mean(area_list),'0.1'))
					
					curr_peak = []
					#six_rows = {"Header": csv_input[i], "Hit 1":csv_input[i+1],"Hit 2":csv_input[i+2],"Hit 3":csv_input[i+3],"Hit 4":csv_input[i+4],"Hit 5":csv_input[i+5]}
					try:
						rows_appended = csv_input[i+1] + csv_input[i+2] + csv_input[i+3] + csv_input[i+4] + csv_input[i+5]
					except IndexError:
						print("#")
					
					"""From https://stackoverflow.com/questions/2600191/how-to-count-the-occurrences-of-a-list-item"""
					count_dict = Counter(rows_appended)
					try:
						for j in range(self.PL_len):
							if len(csv_input[i+1][6+9*j]) > 0:
								csv_input[i+1][1+9*j] = str(count_dict[csv_input[i+1][6+9*j]])
							if len(csv_input[i+2][6+9*j]) > 0:
								csv_input[i+2][1+9*j] = str(count_dict[csv_input[i+2][6+9*j]])
							if len(csv_input[i+3][6+9*j]) > 0:
								csv_input[i+3][1+9*j] = str(count_dict[csv_input[i+3][6+9*j]])
							if len(csv_input[i+4][6+9*j]) > 0:
								csv_input[i+4][1+9*j] = str(count_dict[csv_input[i+4][6+9*j]])
							if len(csv_input[i+5][6+9*j]) > 0:
								csv_input[i+5][1+9*j] = str(count_dict[csv_input[i+5][6+9*j]])
							
							for k in range(1,6):
								match_list.append([csv_input[i+k][6+9*j],csv_input[i+k][7+9*j],csv_input[i+k][1+9*j]] + ["","",""]*(j) + [csv_input[i+k][2+9*j],csv_input[i+k][4+9*j],csv_input[i+k][5+9*j]] + ["","",""]*(self.PL_len-1-j))
						
					except IndexError:
						print("\r",end='')
									
					"""prepare final output"""
					from operator import itemgetter
					sorted_match_list = sorted(match_list, key=itemgetter(1))
					merged_hit_list = []
					
					while len(sorted_match_list)>1:
						hit_name = sorted_match_list[0][0]
						merged_hit = sorted_match_list[0]

						if sorted_match_list[1][0] == merged_hit[0]:
							for index,val in enumerate(sorted_match_list[1]):
								if merged_hit[index] == '':
									merged_hit[index] = val
						
							del sorted_match_list[1]
							
						else:
							MF_list = [] # List of Match Factors
							RMF_list = [] # List of Reverse Match Factors
							hit_num_list = [] # List of hit numbers
							for x in range(self.PL_len):
								if merged_hit[4+(x*3)] != '':
									MF_list.append(int(merged_hit[4+(x*3)]))
								if merged_hit[5+(x*3)] != '':
									RMF_list.append(int(merged_hit[5+(x*3)]))
								if merged_hit[5+(x*3)] != '':
									hit_num_list.append(int(merged_hit[3+(x*3)]))
							with warnings.catch_warnings():
								if not verbose:
									warnings.simplefilter("ignore")						
								if len(MF_list) > 0:
									merged_hit.append("")
									merged_hit.append(str(numpy.mean(MF_list)))	# Match Factor Average
									merged_hit.append(str(numpy.std(MF_list, ddof=1)))	# Match Factor STDEV # error occurs here
									merged_hit.append(str(numpy.std(MF_list, ddof=1)/numpy.mean(MF_list)))	# Match Factor %RSD
								if len(RMF_list) > 0:
									merged_hit.append("")
									merged_hit.append(str(numpy.mean(RMF_list)))	# Reverse Match Factor Average
									merged_hit.append(str(numpy.std(RMF_list, ddof=1)))	# Reverse Match Factor STDEV
									merged_hit.append(str(numpy.std(RMF_list, ddof=1)/numpy.mean(RMF_list)))	# Reverse Match Factor %RSD
								if len(hit_num_list) > 0:
									merged_hit.append("")
									merged_hit.append(str(numpy.mean(hit_num_list)))	# Hit Average
									merged_hit.append(str(numpy.std(hit_num_list, ddof=1)))	# Hit STDEV
									merged_hit.append(str(numpy.std(hit_num_list, ddof=1)/numpy.mean(hit_num_list)))	# Hit %RSD
							merged_hit.append("")
							merged_hit_list.append(merged_hit)
							del sorted_match_list[0]
							
					# End of While
					
					merged_hit = sorted_match_list[0]
					MF_list = [] # List of Match Factors
					RMF_list = [] # List of Reverse Match Factors
					hit_num_list = [] # List of hit numbers
					for x in range(self.PL_len):
						if merged_hit[4+(x*3)] != '':
							MF_list.append(int(merged_hit[4+(x*3)]))
						if merged_hit[5+(x*3)] != '':
							RMF_list.append(int(merged_hit[5+(x*3)]))
						if merged_hit[5+(x*3)] != '':
							hit_num_list.append(int(merged_hit[3+(x*3)]))

					if len(MF_list) > 0:
						merged_hit.append("")
						merged_hit.append(str(numpy.mean(MF_list)))	# Match Factor Average
						merged_hit.append(str(numpy.std(MF_list, ddof=1)))	# Match Factor STDEV
						merged_hit.append(str(numpy.std(MF_list, ddof=1)/numpy.mean(MF_list)))	# Match Factor %RSD
					if len(RMF_list) > 0:
						merged_hit.append("")
						merged_hit.append(str(numpy.mean(RMF_list)))	# Reverse Match Factor Average
						merged_hit.append(str(numpy.std(RMF_list, ddof=1)))	# Reverse Match Factor STDEV
						merged_hit.append(str(numpy.std(RMF_list, ddof=1)/numpy.mean(RMF_list)))	# Reverse Match Factor %RSD
					if len(hit_num_list) > 0:
						merged_hit.append("")
						merged_hit.append(str(numpy.mean(hit_num_list)))	# Hit Average
						merged_hit.append(str(numpy.std(hit_num_list, ddof=1)))	# Hit STDEV
						merged_hit.append(str(numpy.std(hit_num_list, ddof=1)/numpy.mean(hit_num_list)))	# Hit %RSD
					merged_hit.append("")
					merged_hit_list.append(merged_hit)
					
					del sorted_match_list[0]
							
					for index, row in enumerate(merged_hit_list):
						if len(row) != 25:
							if row.count('') == len(row):
								del merged_hit_list[index]
					
					hit_output = sorted(list(merged_hit_list), key=itemgetter(2,13))
					hit_output.reverse()
					
					"""write final output"""
					with open(os.path.join(self.OUTPUT_DIRECTORY, self.lot_name + "_FINAL.csv"),"a") as final_csv_output:

						final_csv_output.write(';'.join(gc_data))
						final_csv_output.write('\n')

						for index, row in enumerate(hit_output):
							if (index < 5) and (row[2] != '1'):
							#if (row[2] != '1'):
								final_csv_output.write(';'.join(row))
								final_csv_output.write(';\n')
						final_csv_output.write('\n')
					
			except TypeError:
				traceback.print_exc()	#print the error
				continue	
			
		"""write counter output"""
		single_list = []
			
		for i in range(len(csv_input)):
			single_list.append((";".join(csv_input[i])))#[:-1])
		
		with open(self.lot_name + "_COUNTER.CSV","w") as f:
			joinedOutput = '\n'.join(single_list)

			f.write(str(joinedOutput))
		
		f.close()
		time.sleep(3)
		
		"""Convert to XLSX and format"""
		append_to_xlsx(self.lot_name + "_COUNTER.CSV", os.path.join(self.RESULTS_DIRECTORY,self.lot_name + "_FINAL.xlsx"), "GC-MS", overwrite=True, seperator=";"),""
		self.formatXLSX(os.path.join(self.RESULTS_DIRECTORY,self.lot_name + "_FINAL.xlsx"))
		append_to_xlsx(os.path.join(self.OUTPUT_DIRECTORY, self.lot_name + "_FINAL.csv"),os.path.join(self.RESULTS_DIRECTORY,self.lot_name + "_FINAL.xlsx"), "Matches",";"),""
		self.finalXLSX(os.path.join(self.RESULTS_DIRECTORY,self.lot_name + "_FINAL.xlsx"), os.path.join(self.OUTPUT_DIRECTORY, self.lot_name + "_FINAL.csv"))
			
		return 0

	def formatXLSX(self, inputFile):	# Formatting for Combined Output
		"""From http://openpyxl.readthedocs.io/en/default/"""
		wb = load_workbook(inputFile)
		
		ws = wb.active # grab the active worksheet

		"""From https://stackoverflow.com/questions/38619471/iterate-through-all-rows-in-specific-column-openpyxl
		and https://stackoverflow.com/questions/14463277/how-to-disable-python-warnings"""

		with warnings.catch_warnings():
			warnings.simplefilter("ignore")
			
			for sample_offset in range (0,11):
				for column_index in [0,1,3,4,5,6,8]:
					#column_name = col_list[column_index + (sample_offset*9)]
					column_name = get_column_letter(column_index + (sample_offset*9) + 1)
				
					col_string = column_name + '{}:' + column_name + '{}'
			
					for row in ws.iter_rows(col_string.format(ws.min_row,ws.max_row)):
						for cell in row:
							cell.alignment = Alignment(horizontal='center',
											vertical='center',
											wrap_text=True)
		
				width_list = [9,11,2.5,0.63,6,6,45,12,6,'']
				
				for column_index2 in range(0,9):
					#column_name2 = col_list[column_index2 + (sample_offset*9)]
					column_name2 = get_column_letter(column_index2 + (sample_offset*9) + 1)
					
					col_width = width_list[column_index2]
					ws.column_dimensions[column_name2].width = col_width
					col_string2 = column_name2 + '{}:' + column_name2 + '{}'
					if column_index2 == 3:
						for row in ws.iter_rows(col_string2.format(ws.min_row,ws.max_row)):
							for cell in row:
								cell.value = None
						ws.column_dimensions[column_name2].hidden = True
					if column_index2 == 0:
						for row in ws.iter_rows(col_string2.format(ws.min_row,ws.max_row)):
							for cell in row:
								try:
									float(cell.internal_value)
									cell.value = float(cell.internal_value)
									cell.number_format = "0.000"
								except:
									cell.value = cell.internal_value
						#ws.merge_cells(column_name2 + '1:' + (col_list[(column_index2 + 8)+(sample_offset*9)]) + '1')
						ws.merge_cells(column_name2 + '1:' + (get_column_letter((column_index2 + 8)+(sample_offset*9) + 1)) + '1')	
						
					if column_index2 in [2,4,5]:
						for row in ws.iter_rows(col_string2.format(ws.min_row,ws.max_row)):
							for cell in row:
								try:
									float(cell.internal_value)
									cell.value = float(cell.internal_value)
									cell.number_format = "0"
								except:
									cell.value = cell.internal_value
				
					if column_index2 == 7:
						for row in ws.iter_rows(col_string2.format(ws.min_row,ws.max_row)):
							for cell in row:
								cell.alignment = Alignment(horizontal='center',
												vertical='center')
		# Save the file
		wb.save(inputFile)

	def finalXLSX(self, inputFile, csv_path):	# Formatting for Final Output
		widgets = [progressbar.widgets.AnimatedMarker(), ' ','Generating XLSX Output.',]
		bar = progressbar.ProgressBar(widgets=widgets, max_value=progressbar.UnknownLength).start()
		bar_step = 1	
		
		"""From http://openpyxl.readthedocs.io/en/default/"""
		wb = load_workbook(inputFile)
		
		# grab the active worksheet
		ws = wb["Matches"]
		
		"""From https://stackoverflow.com/questions/38619471/iterate-through-all-rows-in-specific-column-openpyxl
		and https://stackoverflow.com/questions/14463277/how-to-disable-python-warnings"""

		with warnings.catch_warnings():
			if not verbose:
				warnings.simplefilter("ignore")
			
			width_list = [45,14,5] + [8,8,14]*self.PL_len + [1.5,9,7.5,9]*4 + [1.5,10.5,12,5.8]
				
			for column_index, width in enumerate(width_list):
				column_name = get_column_letter(column_index+1)
				col_width = width_list[column_index]
				ws.column_dimensions[column_name].width = float(col_width)
				
				for row in ws.iter_rows('{0}{1}:{0}{2}'.format(column_name,ws.min_row,ws.max_row)):
					for index, cell in enumerate(row):
						if column_index > 1:
							cell.alignment = Alignment(horizontal='center',
													vertical='center',
													wrap_text=True)
						elif column_index == 1:
							cell.alignment = Alignment(horizontal='center')
						else:
							cell.alignment = Alignment(wrap_text=True)
						try:
							float(cell.internal_value)
							cell.value = float(cell.internal_value)
							if column_index > 12:
								if column_index in [2,3]:
									cell.number_format = "0"
								elif width == 5.5:
									cell.number_format = "0.0"
								elif width == 9:
									cell.number_format = "0.000%"
								else:
									cell.number_format = "0.000"
						except:
							cell.value = cell.internal_value

				for x in range(self.PL_len):
					if column_index == 3+(3*x):
						#ws.merge_cells(column_name + '1:' + (col_list[(column_index + 2)]) + '1')
						ws.merge_cells(column_name + '1:' + (get_column_letter(column_index + 3)) + '1')
				for x in range(5):
					if column_index == (3*self.PL_len)+4+(4*x):
						#ws.merge_cells(column_name + '2:' + (col_list[(column_index + 2)]) + '2')	
						ws.merge_cells(column_name + '2:' + (get_column_letter(column_index + 3)) + '2')	
				bar.update(bar_step); bar_step += 1	
				
			ws.cell(row=2,column=1).alignment = Alignment(horizontal='right') # Format cell A2
			
			"""Statistics Sheet"""
			ws2 = wb.create_sheet("Statistics_Full")
			f = open(csv_path)
			csv_input_raw = csv.reader(f, delimiter=";")
			csv_input = []
			for row in csv_input_raw:
				csv_input.append(row)
			
			f.close()
			
			xls_row_num = 3
			col_indices = [0,1] + [x for x in range (-22,0)]
			for row_index, row in enumerate(csv_input):
				if row_index == 0:
					ws2.cell(row=1,column=1).value = row[0] # Cell A1
				if row_index == 2:
					ws2.cell(row=2,column=1).value = "Name" # Cell A2
					ws2.cell(row=2,column=2).value = "CAS Number" # Cell B2
					#ws2.cell(row=2,column=3).value = "Confidence" # Cell C2
					for index1, index2 in enumerate([x for x in range (-22,0)]):
						ws2.cell(column=(index1+4), row=2).value = row[index2]
				
				ws2.cell(row=1,column=13).value = "Match Factor" # Cell M1
				ws2.cell(row=1,column=17).value = "Reverse Match Factor" # Cell Q1
				ws2.cell(row=1,column=21).value = "Hit Number" # Cell U1
				ws2.cell(row=1,column=25).value = "MS Comparison" # Cell Y1
				ws2.cell(row=2,column=25).value = "Mean" # Cell Y2
				ws2.cell(row=2,column=26).value = "STDEV" # Cell Z2
				ws2.cell(row=2,column=27).value = "%RSD" # Cell AA2
				ws2.cell(row=1,column=5).value = "Retention Time" # Cell E1
				ws2.cell(row=1,column=9).value = "Peak Area" # Cell I1	
				
				"""From https://stackoverflow.com/questions/4843158/check-if-a-python-list-item-contains-a-string-inside-another-string"""
				if any(" of " in s for s in row): 
					try:
						#check if the name is a "similar compound", and if so override
						if csv_input[row_index+1][0] == "Oxamide, N,N'-diethyl-N,N'-diphenyl-":
							ws2.cell(column=1, row=(xls_row_num)).value = "N,N'-Diethyl-N,N'-diphenylurea" # Column A
							ws2.cell(column=2, row=(xls_row_num)).value = "85-98-3" # Column B
						else:
							ws2.cell(column=1, row=(xls_row_num)).value = csv_input[row_index+1][0] # Name, Column A
							ws2.cell(column=2, row=(xls_row_num)).value = csv_input[row_index+1][1] # CAS, Column B
						
						ws2.cell(column=3, row=(xls_row_num)).value = csv_input[row_index+1][2] # Frequency, Column C

						ws2.cell(column=5, row=(xls_row_num)).value = csv_input[row_index][-7] # RT and Area Data, Column E
						ws2.cell(column=6, row=(xls_row_num)).value = csv_input[row_index][-6] # Column F
						ws2.cell(column=7, row=(xls_row_num)).value = csv_input[row_index][-5] # Column G
						ws2.cell(column=8, row=(xls_row_num)).value = csv_input[row_index][-4] # Column H
						ws2.cell(column=9, row=(xls_row_num)).value = csv_input[row_index][-3] # Column I
						ws2.cell(column=10, row=(xls_row_num)).value = csv_input[row_index][-2] # Column J
						ws2.cell(column=11, row=(xls_row_num)).value = csv_input[row_index][-1] # Column K
	 
						ws2.cell(column=13, row=(xls_row_num)).value = csv_input[row_index+1][-13] # Match Data, Column M
						ws2.cell(column=14, row=(xls_row_num)).value = csv_input[row_index+1][-12] # Column N
						ws2.cell(column=15, row=(xls_row_num)).value = csv_input[row_index+1][-11] # Column O
						ws2.cell(column=16, row=(xls_row_num)).value = csv_input[row_index+1][-10] # Column P
						ws2.cell(column=17, row=(xls_row_num)).value = csv_input[row_index+1][-9] # Column Q
						ws2.cell(column=18, row=(xls_row_num)).value = csv_input[row_index+1][-8] # Column R
						ws2.cell(column=19, row=(xls_row_num)).value = csv_input[row_index+1][-7] # Column S
						ws2.cell(column=20, row=(xls_row_num)).value = csv_input[row_index+1][-6] # Column T
						ws2.cell(column=21, row=(xls_row_num)).value = csv_input[row_index+1][-5] # Column U
						ws2.cell(column=22, row=(xls_row_num)).value = csv_input[row_index+1][-4] # Column V
						ws2.cell(column=23, row=(xls_row_num)).value = csv_input[row_index+1][-3] # Column W
						ws2.cell(column=24, row=(xls_row_num)).value = csv_input[row_index+1][-2] # Column X
						ws2.cell(column=25, row=(xls_row_num)).value = csv_input[row_index+1][-1] # Column Y
						
						for prefix in range(self.PL_len):
							ws2.cell(column=(29+(prefix)), row=(xls_row_num)).value = float(csv_input[row_index][5+(3*prefix)])
							
						
						xls_row_num = xls_row_num + 1
					except IndexError:
						print("\r",end='')

				bar.update(bar_step); bar_step += 1

			for prefix in range(self.PL_len):
				ws2.column_dimensions[get_column_letter(29+(prefix))].hidden= True
			
			bar.update(bar_step); bar_step += 1
			
			wb.save(inputFile)
			
			"""Mass Spectra comparison for each retention time"""
			ws30 = wb["GC-MS"]
			rt_dict = {}
			for index, prefix in enumerate(self.prefixList):
				rt_list2 = []
				#for row in ws30.iter_rows("{0}{1}:{0}{2}".format(col_list[0+(9*index)],ws30.min_row,ws30.max_row)):
				for row in ws30.iter_rows("{0}{1}:{0}{2}".format(get_column_letter(1+(9*index)),ws30.min_row,ws30.max_row)):
					for cell in row:
						if RepresentsInt(cell.value):
							rt_list2.append(float(cell.value))

				rt_dict[prefix] = rt_list2

				bar.update(bar_step); bar_step += 1
					
			# average_list = {}
			# for i in range(len(rt_list2)):
				# verbosePrint(i)
				# list_for_mean = []
				# for item in rt_dict.values():
					# verbosePrint(item)
					# verbosePrint(item[i])
					# list_for_mean.append(item[i])
				# average_list[i+1] = str(rounders(str(numpy.mean(list_for_mean)),"0.000"))
				# bar.update(bar_step); bar_step += 1

			# bar.finish()
			
			# Version 2
			average_list = {}
			i=1
			ws50 = wb["Statistics_Full"]
			for row in ws50.iter_rows("E3:E{}".format(ws50.max_row)):
				for cell in row:
					average_list[i] = str(rounders(cell.value,"0.000"))
					bar.update(bar_step); bar_step += 1		
					i = i+1
			bar.finish()
			
			#Actual Comparison	
			self.get_mass_spectra()
			print("\n")
			if len(average_list) == 0:
				print("No peaks were found in common.")
				sys.exit(1)
			self.generate_mass_spectra(average_list)
			self.generate_nist_library()
			MS_comparison_data = self.nist_ms_comparison_2(average_list)
			verbosePrint(MS_comparison_data)
			
			#Write Comparison to end of "Statstics" sheets
			for sheet in ["Statistics_Full"]:#, "Statistics - Lit Only"]:
				ws31 = wb[sheet]
				for row in ws31.iter_rows("E3:E{}".format(ws31.max_row)):
					for cell in row:
						RT = str(rounders(cell.value,"0.000"))
						verbosePrint(RT)
						if RT in MS_comparison_data:
							format_dict = {"Y":[0,"0.0"],"Z":[1,"0.000"],"AA":[2,"0.000%"]}
							for col in format_dict:
								ws31.cell(column=column_index_from_string(col), row=cell.row).value = MS_comparison_data[RT][format_dict[col][0]]
								ws31.cell(column=column_index_from_string(col), row=cell.row).number_format = format_dict[col][1]
								verbosePrint(MS_comparison_data[RT][format_dict[col][0]])

			stat_format(ws2)
			
			wb.save(inputFile)
			
			"""Remove peaks that do not match between samples"""
			ws35 = wb.create_sheet("Statistics")
			ws35_row = 3
			
			for row in ws2.iter_rows('A1:{}{}'.format(get_column_letter(ws2.max_column),ws2.max_row)):
				try:	
					if row[0].row == 1:
						for cell in row:
							ws35.cell(column = cell.col_idx, row=1).value = ws2.cell(column=cell.col_idx, row=1).internal_value
							
					elif row[0].row == 2:
						for cell in row:
							ws35.cell(column = cell.col_idx, row=2).value = ws2.cell(column = cell.col_idx, row=2).internal_value
						
					elif float(row[24].value) >= 650 and int(row[2].value) > (self.PL_len/2):
						for cell in row:
							ws35.cell(column = cell.col_idx, row=(ws35_row)).value = cell.internal_value
						ws35_row = ws35_row+1
						
				except (AttributeError, TypeError, ValueError) as e:
					continue
					# Save the file
				#	wb.save(inputFile)

			ws2 = wb["Statistics"]
			wb.save(inputFile)
			stat_format(ws2)
			
			wb.save(inputFile)
			
			"""Get list of CAS Numbers for compounds reported in literature"""
			with open("./lib/CAS.txt","r") as f:
				CAS_list = f.readlines()
			for index, CAS in enumerate(CAS_list):
				CAS_list[index] = CAS.rstrip("\r\n")
				
			"""Produce Statistics for compounds reported in literature"""
			ws4 = wb.create_sheet("Statistics - Lit Only")
			ws4_row = 3
			
			
			for row in ws2.iter_rows('A1:{}{}'.format(get_column_letter(ws2.max_column),ws2.max_row)):
				try:	
					if row[0].row == 1:
						for cell in row:
							ws4.cell(column = cell.col_idx, row=1).value = ws2.cell(column = cell.col_idx, row=1).internal_value
					
					elif row[0].row == 2:
						for cell in row:
							ws4.cell(column = cell.col_idx, row=2).value = ws2.cell(column = cell.col_idx, row=2).internal_value
					
					elif row[1].value.replace("-","") in CAS_list: 
						for cell in row:
							ws4.cell(column = cell.col_idx, row=(ws4_row)).value = cell.internal_value
						ws4_row = ws4_row+1
				except AttributeError:
					continue
				bar.update(bar_step); bar_step += 1
				
			stat_format(ws4)
			bar.update(bar_step); bar_step += 1
			
			def mean_pa_graph(wsX, title, radar=False):
				"""Graph of Mean Peak Area"""
				ws3 = wb.create_sheet(title + " Data")
				for row in wsX.iter_rows('A{}:A{}'.format(wsX.min_row,wsX.max_row)):
					for cell in row:
						ws3.cell(column = 1, row=(cell.row)).value = cell.internal_value
				for row in wsX.iter_rows('I{}:I{}'.format(wsX.min_row,wsX.max_row)):
					for cell in row:
						ws3.cell(column = 2, row=(cell.row)).value = cell.internal_value
						try:
							ws3.cell(column = 3, row=(cell.row)).value = math.log10(cell.internal_value)
						except TypeError:
							print("LogError {} {}".format(cell.internal_value,str(cell.row)))
							ws3.cell(column = 3, row=(cell.row)).value = cell.internal_value
						ws3.cell(column=3, row=2).value = "Log(Mean)"
				
				width_list3 = [60,13,13]
					
				for column_index, width in enumerate(width_list3):
					column_name = get_column_letter(column_index+1)
					col_width = width_list3[column_index]
					ws3.column_dimensions[column_name].width = float(col_width)
				
				ws3.sheet_state = "hidden"
				
				cs = wb.create_chartsheet(title.replace("PA","Peak Area"))
				
				chart1 = BarChart()
				chart1.type = "col"
				chart1.grouping = "percentStacked"
				chart1.overlap = 100
				chart1.title = 'Mean Peak Area'
				chart1.y_axis.title = 'Peak Area'
				chart1.x_axis.title = ''

				data = Reference(ws3, min_col=2, min_row=1, max_row=ws3.max_row, max_col=3)
				cats = Reference(ws3, min_col=1, min_row=2, max_row=ws3.max_row)
				chart1.add_data(data, titles_from_data=True)
				chart1.set_categories(cats)
				chart1.shape = 4
				cs.add_chart(chart1)
				
				if radar:
					"""Radar Chart"""
					cs = wb.create_chartsheet("Radar Log(PA)")
					chart1 = RadarChart()
					chart1.type = "marker"
					chart1.title = 'Log(Peak Area)'
					chart1.legend = None

					data = Reference(ws3, min_col=3, min_row=2, max_row=ws3.max_row, max_col=3)
					labels = Reference(ws3, min_col=1, min_row=3, max_row=ws3.max_row)
					chart1.add_data(data, titles_from_data=True)
					chart1.set_categories(labels)
					cs.add_chart(chart1)
			
			mean_pa_graph(ws2, "Mean PA") # All Compounds
			bar.update(bar_step); bar_step += 1
			mean_pa_graph(ws4, "Lit Mean PA", True) # Literature only
			bar.update(bar_step); bar_step += 1
			
			"""Graph of all Peak Areas"""
			index_list = []
			ws20 = wb["Statistics"]
			
			wb.save(inputFile)
			
			for sheet in ["All PA", "All Log(PA)"]:
				ws5 = wb.create_sheet(sheet + " Data")

				for row_index, row in enumerate(ws20.iter_rows('A{}:B{}'.format(ws20.min_row,ws20.max_row))):
					if row_index > 1:
						for cell in row:
							if cell.column == "A":
								ws5.cell(column = 2, row=(row_index)).value = cell.internal_value
							elif cell.column == "B":
								ws5.cell(column = 1, row=(row_index)).value = cell.internal_value
				
				for index, prefix in enumerate(self.prefixList):
					source_column = get_column_letter(29+index)
					destination_column = 3+index
					for row_index, row in enumerate(ws20.iter_rows('{0}{1}:{0}{2}'.format(source_column,ws20.min_row,ws20.max_row))):
						if row_index > 1:
							for cell in row:
								if cell.value not in [None,'']:
									if sheet == "All PA":
										ws5.cell(column=destination_column, row=row_index).value = float(cell.value)
									elif sheet == "All Log(PA)":
										ws5.cell(column=destination_column, row=row_index).value = math.log10(float(cell.value))
					
					ws5.cell(column=destination_column, row=1).value = prefix.lower().replace("_"," ").replace("subtract","").replace("  "," ").capitalize()
				
				ws5.cell(column = 1, row=1).value = "CAS Number" # Cell A1
				ws5.cell(column = 2, row=1).value = "Name" # Cell B1

				ws5.sheet_state = "hidden"  # re-enable once finished

				"""Graph Data for compounds reported in literature"""
				ws5 = wb['All PA Data']
				ws6 = wb.create_sheet("Lit " + sheet + " Data")
				ws6.cell(column = 1, row=1).value = "CAS Number" # Cell A1
				ws6.cell(column = 2, row=1).value = "Name" # Cell B1
				for x in range(self.PL_len):

					ws6.cell(column=(x+2), row=1).value = self.prefixList[x].lower().replace("_"," ").replace("subtract","").replace("  "," ").capitalize()
					
					ws6_row = 2
					
					for row in ws5.iter_rows('A1:{}{}'.format(get_column_letter(ws5.max_column),ws5.max_row)): #####
						try:	
							if row[0].row == 1:
								for cell in row:
									ws6.cell(column=cell.col_idx, row=1).value = ws5.cell(column=cell.col_idx, row=1).internal_value
						
							elif row[0].value.replace("-","") in CAS_list: 
								for cell in row:
									ws6.cell(column=cell.col_idx, row=(ws6_row)).value = cell.value
								ws6_row = ws6_row+1
						except AttributeError:
							next
					bar.update(bar_step); bar_step += 1
										
				if sheet == "All PA":
					width_list5 = [12,60,12,12,12]
				elif sheet == "All Log(PA)":
					width_list5 = [12,60,9,9,9]
					
				def format_and_chart(wsX, title):
					for column_index, width in enumerate(width_list5):
						column_name = get_column_letter(column_index+1)
						col_width = width_list5[column_index]
						wsX.column_dimensions[column_name].width = float(col_width)
						for row in wsX.iter_rows('{0}{1}:{0}{2}'.format(column_name,wsX.min_row,wsX.max_row)):
							for index, cell in enumerate(row):
								if col_width == 9:
									cell.number_format = "0.000"
						for x in range(self.PL_len):
							wsX.cell(column=(x+3), row=1).alignment = Alignment(text_rotation="45")
					wsX.sheet_state="hidden"	#re-enable once finished
				
					if "Log" in title:
						cs3 = wb.create_chartsheet(title)
					else:
						cs3 = wb.create_chartsheet(title.replace("PA", "Peak Area"))
					chart3= BarChart()
					chart3.type = "col"
					chart3.grouping = "percentStacked"
					chart3.overlap = 100
					#chart3.title = 'Mean Peak Area'
					chart3.title = title
					chart3.y_axis.title = 'Peak Area'
					chart3.x_axis.title = ''

					data = Reference(wsX, min_col=3, min_row=1, max_row=wsX.max_row, max_col=(self.PL_len+2))
					cats = Reference(wsX, min_col=2, min_row=2, max_row=wsX.max_row)
					chart3.add_data(data, titles_from_data=True)
					chart3.set_categories(cats)
					cs3.add_chart(chart3)
					
					wb.save(inputFile)
					
					"""Line Chart"""
					if "Log" in title and "Lit" in title:
					#if "Lit" in title:
						values = []
						for row in wsX.iter_rows('C2:{0}{1}'.format(get_column_letter(wsX.max_column),wsX.max_row)): #####
							for cell in row:
								#if type(cell.value) != str:
									values.append(cell.value)
						
						values.sort()
						values = filter(None, values)
						
						cs3 = wb.create_chartsheet(title + " Line")
						
						c1 = LineChart()
						c1.title = "Peak Area"
						if "Log" in title:
							c1.y_axis.title = 'Log(Peak Area)'
						else:
							c1.y_axis.title = 'Peak Area'
						c1.x_axis.title = ''
						if "Log" not in title:
							c1.y_axis.scaling.logBase = 10

						data = Reference(wsX, min_col=3, min_row=1, max_row=wsX.max_row, max_col=(self.PL_len+2))
						cats = Reference(wsX, min_col=2, min_row=2, max_row=wsX.max_row)
						c1.add_data(data, titles_from_data=True)
						c1.set_categories(cats)

						if "Log" in title:
							c1.y_axis.scaling.min = int(min(values))
							c1.y_axis.scaling.max = int(max(values))+1
						else:
							c1.y_axis.scaling.min = int(min(values))-1000
							c1.y_axis.scaling.max = int(max(values))+1000
						
						
						from openpyxl.chart.text import RichText
						from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font    
						
			#			#chart title
			#			cp = CharacterProperties(latin=Font(typeface='Calibri'), sz=1400)
			#			c1.title.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
			#			c1.title.tx.rich.p[0].r.rPr = cp
						
			#			#y-axis title
			#			cp = CharacterProperties(latin=Font(typeface='Calibri'), sz=1100)
			#			c1.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
			#			c1.title.tx.rich.p[0].r.rPr = cp
						
						cs3.add_chart(c1)
						#cs3.sheet_view.zoomScale = 80
					
				format_and_chart(ws5, sheet) # All Compounds
				bar.update(bar_step); bar_step += 1
				
				wb.save(inputFile)
				format_and_chart(ws6, "Lit " + sheet) # Reported in Lit #here
				bar.update(bar_step); bar_step += 1	
		bar.finish()
							
		"""Contents Page"""
		ws = wb.create_sheet("Index",0)
		contents = [["GC-MS","Combined GC-MS data matched by retention time.",True],
			["Matches", "List of possible matching compounds for each retention time, based on all samples.",True],
			["Statistics_Full", "Statistics for the top hit for each retention time.",True],
			["Statistics", "Statistics for the top hit for each retention time, where the mass spectra match between samples.",True],
			["Statistics - Lit Only", "As above, but only for compounds reported in literature as being present in propellent or GSR.",True],
			["Mean Peak Area", "Graph showing average peak area and Log(PA) for all compounds.",False],
			["Lit Mean Peak Area", "As above, but only for compounds reported in literature as being present in propellent or GSR.",False],
			["Radar Log(PA)", "Radar chart of Log(PA), ONLY for compounds reported in literature.",False],
			["All Peak Area", "Graph showing peak area for all compounds in each sample.",False],
			["Lit All Peak Area", "Graph showing peak area for compounds in each sample reported in literature.",False],
			["All Log(PA)", "Graph showing Log(PA) for all compounds in each sample.",False],
			["Lit All Log(PA)", "Graph showing Log(PA) for compounds in each sample reported in literature.",False],
			["Lit All Log(PA) Line", "Line graph showing Log(PA) for compounds in each sample reported in literature.",False]
		]
		
		#for prefix in prefixList:
		#	contents.append([prefix,"Combined GC-MS data showing 80 largest peaks.",True])
		
		
		print("\nThe worksheets in the output xlsx file are as follows:")
		ws.append(["GunShotMatch Version {} Output".format(__version__)])
		ws.append(["Note that hyperlinks to charts do not work at this time"])
		for row in contents: 
			print(row[0] + " "*(24-len(row[0])) + row[1])
			if row[2]:
				ws.append(["",'=HYPERLINK("#\'{0}\'!A1","{0}")'.format(row[0]),row[1]])
			else:
				ws.append(["","{0}".format(row[0]),row[1]])	
		
		ws.column_dimensions["B"].width = 50.0
		for row in ws.iter_rows("B2:B{}".format(len(contents)+1)):
			for cell in row:
				if cell.value:
					if cell.value.startswith("=HYPERLINK"):
						cell.font = Font(color="0000EE",
									underline="single")
		
		ws.cell(column=4, row=2).value = "Sample list:"
		for index, prefix in enumerate(self.prefixList):
			ws.cell(row=2, column=(5+index)).value = prefix
		
		# Save the file
		wb.save(inputFile)

	def make_archive(self): 
		# Creates single tarball containing main results
		import tarfile 
		tar = tarfile.open(os.path.join(self.RESULTS_DIRECTORY,(self.lot_name + "_" + str(datetime.datetime.now())[:-7].replace("-","").replace(":","").replace(" ","")+'.tar.gz')), mode='w:gz')
		print(self.RESULTS_DIRECTORY)
		print(os.path.join(self.RESULTS_DIRECTORY,self.lot_name + "_FINAL.xlsx"))
		tar.add(os.path.join(self.RESULTS_DIRECTORY,self.lot_name + "_FINAL.xlsx"), arcname=(self.lot_name + "_FINAL.xlsx"))
		tar.add(os.path.join(self.MSP_DIRECTORY,self.lot_name), arcname='MSP')
		tar.add(os.path.join(self.COMPARISON_DIRECTORY,self.lot_name), arcname='MS Comparisons')
		tar.add("ms_comparisons_{}.txt".format(self.lot_name))
		for prefix in self.prefixList:
			try:
				tar.add(os.path.join(self.OUTPUT_DIRECTORY,prefix + ".CSV"), arcname=(prefix + ".CSV"))
			except:
				print("?")
				tar.add(os.path.join(self.OUTPUT_DIRECTORY,prefix + ".csv"), arcname=(prefix + ".CSV"))

		
	"""MS Comparison Functions"""
	def get_mass_spectra(self):
		widgets = [progressbar.widgets.AnimatedMarker(), ' ','Gathering Mass Spectra.', ' ', progressbar.Timer(),]
		bar = progressbar.ProgressBar(widgets=widgets, max_value=progressbar.UnknownLength).start()
		bar_step = 1
		
		"""OpenChrom sometimes saves files as ".CDF.csv". This renames them to just ".csv"."""
		for prefix in self.prefixList:
			if os.path.exists(os.path.join(self.SPECTRA_DIRECTORY,prefix.rstrip("\n\r ")+".CDF.csv")):
				os.rename(os.path.join(self.SPECTRA_DIRECTORY,prefix.rstrip("\n\r ")+".CDF.csv"),os.path.join(self.SPECTRA_DIRECTORY,prefix.rstrip("\n\r ")+".csv"))

		"""Get list of retention times from combined output"""
		#if lot_name == re.sub(r'\d+', '', str(prefixList[1].rstrip("\n\r "))).replace("__","_"):
		#	input_file_name = lot_name + "_COMBINED.CSV"
		#else:
		#	input_file_name = "MERGED_OUTPUT.CSV"
		
		input_file_name = self.lot_name + "_COMBINED.CSV"
		
		with open(input_file_name,"rb") as f:
			csvfile = csv.reader(f, delimiter=";")

			rt_dict = {}
			
			for sample_number in range(self.PL_len):
				for i, row in enumerate(csvfile):
					if i == 0:
						sample_name = row[0+(9*sample_number)]
						rt_dict[sample_name] = []
					elif i > 1:
						if len(row[0+(9*sample_number)])>0:
							"""From https://stackoverflow.com/questions/26367812/appending-to-list-in-python-dictionary"""
							rt_dict.setdefault(sample_name, []).append(row[0+(9*sample_number)])
				f.seek(0)
				bar.update(bar_step); bar_step += 1
		
		for prefix in self.prefixList:		
			"""Pull mass spectra for those retention times"""
			with open(os.path.join(self.SPECTRA_DIRECTORY,prefix.rstrip("\n\r ")+".csv"),"r") as f:
				inputCSVspectra = f.readlines()
			
			outputCSV = [inputCSVspectra[0]]
			for r in rt_dict[prefix.rstrip("\n\r ")]:
				
				if r.lower() == "no peak":
					outputCSV.append("No Peak,No Peak\n")
				else:
					"""From https://gist.github.com/jackiekazil/6201722"""
					retention_time = str(Decimal(Decimal(r).quantize(Decimal('.001'), rounding=ROUND_HALF_UP)))

					for lineNo in range(len(inputCSVspectra)):
						if lineNo > 0:
							try:
								line = inputCSVspectra[lineNo].split(";")
								"""From https://gist.github.com/jackiekazil/6201722"""
								if str(Decimal(Decimal(line[1]).quantize(Decimal('.001'), rounding=ROUND_HALF_UP))) == retention_time: #dec_scan:
									outputCSV.append(inputCSVspectra[lineNo])
							except:
								print(line)
								print(line[1])
								raise
						bar.update(bar_step); bar_step += 1
					
			with open("./"+prefix.rstrip("\n\r ")+"_SPECTRA.CSV","w") as f:
				f.write("".join(outputCSV))	
		bar.finish()
		
	def generate_mass_spectra(self, rt_list, generate_images=True):	
		import utils.MassSpectraPlot as msplot
		if args.no_images:
			print("Generating mass spectra files. Please wait.")
		else:
			print("Generating mass spectra images. Please wait.")
		bar = progressbar.ProgressBar(max_value=len(rt_list))
		
		if not os.path.exists(os.path.join(self.COMPARISON_DIRECTORY,self.lot_name)):
			os.makedirs(os.path.join(self.COMPARISON_DIRECTORY,self.lot_name))

		#Check if the images need to be updated
		images_changed = True
		
		if os.path.exists("./lib/hashes.ini"):
			hash_config = ConfigParser.RawConfigParser()
			hash_config.read("./lib/hashes.ini")
			
			if os.path.isdir(os.path.join(self.COMPARISON_DIRECTORY,self.lot_name)):
				try:
					#if hash_config.get("main","{}_images".format(self.lot_name.replace(" ","_"))) == DirectoryHash.GetHashofDirs(os.path.join(self.COMPARISON_DIRECTORY,self.lot_name)):
					if hash_config.get("main","{}_images".format(hasher(self.prefixList))) == DirectoryHash.GetHashofDirs(os.path.join(self.COMPARISON_DIRECTORY,self.lot_name)):
						print("Images are unchanged. Skipping.")
						images_changed = False
				except ConfigParser.NoOptionError:
					pass
		
		else:
			hash_file = open("./lib/hashes.ini")
			hash_file.write("[main]")
			hash_file.close()
		
		if images_changed: #and not args.no_images:	 # fix
			for prefix in self.prefixList:
				for filename in os.listdir(os.path.join(self.COMPARISON_DIRECTORY,self.lot_name)):
					if prefix in filename:
						os.unlink(os.path.join(self.COMPARISON_DIRECTORY,self.lot_name + "/" + filename))
			
		"""Grab spectra for the matching retention times and package into individual files"""
		for index in rt_list:##here
			output_dir = os.path.join(self.SPECTRA_DIRECTORY,"Output")
			if not os.path.exists(output_dir):
				os.makedirs(output_dir)
			file_path = os.path.join(output_dir,self.lot_name+"_" + str(rt_list[index])+".csv")
			open(file_path, 'w').close() # Initialise output
			output_file = open(file_path, 'a')
			for i, prefix in enumerate(self.prefixList):
				with open("./"+prefix.rstrip("\n\r ")+"_SPECTRA.CSV","r") as f:
					input_csv = f.readlines()
				if i == 0:
					output_file.write(";" + input_csv[0])
				output_file.write(prefix.rstrip("\n\r ") + ";" + input_csv[index])
			output_file.close()
			
			"""generate .xy files"""
			intensity_dict = {}
			
			spectra_file = open(file_path, 'r')
			spectra = spectra_file.readlines()
			mass_list = spectra[0].replace("\n","").split(";")[4:]
			
			for i, raw_row in enumerate(spectra):
				if i >0:
					row = raw_row.replace("\n","").split(";")
					intensity_dict[row[0]] = row[3:] 

			for prefix in self.prefixList:
				intensity_list = intensity_dict[prefix.rstrip("\n\r ")]
				xy_output = []
				if len(intensity_list) > 0:
					for i in range(len(mass_list)):
							xy_output.append(mass_list[i]+".0")
							xy_output.append(" ")

							if len(intensity_list[i]) > 0:
								xy_output.append(intensity_list[i])
							else:
								xy_output.append('0')
														
							xy_output.append("\n")

							for j in range(1,10):
								xy_output.append(mass_list[i]+"."+str(j))
								xy_output.append(" ")
								xy_output.append("0.0")
								xy_output.append("\n")
					
					with open(os.path.join(self.XY_DIRECTORY, prefix + "_" + rt_list[index] + ".xy"),"w") as f:
						f.write("".join(xy_output))
					with open(os.path.join(self.XY_DIRECTORY, prefix + "_" + rt_list[index] + ".txt"),"w") as f:
						f.write("")
					
					if images_changed and not args.no_images:
						msplot.generate_massspectra_plot_automatic_labels(os.path.join(self.XY_DIRECTORY,prefix + "_" + rt_list[index]), os.path.join(self.COMPARISON_DIRECTORY,self.lot_name+"/"+ prefix + "_" + rt_list[index]))
						
					"""Generate .MSP files for NIST MS Search"""
					xy_output = [] #reset xy output
					
					for i in range(len(intensity_list)-1):
						xy_output.append(mass_list[i]+".0")
						xy_output.append(" ")
						if len(intensity_list[i]) > 0:
							xy_output.append(intensity_list[i])
						else:
							xy_output.append('0')
						xy_output.append("\n")
					
					
					
					if not os.path.exists(os.path.join(self.MSP_DIRECTORY,self.lot_name)):
						os.makedirs(os.path.join(self.MSP_DIRECTORY,self.lot_name))
					msp_file = open(os.path.join(self.MSP_DIRECTORY,self.lot_name,prefix.rstrip("\n\r ") + "_" + rt_list[index] + ".MSP"),"w")
					msp_file.write("Name: " + prefix.rstrip("\n\r ") + "_" + rt_list[index] + "\n")
					msp_file.write("Num Peaks: " + str(len(intensity_list)-1) + "\n")
					for row in "".join(xy_output):
						msp_file.write(row.replace("\n",",\n"))
					msp_file.close()
					bar.update(index)
			
			spectra_file.close()
		
		
		if images_changed:
			#Take hash of images to save time in the future
			
			hash_config = ConfigParser.RawConfigParser()
			#hash_config.add_section('main')
			hash_config.read("./lib/hashes.ini")
			#hash_config.set("main","{}_images".format(lot_name.replace(" ","_")), "{}".format(DirectoryHash.GetHashofDirs(os.path.join(self.COMPARISON_DIRECTORY,self.lot_name))))
			hash_config.set("main","{}_images".format(hasher(self.prefixList)), "{}".format(DirectoryHash.GetHashofDirs(os.path.join(self.COMPARISON_DIRECTORY,self.lot_name))))
			with open("./lib/hashes.ini","wb") as configfile:
				hash_config.write(configfile)
		
		return rt_list
		
	def generate_nist_library(self):
		"""Merge MSP Files together for create NIST Library"""
		"""From https://stackoverflow.com/questions/13613336/python-concatenate-text-files"""
		with open(os.path.join(self.MSP_DIRECTORY,self.lot_name+'.msp'), 'w') as outfile:
			for filename in os.listdir(os.path.join(self.MSP_DIRECTORY,self.lot_name)):
				if filename != self.lot_name+'.msp':
					with open(os.path.join(self.MSP_DIRECTORY,self.lot_name + '/' +filename)) as infile:
						for line in infile:
							outfile.write(line)
		
	#	print(lib2nist_path)

		#Check if the library needs to be updated
	#	import ConfigParser
	#	from utils import DirectoryHash
		
		library_changed = True
		
		if os.path.exists("./lib/hashes.ini"):
			hash_config = ConfigParser.RawConfigParser()
			hash_config.read("./lib/hashes.ini")
			
			if os.path.isdir(os.path.join(self.nist_path,self.lot_name)):
				try:
					#if hash_config.get("main","{}_library".format(lot_name.replace(" ","_"))) == DirectoryHash.GetHashofDirs(os.path.join(self.nist_path,self.lot_name)):
					if hash_config.get("main","{}_library".format(hasher(self.prefixList))) == DirectoryHash.GetHashofDirs(os.path.join(self.nist_path,self.lot_name)):
						print("Library is unchanged. Skipping.")
						library_changed = False
				except ConfigParser.NoOptionError:
					pass
		else:
			hash_file = open("./lib/hashes.ini")
			hash_file.write("[main]")
			hash_file.close()
			
			
		if library_changed:
			#Generate NIST Library
			print("\n\nClick exit once the library has been generated. If nothing happens just click exit anyway.")
			if platform.system() == "Linux": 
				print("On Linux you need to navigate to the .MSP file manually. It should be located at {}".format(os.path.join(os.getcwd(),os.path.join(self.MSP_DIRECTORY,self.lot_name + '.msp'))))
			os.system("{}".format("wine " if platform.system() == "Linux" else "") + os.path.join(self.lib2nist_path,"lib2nist.exe") + " " + os.path.join(self.MSP_DIRECTORY,self.lot_name + '.msp'))
				
			#Copy library to NIST directory
			if os.path.isdir(os.path.join(self.nist_path,self.lot_name)):
				shutil.rmtree(os.path.join(self.nist_path,self.lot_name)) #delete existing directory
			shutil.copytree(os.path.join(self.lib2nist_path,self.lot_name), os.path.join(self.nist_path,self.lot_name))
			
			#Take hash of library to save time in the future
			
			hash_config = ConfigParser.RawConfigParser()
			#hash_config.add_section('main')
			hash_config.read("./lib/hashes.ini")
			#hash_config.set("main","{}_library".format(lot_name.replace(" ","_")), "{}".format(DirectoryHash.GetHashofDirs(os.path.join(self.nist_path,self.lot_name))))
			hash_config.set("main","{}_library".format(hasher(self.prefixList)), "{}".format(DirectoryHash.GetHashofDirs(os.path.join(self.nist_path,self.lot_name))))
			with open("./lib/hashes.ini","wb") as configfile:
				hash_config.write(configfile)
			
	def nist_ms_comparison(self, rt_list):
		data_dict = {}

		try:
			generate_ini(self.nist_path, self.lot_name, 50)
			
			perms = []
			for i in permutations(self.prefixList,2):
				"""from https://stackoverflow.com/questions/10201977/how-to-reverse-tuples-in-python"""
				if i[::-1] not in perms:
					perms.append(i)
			
			def remove_chars(input_string):
				return input_string.replace("Hit 10",""
					).replace("Hit 11","").replace("Hit 12","").replace("Hit 13","").replace("Hit 14","").replace("Hit 15",""
					).replace("Hit 16","").replace("Hit 17","").replace("Hit 18","").replace("Hit 19","").replace("Hit 20",""
					).replace("Hit 21","").replace("Hit 22","").replace("Hit 23","").replace("Hit 24","").replace("Hit 25",""
					).replace("Hit 26","").replace("Hit 27","").replace("Hit 28","").replace("Hit 29","").replace("Hit 30",""
					).replace("Hit 31","").replace("Hit 32","").replace("Hit 33","").replace("Hit 34","").replace("Hit 35",""
					).replace("Hit 36","").replace("Hit 37","").replace("Hit 38","").replace("Hit 39","").replace("Hit 40",""
					).replace("Hit 41","").replace("Hit 42","").replace("Hit 43","").replace("Hit 44","").replace("Hit 45",""
					).replace("Hit 46","").replace("Hit 47","").replace("Hit 48","").replace("Hit 49","").replace("Hit 50",""
					).replace("Hit 1","").replace("Hit 2","").replace("Hit 3","").replace("Hit 4","").replace("Hit 5",""
					).replace("Hit 6","").replace("Hit 7","").replace("Hit 8","").replace("Hit 9","").replace("MF:",""
					).replace(":","").replace("<","").replace(">","").replace(" ","").replace("_"+rt_list[index],"")
			
			print("\nMass Spectrum Comparison in progress. Please wait.")
			bar = progressbar.ProgressBar(max_value=len(rt_list))
			bar.update(0)
			with open("./ms_comparisons_{}.txt".format(lot_name),"w") as f:
				for index in range(1,len(rt_list)+1):
					compare_MF = []
					f.write(rt_list[index])
					f.write("\n")
					for prefix in prefixList:
						filename = prefix.rstrip("\r\n") + "_" + rt_list[index] + ".MSP"

						search_results = [prefix.rstrip("\r\n")]

						raw_output = nist_db_connector(self.nist_path,os.path.join(self.MSP_DIRECTORY,self.lot_name,filename))	
							

						for row_index, row in enumerate(raw_output.split("\n")):

							if row_index > 0:
								if rt_list[index] in row:
									if prefix.rstrip("\r\n") not in row:
										search_results.append(remove_chars(row.split(";")[0]))
										search_results.append(remove_chars(row.split(";")[2]))
					
						
	#	for diagnostics				if len(search_results) != 5:
	#						print("RT = {}".format(rt_list[index]))
	#						print("")
	#						print(raw_output)
	#						print(len(raw_output))
	#						print("")
	#						print(search_results)
	#						raw_input(">")
						
						for combo in perms:
							if combo[0] == search_results[0]:
								if combo[1] in search_results:
									f.write(str(combo) + " " + search_results[1+search_results.index(combo[1])]+"\n")
									compare_MF.append(int(search_results[1+search_results.index(combo[1])]))
								else:
									f.write(str(combo) + " Not Found\n")
									compare_MF.append(0)
						bar.update(index)
						
					MF_mean = numpy.mean(compare_MF)
					MF_std = numpy.std(compare_MF, ddof=1)
					MF_rsd = numpy.std(compare_MF, ddof=1)/numpy.mean(compare_MF)
					f.write(str(MF_mean)+"\n")
					f.write((str(MF_std))+"\n")
					f.write(str(MF_rsd)+"\n")
					data_dict[rt_list[index]] = [MF_mean, MF_std, MF_rsd]
					f.write("\n")
					f.write("\n")
		except:
			traceback.print_exc()	#print the error
			nist_cleanup(nist_path)
			sys.exit(1)
		
		print("\n")
		nist_cleanup(self.nist_path)
		return data_dict

	def nist_ms_comparison_2(self, rt_list):
		data_dict = {}

		try:
			generate_ini(self.nist_path, self.lot_name, 50)
			
			perms = []
			for i in permutations(self.prefixList,2):
				"""from https://stackoverflow.com/questions/10201977/how-to-reverse-tuples-in-python"""
				if i[::-1] not in perms:
					perms.append(i)
			
			def remove_chars(input_string):
				return input_string.replace("Hit 10",""
					).replace("Hit 11","").replace("Hit 12","").replace("Hit 13","").replace("Hit 14","").replace("Hit 15",""
					).replace("Hit 16","").replace("Hit 17","").replace("Hit 18","").replace("Hit 19","").replace("Hit 20",""
					).replace("Hit 21","").replace("Hit 22","").replace("Hit 23","").replace("Hit 24","").replace("Hit 25",""
					).replace("Hit 26","").replace("Hit 27","").replace("Hit 28","").replace("Hit 29","").replace("Hit 30",""
					).replace("Hit 31","").replace("Hit 32","").replace("Hit 33","").replace("Hit 34","").replace("Hit 35",""
					).replace("Hit 36","").replace("Hit 37","").replace("Hit 38","").replace("Hit 39","").replace("Hit 40",""
					).replace("Hit 41","").replace("Hit 42","").replace("Hit 43","").replace("Hit 44","").replace("Hit 45",""
					).replace("Hit 46","").replace("Hit 47","").replace("Hit 48","").replace("Hit 49","").replace("Hit 50",""
					).replace("Hit 1","").replace("Hit 2","").replace("Hit 3","").replace("Hit 4","").replace("Hit 5",""
					).replace("Hit 6","").replace("Hit 7","").replace("Hit 8","").replace("Hit 9","").replace("MF:",""
					).replace(":","").replace("<","").replace(">","").replace(" ","").replace("_"+rt_list[index],"")
			
			print("\nMass Spectrum Comparison in progress. Please wait.")
			bar = progressbar.ProgressBar(max_value=len(rt_list))
			bar.update(0)
			with open("./ms_comparisons_{}.txt".format(self.lot_name),"w") as f:
				for index in range(1,len(rt_list)+1):
					compare_MF = []
					f.write(rt_list[index])
					f.write("\n")
					
					"""From https://stackoverflow.com/questions/13613336/python-concatenate-text-files"""
					with open(os.path.join(self.MSP_DIRECTORY,self.lot_name+'_{}.msp'.format(rt_list[index])), 'w') as outfile:
						for filename in os.listdir(os.path.join(self.MSP_DIRECTORY,self.lot_name)):
							if filename != self.lot_name+'.msp':
								if rt_list[index] in filename:
									with open(os.path.join(self.MSP_DIRECTORY,self.lot_name + '/' +filename)) as infile:
										for line in infile:
											outfile.write(line)
					
					
					raw_output = nist_db_connector(self.nist_path,os.path.join(self.MSP_DIRECTORY,self.lot_name+'_{}.msp'.format(rt_list[index])),self.PL_len)	
							

					for prefix in self.prefixList:
						search_results = [prefix.rstrip("\r\n")]
						start_row = -1
						for row_index, row in enumerate(raw_output.split("\n")):
							if "Unknown: {}".format(prefix) in row:
								start_row = row_index
							elif start_row >=0:
								if row_index > start_row:
									if row.startswith("Unknown"):
										break
									elif rt_list[index] in row:
										if prefix.rstrip("\r\n") not in row:
												search_results.append(remove_chars(row.split(";")[0]))
												search_results.append(remove_chars(row.split(";")[2]))
						
	#	for diagnostics				if len(search_results) != 5:
	#						print("RT = {}".format(rt_list[index]))
	#						print("")
	#						print(raw_output)
	#						print(len(raw_output))
	#						print("")
	#						print(search_results)
	#						raw_input(">")
						
						#print search_results
						#raw_input(">")
						for combo in perms:
								if combo[0] == search_results[0]:
									if combo[1] in search_results:
										f.write(str(combo) + " " + search_results[1+search_results.index(combo[1])]+"\n")
										compare_MF.append(int(search_results[1+search_results.index(combo[1])]))
									else:
										f.write(str(combo) + " Not Found\n")
										compare_MF.append(0)
					bar.update(index)
							
					MF_mean = numpy.mean(compare_MF)
					MF_std = numpy.std(compare_MF, ddof=1)
					MF_rsd = numpy.std(compare_MF, ddof=1)/numpy.mean(compare_MF)
					f.write(str(MF_mean)+"\n")
					f.write((str(MF_std))+"\n")
					f.write(str(MF_rsd)+"\n")
					data_dict[rt_list[index]] = [MF_mean, MF_std, MF_rsd]
					f.write("\n")
					f.write("\n")
		except:
			traceback.print_exc()	#print the error
			nist_cleanup(self.nist_path)
			sys.exit(1)
		
		print("\n")
		nist_cleanup(self.nist_path)
		return data_dict
	
	
	def doAuto1(self):
		print("Auto = " + str(self.auto1))
		if "combine" in self.auto1:
			print("Combine")
			self.doCombine()
		if "merge" in self.auto1:
			print("Merge")
			self.doMerge()
		if "open_lo" in self.auto1:
			print("Merge")
			self.do_open_lo()
		if "jigsaw" in self.auto1:
			print("Jigsaw")
			self.doJigsaw()
		if "counter" in self.auto1:
			print("Counter")
			if "open_lo" in self.auto1:
				self.doCounter(seperator=",")
			else:
				self.doCounter()
	def doAuto2(self):
		print("Auto = " + str(self.auto2))
		if "combine" in self.auto2:
			print("Combine")
			self.doCombine()
		if "merge" in self.auto2:
			print("Merge")
			self.doMerge()
		if "open_lo" in self.auto2:
			print("Merge")
			self.do_open_lo()
		if "jigsaw" in self.auto2:
			print("Jigsaw")
			self.doJigsaw()
		if "counter" in self.auto2:
			print("Counter")
			if "open_lo" in self.auto2:
				self.doCounter(seperator=",")
			else:
				self.doCounter()
	def doAuto3(self):
		print("Auto = " + str(self.auto3))
		if "combine" in self.auto3:
			print("Combine")
			self.doCombine()
		if "merge" in self.auto3:
			print("Merge")
			self.doMerge()
		if "open_lo" in self.auto3:
			print("Merge")
			self.do_open_lo()
		if "jigsaw" in self.auto3:
			print("Jigsaw")
			self.doJigsaw()
		if "counter" in self.auto3:
			print("Counter")
			if "open_lo" in self.auto3:
				self.doCounter(seperator=",")
			else:
				self.doCounter()


	
	"""Run Combine"""
	def doCombine(self):
		print("\nCombine\n")
		for prefix in self.prefixList:
			GCMScombine(prefix.rstrip("\n\r "),self.OUTPUT_DIRECTORY, self.CSV_DIRECTORY)

	"""Run Spacer"""
	def doSpace(self):
		print("\nSpace\n")
		batch_spacer(self.prefixList,self.OUTPUT_DIRECTORY,self.CSV_DIRECTORY)
	
	"""Run Merge"""
	"""This also runs Spacer"""
	def doMerge(self):
		print("\nMerge\n")
		self.file_for_lo = self.Merge()

	"""Open LibreOffice"""
	def do_open_lo(self):
		self.open_lo(self.file_for_lo,)
		
	"""Run Jigsaw"""
	def doJigsaw(self):
		self.jigsaw()
		
	"""Run Counter"""
	def doCounter(self,seperator=";"):
		print("\nCounter\n")
		self.Match_Counter(seperator)
		self.make_archive()
		
		

def create_logger(verbose=False):
	# Configure Logging
	# from https://docs.python.org/2/howto/logging-cookbook.html
	logger = logging.getLogger("GunShotMatch") # create logger

	logger.setLevel(logging.DEBUG)

	# create file handler which logs even debug messages
	fh = logging.FileHandler('GunShotMatch.log')
	fh.setLevel(logging.DEBUG)
	# create console handler with a higher log level
	ch = logging.StreamHandler()
	if verbose:
		ch.setLevel(logging.DEBUG)
	else:
		ch.setLevel(logging.ERROR)
	# create formatter and add it to the handlers
	formatter = logging.Formatter('%(asctime)s - %(levelname)s: %(message)s',"%H:%M:%S")
	fh.setFormatter(formatter)
	ch.setFormatter(formatter)
	# add the handlers to the logger
	logger.addHandler(fh)
	logger.addHandler(ch)
	
	return logger





if __name__ == '__main__':

	clear()		#clear the display
	
	startup_string = """\n\n{0} Version {1} running on {3}.
Copyright {2} Dominic Davis-Foster.
""".format(program_name,__version__,copyright,platform.system())
#Loading. Please wait...""".format(program_name,__version__,copyright,platform.system())
	
	print(startup_string)
	
	openLO = False

	#Command line switches
	parser = argparse.ArgumentParser()
	parser.add_argument("--info",help="Show program info.", action='store_true')
	parser.add_argument("--combine",help="Combine GS and MS data for the samples", action='store_true')
	parser.add_argument("--space",help="Space the samples", action='store_true')
	parser.add_argument("--merge",help="Run Space, then Merge the files into one.", action='store_true')
	parser.add_argument("-v","--verbose",help="Turns on verbosity for diagnostics", action='store_true')
	parser.add_argument("-s","--samples",help="List of samples to run.",nargs='+')
	parser.add_argument("--dependencies",help="Check for and list dependencies.",action='store_true')
	parser.add_argument("--counter",help="Count the number of times a hit appears.",action='store_true')
	parser.add_argument("--jigsaw",help="Combine the output in the terminal.",action='store_true')
	parser.add_argument("--auto",help="Autoprocess list 1.",action='store_true')
	parser.add_argument("--auto1",help="Autoprocess list 1.",action='store_true')
	parser.add_argument("--auto2",help="Autoprocess list 2.",action='store_true')
	parser.add_argument("--auto3",help="Autoprocess list 3.",action='store_true')
	#parser.add_argument("--config",help="Config file to use.",nargs='+')
	parser.add_argument("--default",help="Reload Default configuration.",action='store_true')
	parser.add_argument("--no-images",help="Don't generate spectra images.",action='store_true')
	
	args = parser.parse_args()
	verbose = args.verbose
	
	logger = create_logger(args.verbose)
	
	if len(sys.argv)==1:
		logger.warning("No options supplied.")
		print('')
		parser.print_help(sys.stderr)
		sys.exit(1)
	
	if len(sys.argv)==2 and verbose:
		logger.warning("No options supplied.")
		print('')
		parser.print_help(sys.stderr)
		sys.exit(1)

	if args.info:
		print(__doc__)
		parser.print_help()	#show help and info
		sys.exit(0)
			
#	import ConfigParser
	config = ConfigParser.ConfigParser()
	config.read("./lib/gsmatch.ini")

	if args.default:
		print("\nReloading Default Configuration")
		shutil.copyfile("./lib/default.ini", "./config.ini")
		config.set("MAIN","last_config","config.ini")
		with open("./lib/gsmatch.ini", 'w') as configfile:
			config.write(configfile)
		sys.exit(0)
			
#	if args.config:
#		config.set("MAIN","last_config","config.ini")
		
	last_config = config.get("MAIN","last_config")

	"""Set up GunShotMatch Class"""
	GSM = GunShotMatch("./config.ini")
			
	"""Sample List"""
	if args.samples:
		GSM.prefixList = args.samples
		# overrides whatever was set from the config file

	print("\nReady")
	print("Samples to process: " + str(GSM.prefixList))

	
	"""Define Exit Functions"""
	if "-h" not in str(sys.argv):
		atexit.register(final_cleanup,GSM.prefixList,GSM.OUTPUT_DIRECTORY, GSM.nist_path,GSM.lot_name)
		atexit.register(reload_ini, GSM.nist_path)

	# Autoprocess1
	auto = []
	if args.auto or args.auto1:
		print("\nAutoprocess 1")
		GSM.doAuto1()
		
	# Autoprocess2
	elif args.auto2:
		print("\nAutoprocess 2")
		GSM.doAuto2()
		
	# Autoprocess3
	elif args.auto3:
		print("\nAutoprocess 3")
		GSM.doAuto3()
		
	else:
		# Run Combine
		if args.combine:
			print("\nCombine\n")
			GSM.doCombine()


		"""Run Spacer"""
		if args.space:
			print("\nSpace\n")
			GSM.doSpace()

		
		"""Run Merge"""
		"""This also runs Spacer"""
		if args.merge:
			print("\nMerge\n")
			GSM.doMerge()


		"""Open LibreOffice"""
		if openLO:
			GSM.do_open_lo()

			
		"""Run Jigsaw"""
		if args.jigsaw:
			GSM.doJigsaw()

			
		"""Run Counter"""
		if args.counter:
			print("\nCounter\n")
			if openLO:
				GSM.doCounter(seperator=",")
			else:
				GSM.doCounter()


	

	

	
	
			
	final_cleanup(GSM.prefixList,GSM.OUTPUT_DIRECTORY, GSM.nist_path, GSM.lot_name)
	print("\nComplete.")
	

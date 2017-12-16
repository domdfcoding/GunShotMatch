#!/usr/bin/env python
# -*- coding: utf-8 -*-
# encoding=utf8
#
#  GSMatch.py
#  
#  Copyright 2017 Dominic Davis-Foster <dominic@davis-foster.co.uk>
#  
#  This program is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 2 of the License, or
#  (at your option) any later version.
#  
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#  
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#   

import os
import sys
import traceback
import argparse
import csv
import warnings
import re

from decimal import Decimal, ROUND_HALF_UP
from collections import Counter
from itertools import chain, izip

"""From https://stackoverflow.com/questions/21129020/how-to-fix-unicodedecodeerror-ascii-codec-cant-decode-byte
Big shout out to fisherman https://stackoverflow.com/users/2309581/fisherman"""
reload(sys)  
sys.setdefaultencoding('utf8')

print("""\rGunShotMatch Version 3.5
Copyright 2017 Dominic Davis-Foster

Ready
""")

verbose = False
	
def check_dependencies():
	"""Based on https://stackoverflow.com/questions/22213997/programmatically-check-if-python-dependencies-are-satisfied"""
	from pkgutil import iter_modules
	modules = set(x[1] for x in iter_modules())
	
	missing_modules = []
	for requirement in ["xlsxwriter","openpyxl"]:
		if not requirement in modules:
			missing_modules.append(requirement)
	if len(missing_modules)==0:
		print "All modules installed"
	else:
		print """\rThe following modules are missing.
Please check the documentation."""
		print missing_modules
	print ""
	sys.exit(0)

def verbosePrint(text):	#print only if the --verbose flag is set
	if verbose:
		print(str(text))

def GCMScombine(inputFile):
	"""Define variables"""
	GCpeaks=[] #Stores relevant data for GC peaks, in this case the RT and peak area
	stripCSVms=[] #Stores the MS data after removing redundant lines
	
	"""Open the output file"""
	while True:
		try:
			outputCSV = open("./GunShotMatch Output/" + inputFile+".CSV","w")
			break
		except IOError:
			print "The file \"" + inputFile + ".CSV\" is locked for editing in another program. Press any key to try again."
			raw_input(">")
	
	"""Read GC Data"""
	with open("./CSV/"+inputFile+"_GC_80.CSV","r") as f:
		inputCSVgc = f.readlines()
	
	"""Pull relevant GC Data"""
	for lineNo in range(2,len(inputCSVgc)): 
		verbosePrint(lineNo)
		line = inputCSVgc[lineNo]
		for x in range(0,81):
			if line[:2]== str((["1,","2,","3,","4,","5,","6,","7,","8,","9,"] + range(10,81)+['',''])[x]):
				GCpeaks.append(line.split(",")[1] + "," + line.split(",")[4] + ",")
				 
				#add extra '+ line.split(",")[x] + ",""' to include additional columns, and change headings later on
		verbosePrint(str(GCpeaks) + "\n")
		
	"""Remove redundant rows from GC-MS"""
	with open("./CSV/"+inputFile+"_MS_80.CSV","r") as f:	#open MS data
		inputCSVms = f.readlines()						#read lines to a list
	
	for lineNo in range(0,len(inputCSVms)):				
		line = inputCSVms[lineNo]
		verbosePrint(line)
		
		#Include only rows that relate to compounds
		for page in range(1,81):	#Replace rows of "Page 1 of 30" with a single cell
			if '"Page '+ str(page) +' of 80","Page '+ str(page) +' of 80"' in line:
				stripCSVms.append("Page "+ str(page) +" of 80,,,,,\n")
		if "mainlib" in line:
			stripCSVms.append(line)

	"""Add spaces to MS Data"""
	verbosePrint(inputCSVms)
	
	for lineNo in range(0,len(stripCSVms)+80):
		try:
			line = stripCSVms[lineNo]
			verbosePrint(line)
		except IndexError:
			next
		
		try:	#add additional lines to ensure regular spacing
			if line[:4] == "Page" and (stripCSVms[lineNo-1])[:4] == "Page":
				for i in range(5):
					stripCSVms.insert(lineNo,",,,,,,\n")
				#print str(lineNo) + "Insert 5"
				lines_to_add = 0
			if line[:4] == "Page" and (stripCSVms[lineNo-1])[:1] == "1":
				for i  in range(4):
					stripCSVms.insert(lineNo,",,,,,,\n")
				#print str(lineNo) + "Insert 4"
			if line[:4] == "Page" and (stripCSVms[lineNo-1])[:1] == "2":
				for i  in range(3):
					stripCSVms.insert(lineNo,",,,,,,\n")
				#print str(lineNo) + "Insert 3"
			if line[:4] == "Page" and (stripCSVms[lineNo-1])[:1] == "3":
				stripCSVms.insert(lineNo,",,,,,,\n")
				stripCSVms.insert(lineNo,",,,,,,\n")
				#print str(lineNo) + "Insert 2"
			if line[:4] == "Page" and (stripCSVms[lineNo-1])[:1] == "4":
				for i  in range(1):
					stripCSVms.insert(lineNo,",,,,,,\n")
				#print str(lineNo) + "Insert 1"
		except IndexError:
			traceback.print_exc()
			
	"""Combine GC and MS Data"""
	verbosePrint(stripCSVms)
	for lineNo in range(0,len(stripCSVms)+20):
		try:
			line = stripCSVms[lineNo]
			#Add extra columns in front of lines
			verbosePrint(line)			# for matches to ensure alignment 
			if line[:4] != "Page":		# after merging GC data
				stripCSVms[stripCSVms.index(line)] = ",,"+line
		
		except IndexError:
			next
			
	
	for GCpeakNO in range(0,81):
		lineNo = 0							#Insert the GC data into the
		while lineNo < len(stripCSVms):		# applicable rows of MS data
			verbosePrint (stripCSVms)
			line = stripCSVms[lineNo]
			if line[:7] == "Page "+str((["1 ","2 ","3 ","4 ","5 ","6 ","7 ","8 ","9 "] + range(10,81) + ["",""])[GCpeakNO]):
				verbosePrint(GCpeaks[GCpeakNO])
				stripCSVms.insert(lineNo,GCpeaks[GCpeakNO])
				verbosePrint(stripCSVms)
				lineNo = lineNo+1
			
			lineNo = lineNo+1		
	
	"""Add column heading"""
	stripCSVms.insert(0,inputFile + "\nRetention Time,Peak Area,,Lib,Match,R Match,Name,CAS Number,Notes\n")
#	stripCSVms.insert(0,"Retention Time,Peak Area,,Lib,Match,R Match,Name,CAS Number,Notes")

	"""Save	"""
	outputCSV.write(''.join(stripCSVms))
	print(inputFile + " Done!")
	
	return

def info():
	print ("""
Program for the analysis of OGSR samples to identify matching compounds 
between samples. Can currently process PerkinElmer GC-MS .CSV files only.
	
Samples to be analysed can be passed following the -f/--samples argument, 
or listed in a file called "list.txt" in the base directory. Each sample 
name must be on a new line. For more information read the documentation.
	
The files must be in the subdirectory CSV and called 
		e.g. GECO_Case2_MS.CSV and GECO_Case2_GC.CSV

PerkinElmer is a trademark of PerkinElmer Inc.

Tested with Python 2 ONLY.
""")
	return
	
def getRTlist(sample):
	tempRTlist = []	#list of retention times for the sample passed as an argument
	
	f = open("./"+sample+".CSV")		#Open GCMS merged output of sample
	csv_f = csv.reader(f)

	for row in csv_f:
		verbosePrint(row)
		try:
			verbosePrint(str(row[0]))
			if row[0] not in ["Retention Time",sample]:
				if len(row[0]) > 0:					#Gathers retention
					tempRTlist.append(row[0])		# time data
		except IndexError:
			traceback.print_exc()		#Print error but continue

		verbosePrint(tempRTlist)
	return tempRTlist

def single_spacer(prefix, GCpeaks30):
	"""Space a single sample"""
	spacedOutput = []
	csv_input = []
	f = open("./GunShotMatch Output/" +(prefix.rstrip("\n\r "))+".CSV")
	csv_f = csv.reader(f)
	
	for row in csv_f:
		csv_input.append(row)

	for row in csv_input:
		if len(row)>7:
			if row[7] == "":
				del row[7]
		while len(row)<9:
			row.append("")
	print GCpeaks30
	
	spacedOutput.append(str(csv_input[0]))
	spacedOutput.append(str(csv_input[1]))
	
	done_times = []
	print GCpeaks30
	raw_input(">")
	for time in GCpeaks30:
		for row_num in range(len(csv_input)):
			if row_num not in [0,1]:
				if len(csv_input[row_num][0]) > 0:
					if csv_input[row_num][0] not in done_times:
						if csv_input[row_num][0] == time:
							spacedOutput.append(str(csv_input[row_num]))
							spacedOutput.append(str(csv_input[row_num+1]))
							spacedOutput.append(str(csv_input[row_num+2]))
							spacedOutput.append(str(csv_input[row_num+3]))
							spacedOutput.append(str(csv_input[row_num+4]))
							spacedOutput.append(str(csv_input[row_num+5]))
							done_times.append(time)
							break
						elif float(time) < float(csv_input[row_num][0]):
							for i in range(6):
								spacedOutput.append(str(['', '', '', '', '', '', '', '', '']))
							break
		done_times.append(time)
			
	with open("./GunShotMatch Output/" + prefix.rstrip("\n\r ") + "_SPACED.CSV","w") as f:
		joinedOutput = '\n'.join(spacedOutput)
		f.write(joinedOutput.replace("[",'').replace("]",'').replace("\"",'\''))	
	return spacedOutput
		
def batch_spacer(prefixList):
	"""Space batch of samples"""
#	with open("./list.txt","r") as f:
#		prefixList = f.readlines()
		
	GCpeaks30 = get_top_30(prefixList)
	for prefix in prefixList:
		single_spacer(prefix, GCpeaks30)

def get_top_30(prefixList):
	"""Get list of top 30 peaks in all samples"""
	GCpeaks30 = []
	for prefix in prefixList:
		with open("./CSV/"+prefix.rstrip("\n\r ")+"_GC_30.CSV","r") as f:
			inputCSVgc30 = f.readlines()
	
		for lineNo in range(2,len(inputCSVgc30)): 
			verbosePrint(lineNo)
			line = inputCSVgc30[lineNo]
			
			for x in range(0,81):
				if line.split(",")[0] == str((range(1,81) + ["",""])[x]):
					GCpeaks30.append(line.split(",")[1])
					#add extra '+ line.split(",")[x] + ",""' to include additional columns, and change headings later on
			print GCpeaks30
			#raw_input(">")
	GCpeaks30 = list(set(GCpeaks30)) #Remove duplicates
	GCpeaks30.sort(key=float) #Sort list of retention times low to high

	return GCpeaks30

def get_mass_spectra(prefixlist):
	"""OpenChrom sometimes saves files as ".CDF.csv". This renames them to just ".csv"."""
	for prefix in prefixList:
		if os.path.exists("./Spectra_CSV/"+prefix.rstrip("\n\r ")+".CDF.csv"):
			os.rename("./Spectra_CSV/"+prefix.rstrip("\n\r ")+".CDF.csv","./Spectra_CSV/"+prefix.rstrip("\n\r ")+".csv")

	"""Get list of retention times from combined output"""
	if re.sub(r'\d+', '', str(prefixList[0].rstrip("\n\r "))).replace("__","_") == re.sub(r'\d+', '', str(prefixList[1].rstrip("\n\r "))).replace("__","_"):
		input_file_name = re.sub(r'\d+', '', str(prefixList[0].rstrip("\n\r "))).replace("__","_") + "_COMBINED.CSV"
	else:
		input_file_name = "MERGED_OUTPUT.CSV"
	
	with open(input_file_name,"rb") as f:
		csvfile = csv.reader(f)

		rt_dict = {}
		
		for sample_number in range(len(prefixList)):
			for i, row in enumerate(csvfile):
				if i == 0:
					sample_name = row[0+(9*sample_number)]
					rt_dict[sample_name] = []
				elif i > 1:
					if len(row[0+(9*sample_number)])>0:
						"""From https://stackoverflow.com/questions/26367812/appending-to-list-in-python-dictionary"""
						rt_dict.setdefault(sample_name, []).append(row[0+(9*sample_number)])
			f.seek(0)
	print rt_dict

		
				
	
	for prefix in prefixList:
		#with open("./CSV/"+prefix.rstrip("\n\r ")+"_GC_80.CSV","r") as f:
		#	inputCSVgc80 = f.readlines()
		#	GCpeaks80 = []
		#	
		#	for lineNo in range(2,len(inputCSVgc80)): 
		#		for x in range(0,81):
		#			if inputCSVgc80[lineNo][:2]== (["1 ","2 ","3 ","4 ","5 ","6 ","7 ","8 ","9 "] + range(10,81) + ["",""])[x]:
		#				GCpeaks80.append(inputCSVgc80[lineNo].split(",")[1])
		#	
		#print GCpeaks80
		
		"""Pull mass spectra for those retention times"""
		with open("./Spectra_CSV/"+prefix.rstrip("\n\r ")+".csv","r") as f:
			inputCSVspectra = f.readlines()
		
		outputCSV = [inputCSVspectra[0]]
		for r in rt_dict[prefix.rstrip("\n\r ")]:
			
			if r.lower() == "no peak":
				outputCSV.append("No Peak,No Peak\n")
			else:
				"""From https://gist.github.com/jackiekazil/6201722"""
				retention_time = str(Decimal(Decimal(r).quantize(Decimal('.001'), rounding=ROUND_HALF_UP)))
				
				#dec_scan = Decimal(float(scan))
				

				for lineNo in range(len(inputCSVspectra)):
					#raw_input(">")
					#print dec_scan
					#print scan
					if lineNo > 0:
						line = inputCSVspectra[lineNo].split(";")
						
					#	print Decimal(Decimal(line[1]).quantize(Decimal('.001'), rounding=ROUND_HALF_UP))
					"""From https://gist.github.com/jackiekazil/6201722"""
						if str(Decimal(Decimal(line[1]).quantize(Decimal('.001'), rounding=ROUND_HALF_UP))) == retention_time: #dec_scan:
					#		print inputCSVspectra[lineNo]
							outputCSV.append(inputCSVspectra[lineNo])
			
		#print outputCSV
		
		#raw_input(">")
		with open("./"+prefix.rstrip("\n\r ")+"_SPECTRA.CSV","w") as f:
			f.write("".join(outputCSV))	

def compare_mass_spectra(prefixList, retentionTime):
	
	no_peaks = []
	rt_list = {}
	
	"""Get list of retention times from Sample 1, unless peak missing, in which case use Sample 2, 3, 4 etc."""
	for i, prefix in enumerate(prefixList):
		with open("./"+prefix.rstrip("\n\r ")+"_SPECTRA.CSV","r") as f:
			csvfile = csv.reader(f, delimiter=";")
			if i == 0:
				for j, row in enumerate(csvfile):
					#print row
					if j > 0:
						if row[1] != "No Peak":
							"""From https://gist.github.com/jackiekazil/6201722"""
							rt_list[j] = str(Decimal(Decimal(row[1]).quantize(Decimal('.001'), rounding=ROUND_HALF_UP)))
							#open(re.sub(r'\d+', '', str(prefix.rstrip("\n\r "))).replace("__","_")+"_" + str(Decimal(Decimal(row[1]).quantize(Decimal('.001'), rounding=ROUND_HALF_UP)))+".csv", 'a').close()
						else:
							no_peaks.append(row)
			if i>0:
				for j, row in enumerate(csvfile):
					for k in no_peaks:
						if j == k:
							if row[1] != "No Peak":
								"""From https://gist.github.com/jackiekazil/6201722"""
								rt_list[j] = str(Decimal(Decimal(row[1]).quantize(Decimal('.001'), rounding=ROUND_HALF_UP)))
								#open(re.sub(r'\d+', '', str(prefix.rstrip("\n\r "))).replace("__","_")+"_" + str(Decimal(Decimal(row[1]).quantize(Decimal('.001'), rounding=ROUND_HALF_UP)))+".csv", 'a').close()
								del no_peaks[k]
	
	"""Grab spectra for the matching retention times and package into individual files"""
	for index in rt_list:
		output_file = open(re.sub(r'\d+', '', str(prefix.rstrip("\n\r "))).replace("__","_")+"_" + str(rt_list[index])+".csv", 'a')
		for i, prefix in enumerate(prefixList):
			with open("./"+prefix.rstrip("\n\r ")+"_SPECTRA.CSV","r") as f:
				input_csv = f.readlines()
			if i == 0:
				output_file.write(";" + input_csv[0])
			output_file.write(prefix.rstrip("\n\r ") + ";" + input_csv[index])
		output_file.close()
		
		"""generate .xy files"""
		intensity_dict = {}
		
		spectra_file = open(re.sub(r'\d+', '', str(prefix.rstrip("\n\r "))).replace("__","_")+"_" + str(rt_list[index])+".csv", 'r')
		spectra = spectra_file.readlines()
		mass_list = spectra[0].replace("\n","").split(";")[4:]
		#print mass_list
		
		for i, raw_row in enumerate(spectra):
			if i >0:
				row = raw_row.replace("\n","").split(";")
				intensity_dict[row[0]] = row[3:] 
		print intensity_dict
		
		
		for prefix in prefixList:
			intensity_list = intensity_dict[prefix.rstrip("\n\r ")]
			xy_output = []
			print intensity_list
			for i in range(456):
				xy_output.append(mass_list[i]+".0")
				xy_output.append(" ")
				xy_output.append(intensity_list[i])
				xy_output.append("\n")
				for j in range(1,10):
					xy_output.append(mass_list[i]+"."+str(j))
					xy_output.append(" ")
					xy_output.append("0.0")
					xy_output.append("\n")
			
			#print xy_output
			
			with open("./xy/" + prefix.rstrip("\n\r ") + "_" + rt_list[index] + ".xy","w") as f:
				f.write("".join(xy_output))
			with open("./xy/" + prefix.rstrip("\n\r ") + "_" + rt_list[index] + ".txt","w") as f:
				f.write("")
		
			"""Generate .MSP files for NIST MS Search"""
			msp_file = open("./MSP/" + prefix.rstrip("\n\r ") + "_" + rt_list[index] + ".MSP","w")
			msp_file.write("Name: " + prefix.rstrip("\n\r ") + "_" + rt_list[index] + "\n")
			msp_file.write("Num Peaks: " + str(len(intensity_list)) + "\n")
			for row in "".join(xy_output):
				msp_file.write(row.replace("\n",",\n"))
			msp_file.close()
			
		sys.exit()

#for i in [2,4,6,8,10,12,14,16,18,20,22,24,26]:
#			try:
#				intensity_dict[outputCSV[i][:-1]] = ";".join(outputCSV[i+1].replace("\n","").split(";")[2:])
#			except IndexError:
#				print ""
#		#print intensity_dict	
	

def convert_to_xlsx(csv_input_file):
	"""From http://coderscrowd.com/app/public/codes/view/201"""

	# if we read f.csv we will write f.xlsx
	wb = xlsxwriter.Workbook(csv_input_file.lower().replace(".csv",".xlsx"))
	ws = wb.add_worksheet(csv_input_file.lower().replace(".csv","")[:29]) 
	with open("./" + csv_input_file,'r') as csvfile:
		table = csv.reader(csvfile)
		i = 0
		# write each row from the csv file as text into the excel file
		# this may be adjusted to use 'excel types' explicitly (see xlsxwriter doc)
		wrap_format = wb.add_format()
		wrap_format.set_text_wrap(True)
		
		for row in table:
			ws.write_row(i, 0, row)#,wrap_format)
			i += 1
	wb.close()

def formatXLSX(inputFile):
	"""From http://openpyxl.readthedocs.io/en/default/"""
	wb = load_workbook(inputFile)
	
	# grab the active worksheet
	ws = wb.active
	
	col_list = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ','BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ','CA','CB','CC','CD','CE','CF','CG','CH','CI','CJ','CK','CL','CM','CN','CO','CP','CQ','CR','CS','CT','CU','CV','CW','CX','CY','CZ','DA','DB','DC','DE','DF'] #Only Supports 12 samples
#	col_list = ['A','B','D','E','F','G','I','J','K','M','N','O','P','R','S','T','V','W','X','Y','AA','AB','AC','AE','AF','AG','AH','AJ','AK','AL','AN','AO','AP','AQ','AS','AT','AU','AW','AX','AY','AZ','BA','BB','BC','BD','BF','BG','BH','BI','BK','BL','BM','BO','BP','BQ','BR','BT','BU','BV','BX','BY','BZ','CA','CC','CD','CE','CG','CH','CI','CJ','CL','CM','CN','CP','CQ','CR','CS','CU','CV','CW','CY','CZ','DA','DB','DE','DF'] #Only Supports 12 samples
	"""From https://stackoverflow.com/questions/38619471/iterate-through-all-rows-in-specific-column-openpyxl
	and https://stackoverflow.com/questions/14463277/how-to-disable-python-warnings"""

	with warnings.catch_warnings():
		warnings.simplefilter("ignore")
		
		for sample_offset in range (0,11):
			for column_index in [0,1,3,4,5,6,8]:
				column_name = col_list[column_index + (sample_offset*9)]
			
				col_string = column_name + '{}:' + column_name + '{}'
		
				for row in ws.iter_rows(col_string.format(ws.min_row,ws.max_row)):
					for cell in row:
						cell.alignment = Alignment(horizontal='center',
										vertical='center',
										wrap_text=True)
	
			width_list = [9,11,2.5,0.63,6,6,45,12,6,'']
			
			for column_index2 in range(0,9):
				column_name2 = col_list[column_index2 + (sample_offset*9)]
				col_width = width_list[column_index2]
				ws.column_dimensions[column_name2].width = col_width
				if column_index2 == 3:
					col_string2 = column_name2 + '{}:' + column_name2 + '{}'
					for row in ws.iter_rows(col_string2.format(ws.min_row,ws.max_row)):
						for cell in row:
							cell.value = None
					ws.column_dimensions[column_name2].hidden = True
				
				if column_index2 == 0:
					ws.merge_cells(column_name2 + '1:' + (col_list[(column_index2 + 8)+(sample_offset*9)]) + '1')

	# Save the file
	wb.save(inputFile)

def Merge(prefixList):
	GCpeaks30 = get_top_30(prefixList)
	
	if len(prefixList) == 1:
		print "Require two or more samples to combine. Check ./list.txt or --samples and try again."
		print ""
		sys.exit(info())

	if len(prefixList) > 1:
		sample_space_list = {"sample_1_spaced": single_spacer(prefixList[0], GCpeaks30), "sample_2_spaced": single_spacer(prefixList[1], GCpeaks30)}

	for i in range(2,12):
		if len(prefixList) > i:
			sample_space_list["sample_" + str(i+1) + "_spaced"] = single_spacer(prefixList[i], GCpeaks30)
		else: sample_space_list["sample_" + str(i+1) + "_spaced"] = ['']*len(sample_space_list["sample_1_spaced"])
	
	
	#print sample_space_list
	#print max((len(v), k) for k,v in sample_space_list.iteritems())[0]
	#print len(sample_space_list["sample_1_spaced"])
	for i in range(1,12):
		"""From https://stackoverflow.com/questions/10895567/find-longest-string-key-in-dictionary"""
		while len(sample_space_list["sample_" + str(i+1) + "_spaced"]) < max((len(v), k) for k,v in sample_space_list.iteritems())[0]:
			"""From https://stackoverflow.com/questions/26367812/appending-to-list-in-python-dictionary"""
			sample_space_list.setdefault("sample_" + str(i+1) + "_spaced", []).append(str(['', '', '', '', '', '', '', '', '']))
		verbosePrint(len(sample_space_list["sample_" + str(i+1) + "_spaced"]))
	
	
	newline_List = ["\n"]*len(sample_space_list["sample_1_spaced"])		# Fix This!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
	
	""" from https://stackoverflow.com/questions/11125212/interleaving-lists-in-python"""	
	single_CSV = list(chain.from_iterable(izip(sample_space_list["sample_1_spaced"],sample_space_list["sample_2_spaced"],sample_space_list["sample_3_spaced"],sample_space_list["sample_4_spaced"],
			sample_space_list["sample_5_spaced"],sample_space_list["sample_6_spaced"],sample_space_list["sample_7_spaced"],sample_space_list["sample_8_spaced"],sample_space_list["sample_9_spaced"],
			sample_space_list["sample_10_spaced"],sample_space_list["sample_11_spaced"],sample_space_list["sample_12_spaced"],newline_List)))

	if re.sub(r'\d+', '', str(prefixList[0].rstrip("\n\r "))).replace("__","_") == re.sub(r'\d+', '', str(prefixList[1].rstrip("\n\r "))).replace("__","_"):
		output_file_name = re.sub(r'\d+', '', str(prefixList[0].rstrip("\n\r "))).replace("__","_") + "_COMBINED.CSV"
	else:
		output_file_name = "MERGED_OUTPUT.CSV"
	
	"""From https://stackoverflow.com/questions/2400504/easiest-way-to-replace-a-string-using-a-dictionary-of-replacements"""
	replace_dictionary = {"']['":'","', "['":'"', "']":'"', '"':"'", "', '":'","', "','":'","', "'\n'":'"\n"'}
	pattern = re.compile('|'.join(re.escape(key) for key in replace_dictionary.keys()))
	
	with open(output_file_name,"w") as f:
		joinedOutput = ''.join(single_CSV)
		#f.write('\"' + joinedOutput.replace("\'][\'",'\",\"').replace("[\'",'\"').replace("\']",'\"').replace('\"',"\'").replace("\', \'",'\",\"').replace("\',\'",'\",\"').replace("\'\n\'",'\"\n\"')[1:])		
		f.write(('\"' + pattern.sub(lambda x: replace_dictionary[x.group()], joinedOutput)[1:]).replace("', '",'","'))		

	with open("./GunShotMatch Output/FULL_" + output_file_name,"w") as f:
		joinedOutput = ''.join(single_CSV)
		#f.write('\"' + joinedOutput.replace("\'][\'",'\",\"').replace("[\'",'\"').replace("\']",'\"').replace('\"',"\'").replace("\', \'",'\",\"').replace("\',\'",'\",\"').replace("\'\n\'",'\"\n\"')[1:])		
		f.write(('\"' + pattern.sub(lambda x: replace_dictionary[x.group()], joinedOutput)[1:]).replace("', '",'","'))		

# functionality moved to counter
#	convert_to_xlsx(output_file_name),""
	
#	formatXLSX(output_file_name.lower()[:-3] + "xlsx")


def Match_Counter(prefixList):

	if len(prefixList) == 1:
		print "Require two or more samples to combine. Check ./list.txt or --samples and try again."
		print ""
		sys.exit(info())
	

	if re.sub(r'\d+', '', str(prefixList[0].rstrip("\n\r "))).replace("__","_") == re.sub(r'\d+', '', str(prefixList[1].rstrip("\n\r "))).replace("__","_"):
		input_file_name = re.sub(r'\d+', '', str(prefixList[0].rstrip("\n\r "))).replace("__","_") + "_COMBINED.CSV"
	else:
		input_file_name = "MERGED_OUTPUT.CSV"
	
	
	f = open("./"+input_file_name)		#Open merged output of lot of samples
	csv_f = csv.reader(f)

	csv_input = []
	
	for row in csv_f:
		csv_input.append(row)

	for i in range(len(csv_input)):
#		print i
#		raw_input(">")
#		if i == 0:
#			row_0 = csv_input[i]
#		if i == 1:
#			row_1 = csv_input[i]
		rows_appended = []
		try:
			"""From https://stackoverflow.com/questions/4843158/check-if-a-python-list-item-contains-a-string-inside-another-string"""
			if any("Page" in s for s in csv_input[i]): #csv_input[i][][:4] == "Page" or :
				#six_rows = {"Header": csv_input[i], "Hit 1":csv_input[i+1],"Hit 2":csv_input[i+2],"Hit 3":csv_input[i+3],"Hit 4":csv_input[i+4],"Hit 5":csv_input[i+5]}
				try:
					rows_appended = csv_input[i+1] + csv_input[i+2] + csv_input[i+3] + csv_input[i+4] + csv_input[i+5]
				except IndexError:
					print "#"
				"""From https://stackoverflow.com/questions/2600191/how-to-count-the-occurrences-of-a-list-item"""
				count_dict = Counter(rows_appended)
				try:
					for j in range(12):
						if len(csv_input[i+1][6+9*j]) > 0:
							csv_input[i+1][1+9*j] = str(count_dict[csv_input[i+1][6+9*j]])
						if len(csv_input[i+2][6+9*j]) > 0:
							csv_input[i+2][1+9*j] = str(count_dict[csv_input[i+2][6+9*j]])
						if len(csv_input[i+3][6+9*j]) > 0:
							csv_input[i+3][1+9*j] = str(count_dict[csv_input[i+3][6+9*j]])
						if len(csv_input[i+4][6+9*j]) > 0:
							csv_input[i+4][1+9*j] = str(count_dict[csv_input[i+4][6+9*j]])
						if len(csv_input[i+5][6+9*j]) > 0:
							csv_input[i+5][1+9*j] = str(count_dict[csv_input[i+5][6+9*j]])
				except IndexError:
					print ""

		except TypeError:
			next
			
	single_list = []
	#single_list.append("'" + ("','".join(row_0)))
	#single_list.append("'" + ("','".join(row_1)) + '"')
	for i in range(len(csv_input)):
		single_list.append("'" + ("','".join(csv_input[i])) + '"')#[:-1])

	
	"""From https://stackoverflow.com/questions/2400504/easiest-way-to-replace-a-string-using-a-dictionary-of-replacements"""
	replace_dictionary = {"\n'":'\n"',"','":'","',"', '":'", "',}
	pattern = re.compile('|'.join(re.escape(key) for key in replace_dictionary.keys()))
	
	with open(input_file_name[:-12] + "COUNTER.CSV","w") as f:
		joinedOutput = '\n'.join(single_list)

		f.write(('\"' + pattern.sub(lambda x: replace_dictionary[x.group()], str(joinedOutput))[1:]).replace("',",'",'))	
#		f.write('\"' + pattern.sub(lambda x: replace_dictionary[x.group()], joinedOutput)[1:])	
	
	convert_to_xlsx(input_file_name[:-12] + "COUNTER.CSV"),""
	
	formatXLSX((input_file_name[:-12] + "COUNTER.CSV").lower()[:-3] + "xlsx")
	
	return 0


if __name__ == '__main__':
	parser = argparse.ArgumentParser()
	
	try:	#Command line switches
		parser.add_argument("--info",help="Show program info.", action='store_true')
		parser.add_argument("-c","--combine",help="Combine GS and MS data for the samples", action='store_true')
		parser.add_argument("-s","--space",help="Space the samples", action='store_true')
		parser.add_argument("-m","--merge",help="Run Space, then Merge the files into one.", action='store_true')
		parser.add_argument("-v","--verbose",help="Turns on verbodity for diagnostics", action='store_true')
		parser.add_argument("-f","--samples",help="List of samples to run.",nargs='+')
		parser.add_argument("--dependencies",help="Check for and list dependencies.",action='store_true')
		parser.add_argument("--counter",help="Count the number of times a hit appears.",action='store_true')
		parser.add_argument("--spectra",help="Gather the spectra for the top 80 peaks.",action='store_true')
		parser.add_argument("--compare",help="Compare spectra for retention time specified.")
		args = parser.parse_args()
		
		if args.dependencies:
			check_dependencies()
		
			"""
GunShotMatch requires xlsxwriter and openpyxl.
To install them run the following command:
pip install XlsxWriter openpyxl
This will also install any dependencies. 
"""
		import xlsxwriter	# http://xlsxwriter.readthedocs.io/
		from openpyxl import  load_workbook	# https://openpyxl.readthedocs.io/en/default/
		from openpyxl.styles import Font, Fill, Alignment
		
		if args.info or len(sys.argv)==1:
			if not args.info:
				print "You must specify an option." 
			parser.print_help()	#show help and info
			info()
			sys.exit(0 if args.info else 1) 
		
		verbose = args.verbose
		verbosePrint(args)
		
		"""Gather the list of samples to run"""
		if args.samples:
			prefixList = args.samples
		elif os.path.isfile("list.txt"):
			with open("./list.txt","r") as f:
					prefixList = f.readlines()
		else:
			print("No samples specified. Please enter samples in list.txt or following the --samples argument\n.")
			sys.exit(1)
		
		if len(prefixList) == 0:
			print("No samples specified. Please enter samples in list.txt or following the --samples argument\n.")
			sys.exit(1)

		print prefixList

		"""Run Combine"""
		if args.combine:
			print "Combine"
			for prefix in prefixList:
				GCMScombine(prefix.rstrip("\n\r "))

		"""Run Spacer"""
		if args.space:
			print "Space"
			batch_spacer(prefixList)
		
		"""Run Merge"""
		"""This also runs Spacer"""
		if args.merge:
			print "Merge"
			Merge(prefixList)
			
		"""Run Counter"""
		if args.counter:
			print "Counter"
			Match_Counter(prefixList)
			
		"""Run get_mass_spectra"""
		if args.spectra:
			print "Get Mass Spectra"
			get_mass_spectra(prefixList)
		
		"""Run compare_mass_spectra"""
		if args.compare:
			compare_mass_spectra(prefixList, args.compare)
		
	except IndexError:	#if a sample is specified that doesn't exist
		traceback.print_exc()	#print the error and then
		parser.print_help()	#show help and info
		info()
		sys.exit(1)		
	

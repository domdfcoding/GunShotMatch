#  !/usr/bin/env python
#   -*- coding: utf-8 -*-
#
#  exporters.py
#
#  This file is part of GunShotMatch
#
#  Copyright © 2020 Dominic Davis-Foster <dominic@davis-foster.co.uk>
#
#  GunShotMatch is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 3 of the License, or
#  (at your option) any later version.
#
#  GunShotMatch is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#

# stdlib
import datetime
import os
import tempfile
from collections import Counter
from numbers import Number

# 3rd party
from domdf_python_tools.utils import as_text
from domdf_spreadsheet_tools import append_to_xlsx, format_header, format_sheet, make_column_property_dict
from mathematical.utils import rounders
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import PageBreak, Paragraph, Spacer, Table, TableStyle

# This package
from GuiV2.GSMatch2_Core.exporters import PDFExporterBase, styles, font_size


__version__ = "TODO"


class InfoPDFExporter(PDFExporterBase):
	no_input_filename_str = "&lt;Unsaved Project&gt;"
	no_input_full_filename_str = "an unsaved Project"

	def __init__(self, project, input_filename, output_filename, title="Project Information"):
		PDFExporterBase.__init__(self, input_filename, output_filename, title)
		
		self.project = project
		self.make_project_info_inner()
		self.build()
	
	def make_project_info_inner(self):
		project_info_col_widths = [self.inner_width * (2.5 / 7), self.inner_width * (4.5 / 7)]
		
		project_info = [["Name", self.project.name]]
		for prop in self.project.all_properties:
			if prop.type == datetime:
				value = prop.format_time()
			else:
				value = prop.value
			
			project_info.append([prop.label, value])
		
		self.elements.append(
			self.make_parameter_table("Project Information", project_info, False, project_info_col_widths))
		self.elements.append(PageBreak())
		
		# Experiments
		for experiment in self.project.experiment_objects:
			expr_info = []
			
			for prop in experiment.all_properties:
				expr_info.append([prop.label, prop.value])
			self.elements.append(
					self.make_parameter_table(f"Experiment: {experiment.name}", expr_info, False))
			self.elements.append(PageBreak())


class AlignmentPDFExporter(PDFExporterBase):
	def __init__(self, alignment_panel, input_filename, output_filename, title="Alignment Report"):
		# Parent: class:AlignmentDataPanel
		PDFExporterBase.__init__(self, input_filename, output_filename, title)
		
		self.alignment_panel = alignment_panel
		self.make_alignment_inner()
		self.build()
	
	def make_alignment_inner(self):
		self.elements.append(self.make_alignment_filter_table())
		self.elements.append(Spacer(1, cm/2))
		self.elements.append(self.make_alignment_table())
		
	def make_alignment_table(self):
		
		alignment_table_style = TableStyle([
				# (x, y)
				('FONTSIZE', (0, 0), (-1, -1), font_size),
				("SPAN", (1, 0), (-1, 0)),
				('ALIGN', (1, 0), (-1, 0), 'CENTER'),
				('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
				('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
				("LINEBELOW", (0, 0), (-1, 1), 1, colors.black),
				("LINEBELOW", (0, 2), (-1, -1), 0.5, colors.lightgrey),
				])
		
		top_row = ['', Paragraph("Retention Time (minutes)", styles["Center"])]
		header_row = [Paragraph("Peak No.", styles["Normal"])]
		rows = [top_row, header_row]
		
		for experiment in self.alignment_panel.project.experiment_name_list:
			header_row.append(Paragraph(experiment, styles["Right"]))
		
		for peak in self.alignment_panel.project.rt_alignment.itertuples():
			row_data = []
			
			for experiment in peak:
				rt = rounders(experiment, "0.00000")
				if rt.is_nan():
					rt = "-"
				row_data.append(rt)
			
			row_data[0] = peak.Index
			if self.alignment_panel.n_experiments - Counter(row_data)["-"] >= self.alignment_panel.filter_min_experiments:
				for experiment in row_data[1:]:
					if experiment != "-":
						if self.alignment_panel.filter_min_rt <= experiment <= self.alignment_panel.filter_max_rt:
							rows.append([Paragraph(str(x), styles["Right"]) for x in row_data])
							break  # As long as one of the peaks is in range, add the peak
		
		peak_no_prop = 1.5
		expr_prop = 2
		base_total = peak_no_prop + expr_prop*len(self.alignment_panel.project.experiment_name_list)
		peak_no_width = self.inner_width * (peak_no_prop / base_total)
		expr_width = self.inner_width * (expr_prop / base_total)
		
		col_widths = [peak_no_width] + [expr_width]*len(self.alignment_panel.project.experiment_name_list)
		
		return Table(
				rows,
				colWidths=col_widths,
				rowHeights=[None] * len(rows),
				style=alignment_table_style,
				hAlign="LEFT",
				repeatRows=2
				)
				
	def make_alignment_filter_table(self):
		if self.alignment_panel.filter_is_default():
			return Paragraph("<font size=12>This report was produced with all peaks shown.</font>", styles["Normal"])
		
		else:
			filter_table_style = TableStyle([
					# (x, y)
					('FONTSIZE', (0, 0), (-1, 0), 12),
					('FONTSIZE', (0, 1), (-1, -1), font_size),
					("SPAN", (0, 0), (-1, 0)),
					('ALIGN', (1, 0), (-1, 0), 'CENTER'),
					('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
					('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
					("LINEBELOW", (0, 0), (-1, 0), 1, colors.black),
					("LINEBELOW", (0, 1), (-1, -1), 0.5, colors.lightgrey),
					])
			
			return Table(
					[
							[
									Paragraph(
										"This report was produced with the following settings applied:",
										styles["Normal"]),
									'',
									],
							[
									Paragraph("Minimum Experiments:", styles["Normal"]),
									Paragraph(f"{self.alignment_panel.filter_min_experiments}", styles["Normal"]),
									],
							[
									Paragraph("Retention Time Range:", styles["Normal"]),
									Paragraph(
											f"{self.alignment_panel.filter_min_rt} - {self.alignment_panel.filter_max_rt} minutes",
											styles["Normal"]),
									],
							],
					colWidths=[5 * cm, 7 * cm],
					rowHeights=[None, None, None],
					style=filter_table_style,
					hAlign="LEFT",
					)


class ConsolidatePDFExporter(PDFExporterBase):
	def __init__(self, consolidate_panel, input_filename, output_filename, title="Consolidated Peaks"):
		PDFExporterBase.__init__(self, input_filename, output_filename, title, orient_landscape=True)
		
		self.consolidate_panel = consolidate_panel
		self.elements[-1] = Spacer(1, cm / 2)
		# self.elements.append(NextPageTemplate('landscape'))
		self.make_consolidate_inner()
		self.build()
	
	def make_consolidate_inner(self):
		# TODO: Option to respect filter options?
		#  Perhaps a dialog after choosing the filename? Or before?
		#  With options for all peaks, use current display, or custom, which has options from filter dialog?
		
		self.add_peak_compound(self.consolidate_panel.peak_list[0])
		
		# self.elements.append(PageBreak())
		
		for peak in self.consolidate_panel.peak_list[1:]:
			self.add_peak_compound(peak)
			
	def add_peak_compound(self, peak):
		
		rsd_width = self.inner_width * (1.9 / 25)
		pm_width = self.inner_width * (0.5 / 25)
		mf_width = self.inner_width * (1.5 / 25)
		
		col_widths = [
				self.inner_width * (0.5 / 25),
				
				self.inner_width * (1 / 25),  # Hit Num.
				pm_width,
				rsd_width,
				
				self.inner_width * (9 / 25),  # Name
				
				self.inner_width * (2.8 / 25),  # CAS
				
				mf_width,  # Match
				pm_width,
				rsd_width,
				
				mf_width,  # R Match
				pm_width,
				rsd_width,
				]
		
		# Final column width should be remaining width
		col_widths.append(self.inner_width - sum(col_widths))
		
		hits_style = TableStyle([
				# (x, y)
				# Whole Table
				('FONTSIZE', (0, 0), (-1, -1), font_size),
				('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
		
				# Lines
				("LINEAFTER", (0, 0), (-2, -1), 0.5, colors.lightgrey),
				("LINEBELOW", (0, 0), (-1, 1), 1, colors.black),
				("LINEABOVE", (0, 0), (-1, 0), 1, colors.black),
				("LINEBELOW", (0, 2), (-1, -1), 0.5, colors.lightgrey),
				
				# rt_area table
				('ALIGN', (0, 0), (0, -1), 'LEFT'),
				('LEFTPADDING', (0, 0), (-1, 0), 0),
				('RIGHTPADDING', (0, 0), (-1, 0), 0),
				("SPAN", (0, 0), (-1, 0)),
				
				# Hit Num
				("SPAN", (1, 1), (3, 1)),
				('ALIGN', (2, 1), (2, -1), 'CENTER'),
				('ALIGN', (1, 1), (1, 1), 'CENTER'),
				('RIGHTPADDING', (2, 1), (2, -1), 0),
				('LEFTPADDING', (2, 1), (2, -1), 0),
				
				# Match
				("SPAN", (6, 1), (8, 1)),
				('ALIGN', (7, 1), (7, -1), 'CENTER'),
				('ALIGN', (6, 1), (6, 1), 'CENTER'),
				('RIGHTPADDING', (7, 1), (7, -1), 0),
				('LEFTPADDING', (7, 1), (7, -1), 0),
				
				# R Match
				("SPAN", (9, 1), (11, 1)),
				('ALIGN', (10, 1), (10, -1), 'CENTER'),
				('ALIGN', (9, 1), (9, 1), 'CENTER'),
				('RIGHTPADDING', (10, 1), (10, -1), 0),
				('LEFTPADDING', (10, 1), (10, -1), 0),
				
				])
		
		rt_string = str(rounders(peak.rt / 60, '0.000000'))
		area_string = f"{rounders(peak.area, '0.000000'):,}"
		peak_no_string = str(peak.peak_number).rjust(4).replace(' ', '&nbsp;')
	
		rt_area_table_style = TableStyle([
				# (x, y)
				('FONTSIZE', (0, 0), (-1, -1), font_size),
				('ALIGN', (0, 0), (0, -1), 'LEFT'),
				('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
				('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
				("LINEAFTER", (2, 0), (2, -1), 0.5, colors.lightgrey),
				("LINEAFTER", (5, 0), (5, -1), 0.5, colors.lightgrey),
				("LINEAFTER", (8, 0), (8, -1), 0.5, colors.lightgrey),
				("LINEAFTER", (10, 0), (10, -1), 0.5, colors.lightgrey),
				("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.lightgrey),
				# ("LINEAFTER", (0, 0), (-1, -1), 0.5, colors.lightgrey),  # Diagnostic lines
				])
		
		rt_area_table_col_widths = [
				self.inner_width * (0.5 / 25) + pm_width + rsd_width,  # RT / Area label

				self.inner_width * (1 / 25),  # xbar
				self.inner_width * (5 / 25),  # mean
				pm_width, 	# ±
				rsd_width, 	# rsd
				
				col_widths[4] - self.inner_width * (5 / 25) - pm_width - rsd_width,  # spacer
				
				col_widths[5],  # Similarity / Peak No. label
				
				self.inner_width * (1 / 25),  # xbar / peak no.
				mf_width, 	# mean
				pm_width, 	# ±
				rsd_width,  # rsd
				]
		
		# End of No. Experiments column should align with end of `R Match x`
		rt_area_table_col_widths.append(sum(col_widths[:-1]) - sum(rt_area_table_col_widths))
		rt_area_table_col_widths.append(self.inner_width * (1 / 25)),  # No. Experiments value
		
		# Final column width should be remaining width
		rt_area_table_col_widths.append(self.inner_width - sum(rt_area_table_col_widths))
		
		rt_area_table_rows = [
						[
								Paragraph(f"Retention Time:", styles["Normal"]),
								Paragraph(f"{self.xbar} =", styles["Right"]),
								Paragraph(f"{rt_string} minutes", styles["Right"]),
								Paragraph(f"±", styles["Center"]),
								Paragraph(f"{peak.rt_stdev / peak.rt:.2%}", styles["Right"]),
								'',
								Paragraph(f"Similarity:", styles["Normal"]),
								Paragraph(f"{self.xbar} =", styles["Right"]),
								Paragraph(f"{peak.average_ms_comparison:.1f}", styles["Right"]),
								Paragraph(f"±", styles["Center"]),
								Paragraph(f"{peak.ms_comparison_stdev / peak.average_ms_comparison:3.2%}", styles["Right"]),
								'',	'', '',
								],
						[
								Paragraph(f"Peak Area:", styles["Normal"]),
								Paragraph(f"{self.xbar} =", styles["Right"]),
								Paragraph(f"{area_string}", styles["Right"]),
								Paragraph(f"±", styles["Center"]),
								Paragraph(f"{peak.area_stdev / peak.area:.2%}", styles["Right"]),
								'',
								Paragraph(f"Peak Number:", styles["Normal"]),
								Paragraph(f"{peak_no_string}", styles["Right"]),
								'', '', '',
								Paragraph(f"No. Experiments:", styles["Normal"]),
								Paragraph(f"{len(peak)}", styles["Right"]),
								'',
								]
						]
		
		align_repeat = dict(
				hAlign="LEFT",
				repeatRows=2
				)
		
		rows = [
				[
						Table(
								rt_area_table_rows,
								colWidths=rt_area_table_col_widths,
								style=rt_area_table_style,
								rowHeights=[None, None],
								**align_repeat,
								),
						'', '', '', '', '',
						],
				[
						'',
						Paragraph(f"Hit Num. {self.xbar}", styles["Center"]), '', '',
						Paragraph("Name", styles["Normal"]),
						Paragraph("CAS", styles["Center"]),
						Paragraph(f"Match {self.xbar}", styles["Center"]), '', '',
						Paragraph(f"R Match {self.xbar}", styles["Center"]), '', '',
						Paragraph("Freq.", styles["Center"]),
						]
				]
		
		hit_list = list(enumerate(peak.hits))
		
		def make_stats_para(mean, stdev):
			rsd = self.convert_spaces(f"{stdev / mean:3.2%}", 7)
			
			return [
					Paragraph(f"{mean:.1f}", styles["Right"]),
					Paragraph("±", styles["Center"]),
					Paragraph(f"{rsd}", styles["Right"]),
					]

		for hit_number, hit in hit_list:
			
			rows.append([
					Paragraph(f'{hit_number + 1}', styles["Center"]),
					*make_stats_para(hit.average_hit_number, hit.hit_number_stdev),
					Paragraph(f"{hit.name}", styles["Normal"]),
					Paragraph(f"{hit.cas}", styles["Center"]),
					*make_stats_para(hit.match_factor, hit.match_factor_stdev),
					*make_stats_para(hit.reverse_match_factor, hit.reverse_match_factor_stdev),
					Paragraph(f"{len(hit)}", styles["Center"]),
					])
		
		self.elements.append(Table(
				rows,
				colWidths=col_widths,
				rowHeights=[None] * len(rows),
				style=hits_style,
				**align_repeat,
				))
		
		self.elements.append(Spacer(1, cm / 3))


class CSVExporterBase:
	def __init__(self, minutes=True):
		self.minutes = minutes
		self.csv = None
	
	def open(self, output_filename):
		"""
		Open a csv file for writing
		
		:param output_filename: The filename to save the CSV file as
		:type output_filename: str or pathlib.Path
		"""
		
		self.csv = open(output_filename, "w")
		
	def close(self):
		"""
		Close the csv file
		"""
		
		self.csv.close()
	
	def convert_rt(self, rt):
		"""
		Converts the retention time to minutes if applicable

		:param rt: The retention time to convert
		:type rt: Number
		"""
		
		if self.minutes:
			if isinstance(rt, (list, tuple)):
				return [seconds / 60 for seconds in rt]
			elif isinstance(rt, Number):
				return rt / 60
		else:
			return rt
	
	def do_write_statistics(self, file_object, mean, stdev):
		"""
		Append a set of statistics to the csv file

		:param file_object: The csv file to write to
		:type file_object: io.TextIOWrapper
		:param mean: The mean (average) value
		:type mean: Number
		:param stdev: The standard deviation value
		:type stdev: Number
		"""
		
		self.do_write_value(file_object, mean)
		self.do_write_value(file_object, stdev)
		self.do_write_value(file_object, stdev / mean)
		self.do_empty_column(file_object)

	def write_statistics(self, mean, stdev):
		"""
		Append a set of statistics to the csv file

		:param mean: The mean (average) value
		:type mean: Number
		:param stdev: The standard deviation value
		:type stdev: Number
		"""
		
		self.do_write_statistics(self.csv, mean, stdev)

	@staticmethod
	def do_newline(file_object):
		"""
		Append a newline character (\n) to the csv file

		:param file_object: The csv file to write to
		:type file_object: io.TextIOWrapper
		"""
		
		file_object.write("\n")

	def newline(self):
		"""
		Append a newline character (\n) to the csv file
		"""
		
		self.do_newline(self.csv)
	
	@staticmethod
	def do_empty_column(file_object, n=1):
		"""
		Append an empty column given value to the csv file

		:param file_object: The csv file to write to
		:type file_object: io.TextIOWrapper
		:param n: The number of columns to append
		:type n: int
		"""
		
		file_object.write(";"*n)
	
	def empty_column(self, n=1):
		"""
		Append an empty column given value to the csv file

		:param n: The number of columns to append
		:type n: int
		"""

		self.do_empty_column(self.csv, n)
	
	@staticmethod
	def do_write_value(file_object, value=''):
		"""
		Append the given value to the csv file

		:param file_object: The csv file to write to
		:type file_object: io.TextIOWrapper
		:param value: The value to write to the file
		:type value: any
		"""
		
		file_object.write(f"{value};")
	
	def write_value(self, value=''):
		"""
		Append the given value to the csv file
		
		:param value: The value to write to the file
		:type value: any
		"""
		
		self.do_write_value(self.csv, value)
	
	def write_values(self, values):
		"""
		Append a sequence of values to the csv file

		:param values: The sequence of values to write to the file
		:type values: iterable
		"""

		for value in values:
			self.do_write_value(self.csv, value)
			

class MatchesCSVExporter(CSVExporterBase):
	"""
	Export the Matches data to a CSV file
	"""
	
	def __init__(self, project, output_filename, minutes=True, n_hits=5):
		"""
		:param project:
		:type project:
		:param output_filename: The filename to save the CSV file as
		:type output_filename: str or pathlib.Path
		:param minutes: Whether the retention times should be given in minutes. Default True
		:type minutes: bool, optional
		:param n_hits: The number of hits to report for each peak. Default 5
		:type n_hits: int, optional
		"""
		
		CSVExporterBase.__init__(self, minutes)
		
		self.open(output_filename)
		
		# Write the names of the project and the experiments
		self.write_name(project.name)
		
		for experiment_name in project.experiment_name_list:
			self.write_name(experiment_name)
		
		self.newline()
		
		# Number of experiments
		n_expr = len(project.experiment_name_list)
		
		# Write the column headers
		self.csv.write(
				"Retention Time;Peak Area;Peak No.;" +
				";RT;Area;" * n_expr +
				";Match Factor;;;;Reverse Match Factor;;;;Hit Number;;;;Retention Time;;;;Peak Area;;\n")
		self.csv.write(
				"Name;CAS Number;Freq.;" +
				"Hit No.;Match Factor;Reverse MF;" * n_expr +
				";Mean;STDEV;%RSD;;Mean;STDEV;%RSD;;Mean;STDEV;%RSD;;Mean;STDEV;%RSD;;Mean;STDEV;%RSD\n")
		
		for peak in project.consolidated_peaks:
			self.write_values([self.convert_rt(peak.rt), peak.area, peak.peak_number])
			
			for rt, area in zip(self.convert_rt(peak.rt_list), peak.area_list):
				self.write_values(['', rt, area])
			
			self.empty_column(13)
			
			# Retention Time Statistics
			self.write_statistics(self.convert_rt(peak.rt), self.convert_rt(peak.rt_stdev))
			
			# Peak Area Statistics
			self.write_statistics(peak.area, peak.area_stdev)
			
			self.newline()

			for hit in peak.hits[:n_hits]:
				self.write_values([hit.name, hit.cas, len(hit)])
				
				for hit_no, mf, rmf in zip(hit, hit.mf_list, hit.rmf_list):
					self.write_values([hit_no, mf, rmf])
				
				self.empty_column()
				
				# Match Factor Statistics
				self.write_statistics(hit.match_factor, hit.match_factor_stdev)
	
				# Reverse Match Factor Statistics
				self.write_statistics(hit.reverse_match_factor, hit.reverse_match_factor_stdev)
				
				# Hit Number Statistics
				self.write_statistics(hit.average_hit_number, hit.hit_number_stdev)
				
				self.newline()
		
		self.close()
	
	def write_name(self, name):
		"""
		Write the given name to the csv file, followed by two empty columns
		
		:type name: str
		"""
		
		self.write_value(name)
		self.empty_column(2)


class StatisticsCSVExporter(CSVExporterBase):
	"""
	Export the statistics to a CSV file
	"""
	
	def __init__(self, project, output_filename, minutes=True, peak_filter=None):
		"""
		:param project:
		:type project:
		:param output_filename: The filename to save the CSV file as
		:type output_filename: str or pathlib.Path
		:param minutes: Whether the retention times should be given in minutes. Default True
		:type minutes: bool, optional
		"""
		
		CSVExporterBase.__init__(self, minutes)
		
		self.open(output_filename)
		
		statistics_header = ";;;;".join([
				project.name, "Retention Time", "Peak Area", "Match Factor",
				"Reverse Match Factor", "Hit Number", "MS Comparison",
				])
		
		statistics_header += "\nName;CAS Number;n;;"
		statistics_header += "Mean;STDEV;%RSD;;"*6
		statistics_header += "\n"
		
		self.csv.write(statistics_header)
		
		for peak in project.consolidated_peaks:
			if peak_filter:
				if not all([
						# Number of samples peak must be present in
						len(peak.hits[0]) >= peak_filter.min_samples,
						
						# Average match factor between samples
						peak.average_ms_comparison > peak_filter.min_similarity,
						
						# Average match factor of samples compared to reference spectrum
						peak.hits[0].match_factor > peak_filter.min_match_factor,
						
						# Average reverse match factor of samples compared to reference spectrum
						peak.hits[0].reverse_match_factor > peak_filter.min_reverse_match_factor,
						
						# Compound in list of those previously reported as being in propellant or OGSR
						peak_filter.check_cas_number(peak.hits[0].cas),
						
						peak_filter.check_rt(peak.rt),  # Retention Time
						peak_filter.check_area(peak.area),  # Peak Area
						]):
					continue
			
			self.write_peak(peak)
				
		self.close()
	
	def write_peak(self, peak):
		"""
		Write the given peak and mass spectrum to the file

		:param peak: The peak to write to file
		:type peak:
		"""
		
		hit = peak.hits[0]
		self.write_value(hit.name)
		self.write_value(hit.cas)
		self.write_value(len(hit))
		self.empty_column()
		
		self.write_statistics(self.convert_rt(peak.rt), self.convert_rt(peak.rt_stdev))
		self.write_statistics(peak.area, peak.area_stdev)
		self.write_statistics(hit.match_factor, hit.match_factor_stdev)
		self.write_statistics(hit.reverse_match_factor, hit.reverse_match_factor_stdev)
		self.write_statistics(hit.average_hit_number, hit.hit_number_stdev)
		self.write_statistics(peak.average_ms_comparison, peak.ms_comparison_stdev)
		
		self.newline()


class XLSXExporterBase:
	def __init__(self, project, output_filename, minutes=True, peak_filter=None):
		"""
		
		:param project:
		:type project:
		:param output_filename: The filename to save the spreadsheet as
		:type output_filename: str or pathlib.Path
		:param minutes:
		:type minutes:
		:param peak_filter:
		:type peak_filter:
		"""
		
		self.project = project
		self.n_experiments = len(self.project.experiment_name_list)
		self.output_filename = output_filename
		
		print('\nGenerating XLSX Output...')

		# TODO: Make domdf_spreadsheet_tools.append_to_xlsx return the workbook at the end of the function
	
	def load_workbook(self):
		self.wb = load_workbook(self.output_filename)
	
	def format_statistics_sheet(self, sheet_name="Statistics"):
		"""
		Formatting for Statistics Sheet
		"""
		
		# Statistics
		number_format_list = {
				'C': '0', 'E': '0.000', 'F': '0.000000', 'G': '0.00%', 'I': '0.00', 'J': '0.00', 'K': '0.00%',
				'M': '0.0', 'N': '0.0000', 'O': '0.00%', 'Q': '0.0', 'R': '0.0000', 'S': '0.00%', 'U': '0.0',
				'V': '0.0000', 'W': '0.00%', 'Y': '0.0', 'Z': '0.000', 'AA': '0.000%'
				}
		width_dict = {
				"B": 12, "C": 2, "D": 1, "E": 8, "F": 11, "G": 8, "H": 1, "I": 13, "J": 12, "K": 10, "L": 1, "M": 8,
				"N": 10, "O": 10, "P": 1, "Q": 8, "R": 10, "S": 10, "T": 1, "U": 6, "V": 8, "W": 10, "X": 1, "Y": 8,
				"Z": 9, "AA": 9
				}
		alignment_list = {"B": "center", "C": "center"}
		
		ws = self.wb[sheet_name]
		format_sheet(ws, number_format_list, width_dict, alignment_list)
		for offset in range(6):
			ws.merge_cells(
					get_column_letter(5 + (4 * offset)) + '1:' + get_column_letter(7 + (4 * offset)) + '1')
		
		format_header(ws, make_column_property_dict({"2": "center"}, repeat=27, length=1), 1, 2)
		
	def save_workbook(self):
		# Save the file
		self.wb.save(self.output_filename)
		return

	def format_combined_sheet(self):
		# GC-MS
		ws = self.wb["GC-MS"]
		number_format_list = make_column_property_dict(
				{'1': '0.000', '2': '0.0', '3': '0', '5': '0', '6': '0'},
				repeat=self.n_experiments,
				length=9
				)
		width_dict = make_column_property_dict(
				{"1": 14, "2": 12, "4": 0, "5": 9, "6": 9, "8": 15, "9": 10},
				repeat=self.n_experiments,
				length=9
				)
		alignment_list = make_column_property_dict({"5": "center", "6": "center", "8": "center"})
		format_sheet(ws, number_format_list, width_dict, alignment_list)
		
		for offset in range(self.n_experiments):
			merge_string = get_column_letter(1 + (9 * offset)) + '1:' + get_column_letter(9 + (9 * offset)) + '1'
			ws.merge_cells(merge_string)
		
		format_header(ws, make_column_property_dict(
				{"1": "center", "2": "center", "5": "center", "6": "center", "8": "center", "9": "center"},
				repeat=self.n_experiments,
				length=9
				), 1, 2)
	
	def format_matches_sheet(self):
		ws = self.wb["Matches"]
		header_number_format_dict = make_column_property_dict(
				{'5': '0.000', '6': '0.00'}, {"A": "0.000", "B": "0.00"},
				{"17": "0.000", "18": '0.000000', "19": '0.00%', "21": "0.00", "22": "0.00", "23": '0.00%'},
				repeat=self.n_experiments,
				length=3
				)
		header_h_alignment_list = make_column_property_dict(
				{'5': "right", '6': "right", }, {"A": "right", "B": "right"},
				{"17": "right", "18": "right", "19": "right", "21": "right", "22": "right", "23": "right"},
				repeat=self.n_experiments,
				length=3
				)
		# header_v_alignment_list = make_column_property_dict(
		# 	{'5': "center", '6': "center"},
		# 	{"A": "center", "B": "center", "AF": "center", "AG": "center", "AH": "center", "AJ": "center",
		# 	 "AK": "center", "AL": "center"})
		hits_number_format_dict = make_column_property_dict(
				{'4': '0', '5': '0', '6': '0'}, {"C": "0"},
				{
						"5": "0.0", "6": "0.0000", "7": "0.00%", "9": "0.0", "10": "0.0000", "11": "0.00%",
						"13": "0.0",
						"14": "0.0000", "15": "0.00%"
						},
				repeat=self.n_experiments,
				length=3
				)
		hits_alignment_dict = make_column_property_dict(
				{'4': 'center', '5': 'center', '6': 'center'},
				{"B": "center", "C": "center"},
				{"13": "center"},
				repeat=self.n_experiments,
				length=3
				)
		width_dict = make_column_property_dict(
				{"4": 8, "5": 8, "6": 11},
				{"B": 12},
				{
						"5": 6, "6": 9, "7": 9, "9": 6, "10": 9, "11": 9, "13": 5, "14": 7,
						"15": 9, "17": 8, "18": 9, "19": 9, "21": 14, "22": 12, "23": 9,
						"4": 1, "8": 1, "12": 1, "16": 1, "20": 1
						},
				repeat=self.n_experiments,
				length=3
				)
		
		for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
			for cell in row:
				cell.alignment = Alignment(vertical="center", wrap_text=False)
		
		headers = [x + 1 for x in range(ws.max_row)][3::6]
		
		for column_cells in ws.columns:
			length = max(len(as_text(cell.value)) for cell in column_cells)
			ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length
		# ws.column_dimensions[column_cells[0].column].bestFit = True
		
		for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
			for cell in row:
				# Column Headers
				if cell.row in headers:
					if header_number_format_dict and (cell.column in header_number_format_dict):
						cell.number_format = header_number_format_dict[cell.column]
					if header_h_alignment_list and (cell.column in header_h_alignment_list):
						cell.alignment = Alignment(
								horizontal=header_h_alignment_list[cell.column],
								vertical="bottom",
								wrap_text=False
								)
				# Hit Data
				else:
					if hits_number_format_dict and (cell.column in hits_number_format_dict):
						cell.number_format = hits_number_format_dict[cell.column]
					if hits_alignment_dict and (cell.column in hits_alignment_dict):
						cell.alignment = Alignment(
								horizontal=hits_alignment_dict[cell.column],
								vertical="center",
								wrap_text=False
								)
		if width_dict:
			for column in width_dict:
				if width_dict[column] == 0:
					ws.column_dimensions[column].hidden = True
				else:
					ws.column_dimensions[column].width = width_dict[column]
		
		format_header(ws, make_column_property_dict(
				{"4": "center"},
				repeat=self.n_experiments,
				length=3
				))  # First Row
		format_header(ws, make_column_property_dict(
				{"2": "center"},
				{"A": "right"},
				repeat=((self.n_experiments * 3) + 25),
				length=1
				), 2, 2)  # Second Row
		format_header(ws, make_column_property_dict(
				{"2": "center"},
				repeat=((self.n_experiments * 3) + 25),
				length=1
				), 3, 3)  # Third Row
		
		for offset in range(self.n_experiments):
			merge_string = get_column_letter(4 + (3 * offset)) + '1:' + get_column_letter(6 + (3 * offset)) + '1'
			ws.merge_cells(merge_string)
		
		offset = self.n_experiments * 3
		for index in [5, 9, 13, 17, 21]:
			merge_string = get_column_letter(index + offset) + '2:' + get_column_letter(index + 2 + offset) + '2'
			ws.merge_cells(merge_string)
	
	def make_contents_page(self):
		"""Contents Page"""
		ws = self.wb.create_sheet("Index", 0)
		contents = [
				[
						"GC-MS",
						"Combined GC-MS data aligned by retention time.",
						True
						],
				[
						"Matches",
						"List of possible matching compounds for each retention time, based on all samples.",
						True
						],
				[
						"Statistics for the top hit for each retention time.",
						True
						],
				]
		
		print("\nThe worksheets in the output xlsx file are as follows:")
		ws.append(["GunShotMatch Version {} Output".format(__version__)])
		for row in contents:
			print(row[0] + " " * (24 - len(row[0])) + row[1])
			if row[2]:
				ws.append(["", '=HYPERLINK("#\'{0}\'!A1","{0}")'.format(row[0]), row[1]])
			else:
				ws.append(["", "{0}".format(row[0]), row[1]])
		
		ws.column_dimensions["B"].width = 50.0
		for row in ws.iter_rows(min_row=2, max_row=len(contents) + 1, min_col=2, max_col=2):
			for cell in row:
				if cell.value:
					if cell.value.startswith("=HYPERLINK"):
						cell.font = Font(
								color="0000EE",
								underline="single"
								)
		
		# TODO:
		# ws.cell(column=4, row=2).value = "Sample list:"
		# for index, prefix in enumerate(prefixList):
		# 	ws.cell(row=2, column=(5+index)).value = prefix


class StatisticsXLSXExporter(XLSXExporterBase):
	def __init__(self, project, output_filename, minutes=True, peak_filter=None):
		
		XLSXExporterBase.__init__(self, project, output_filename, minutes, peak_filter)
		
		with tempfile.TemporaryDirectory() as tempdir:
			csv_file = os.path.join(tempdir, "statistics.csv")
			
			StatisticsCSVExporter(project, csv_file, minutes, peak_filter)
			
			append_to_xlsx(
					csv_file,
					output_filename,
					"Statistics",
					separator=";",
					toFloats=True,
					overwrite=True
					)
		
		# TODO: Make domdf_spreadsheet_tools.append_to_xlsx return the workbook at the end of the function
		self.load_workbook()
		self.format_statistics_sheet()
		self.save_workbook()


class MatchesXLSXExporter(XLSXExporterBase):
	"""
	Export the matches data to an XLSX spreadsheet
	"""
	
	def __init__(self, project, output_filename, minutes=True, n_hits=5):
		"""
		
		:param project:
		:type project:
		:param output_filename: The filename to save the spreadsheet as
		:type output_filename: str or pathlib.Path
		:param minutes:
		:type minutes:
		:param n_hits:
		:type n_hits:
		"""
		
		XLSXExporterBase.__init__(self, project, output_filename, minutes, n_hits)
		
		with tempfile.TemporaryDirectory() as tempdir:
			csv_file = os.path.join(tempdir, "matches.csv")
			
			MatchesCSVExporter(project, csv_file, minutes, n_hits)
			
			append_to_xlsx(
					csv_file,
					output_filename,
					"Matches",
					separator=";",
					toFloats=True,
					overwrite=True
					)
		
		# TODO: Make domdf_spreadsheet_tools.append_to_xlsx return the workbook at the end of the function
		self.load_workbook()
		self.format_statistics_sheet()
		self.save_workbook()

# TODO: CombinedCSVExporter
# class CombinedXLSXExporter(XLSXExporterBase):
# 	def __init__(self, project, output_filename, minutes=True, peak_filter=None):
#
# 		XLSXExporterBase.__init__(self, project, output_filename, minutes, peak_filter)
#
# 		with tempfile.TemporaryDirectory() as tempdir:
# 			csv_file = os.path.join(tempdir, "combined.csv")
#
# 			CombinedCSVExporter(project, csv_file, minutes, peak_filter)
#
# 			append_to_xlsx(
# 					csv_file,
# 					output_filename,
# 					"combined",
# 					separator=";",
# 					toFloats=True,
# 					overwrite=True
# 					)
#
# 		# TODO: Make domdf_spreadsheet_tools.append_to_xlsx return the workbook at the end of the function
# 		self.load_workbook()
# 		self.format_statistics_sheet()
# 		self.save_workbook()


class CompleteXLSXExporter(XLSXExporterBase):
	def __init__(self, project, output_filename, minutes=True, peak_filter=None, matches_n_hits=5):

		XLSXExporterBase.__init__(self, project, output_filename, minutes, peak_filter)

		with tempfile.TemporaryDirectory() as tempdir:
			# Combined Sheet
			combiend_csv_file = os.path.join(tempdir, "combined.csv")
#
# 			CombinedCSVExporter(project, csv_file, minutes, peak_filter)
#
# 			append_to_xlsx(
# 					csv_file,
# 					output_filename,
# 					"combined",
# 					separator=";",
# 					toFloats=True,
# 					overwrite=True
# 					)
#
			# Matches Sheet
			matches_csv_file = os.path.join(tempdir, "matches.csv")
			
			MatchesCSVExporter(project, matches_csv_file, minutes, matches_n_hits)
			
			append_to_xlsx(
					matches_csv_file,
					output_filename,
					"Matches",
					separator=";",
					toFloats=True,
					)
			
			# Statistics Sheet
			statistics_csv_file = os.path.join(tempdir, "statistics.csv")
			
			StatisticsCSVExporter(project, statistics_csv_file, minutes, peak_filter)
			
			append_to_xlsx(
					statistics_csv_file,
					output_filename,
					"Statistics",
					separator=";",
					toFloats=True,
					)
		
		self.load_workbook()
		# self.format_combined_sheet()
		self.format_matches_sheet()
		self.format_statistics_sheet()
		
		self.make_contents_page()
		
		self.save_workbook()


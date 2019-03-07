#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
#  GSM_plot.py
program_name ="GunShotMatch Comparison Plot"
_version = '0.2'
copyright = 2018

#  Copyright 2018 Dominic Davis-Foster <dominic@davis-foster.co.uk>
#  
#  This program is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 2 of the License, or
#  (at your option) any later version.
#  
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#  
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#  
#  

# adapted from https://stackoverflow.com/questions/48191792/scatter-plot-with-multiple-y-values-with-line-for-each-category-x-label
# https://matplotlib.org/tutorials/intermediate/legend_guide.html
# https://matplotlib.org/devdocs/gallery/text_labels_and_annotations/custom_legends.html

samples_to_compare = [
	("UNIQUE_SUBTRACT_20180214203141","Unique"),
#	("ELEY_SUBTRACT_20180214195019","Eley Contact"),
#	("ELEY_CASE_SUBTRACT_20180214195713","Eley Case"),
#	("WINCHESTER_SUBTRACT_20180214193908","Winchester Pistol"),
#	("WINCHESTER_CASE_SUBTRACT_20180214194606","Winchester Case"),
#	("GECO_SUBTRACT_20180214200701","Geco"),
#	("GECO_CASE_SUBTRACT_20180214201631","Geco Case"),
	("ELEY_SHOTGUN_SUBTRACT_20180214202526","Eley Hawk"),
	] #must be in order you want them on the graph

# Options
show_outliers = True
show_raw_data = False
use_mean = True
use_median = False
#mode = "sides" # sides/top
comparison = True
outlier_mode = "2stdev" #mad,2stdev,spss,quartile		SPSS not currently working
leg_cols = 1 # Number of columns for legend

if use_mean and use_median:
	raise ValueError("Cannot specify both use_mean and use_median at the same time")
elif not use_mean and not use_median:	
	raise ValueError("Must specify either use_mean or use_median")
	
metric = "mean" if use_mean else "median"

from utils import timing # Times the program
import sys, os

from utils.helper import clear, check_dependencies
clear()		#clear the display

missing_modules = check_dependencies(["progressbar", "openpyxl", "matplotlib"], prt=False)
if len(missing_modules) > 0:
	for mod in missing_modules:
		print("{} is not installed. Please install it and try again".format(mod))
		sys.exit(1)
		
import progressbar

print("""\r{0} Version {1} is loading. Please wait...
Copyright {2} Dominic Davis-Foster""".format(program_name,_version,copyright,))

"""Imports"""
if "-h" not in str(sys.argv):
	bar = progressbar.ProgressBar(max_value=24) #progressbar for imports

	import time ; time.sleep(0.1); bar.update(1)
	import argparse	; time.sleep(0.1); bar.update(2)
#	import traceback ; time.sleep(0.1); bar.update(3)
	import warnings	; time.sleep(0.1); bar.update(4)
	import tarfile	; time.sleep(0.1); bar.update(5)
	import numpy	; time.sleep(0.1); bar.update(7)
	import math		; time.sleep(0.1); bar.update(8)
	import ConfigParser	; time.sleep(0.1); bar.update(9)
	import shutil	; time.sleep(0.1); bar.update(10)
	import atexit	; time.sleep(0.1); bar.update(11)	
	import operator	; time.sleep(0.1); bar.update(12)	
	import itertools; time.sleep(0.1); bar.update(13)
	
	from scipy import stats	; time.sleep(0.1); bar.update(14)

	from openpyxl import  Workbook, worksheet, load_workbook	# https://openpyxl.readthedocs.io/en/default/
	time.sleep(0.1); bar.update(15)
	from openpyxl.utils import get_column_letter
	time.sleep(0.1); bar.update(17)

	from utils.helper import entry, mean_none, std_none, copytree
	time.sleep(0.1); bar.update(18)
	from utils import load_config, within1min, outliers
	
	import matplotlib; matplotlib.use("Cairo")
	time.sleep(0.1); bar.update(19)
	
	import matplotlib.pyplot as plt	; time.sleep(0.1); bar.update(20)
	import matplotlib.ticker as ticker; time.sleep(0.1); bar.update(21)
	
	from matplotlib.lines import Line2D; time.sleep(0.1); bar.update(22)
else:
	import argparse
	import warnings


"""From https://stackoverflow.com/questions/21129020/how-to-fix-unicodedecodeerror-ascii-codec-cant-decode-byte
Big shout out to fisherman https://stackoverflow.com/users/2309581/fisherman"""
reload(sys)  
sys.setdefaultencoding('utf8')
time.sleep(0.1); bar.update(23)


def comparison():
	
	styles = ["D", #diamond
	"s", #square
	"X", #bold cross
	"^", #triangle
	"d", #diamond
	"h", #hexagon
	"o", #dot
	"v", #down triangle
	"<", #left triangle
	">"] #right triangle
		
	#colours = ['b', 'g', 'r', 'c', 'm', 'y', 'k']
	colours = [
		"DarkRed",
		"DeepSkyBlue",	
		"Green",
		"Purple",
		"Black",
		"Sienna",
		"MediumTurquoise",
		"DarkOliveGreen",
		"m",
		"DarkSlateBlue",
		"OrangeRed",
		"CadetBlue",
		"Olive",
		"Red",
		"Blue",
		"PaleVioletRed",
		"Teal",
		"y",
		"RoyalBlue",
	]
	
	
	
	import os
	for directory in ["./.temp", "charts"]:
		if not os.path.exists(directory):
			os.makedirs(directory)
	
	sample_data = {}
	all_sample_names = []
	batch_names = {}
	
	for index, sample in enumerate(samples_to_compare):	
		sample = sample[0]
		print sample
		
		if not os.path.isfile("./.temp/{0}/{0}_FINAL.xlsx".format(sample[:-15])):
			"""from https://stackoverflow.com/questions/31163668/how-do-i-extract-a-tar-file-using-python-2-4"""
			tar = tarfile.open("{}.tar.gz".format(sample))
			#tar.extractall(path='./.temp/{}'.format(sample[:-15]))
			tar.extract("{}_FINAL.xlsx".format(sample[:-15]), path='./.temp/{}'.format(sample[:-15]))
			tar.close()
	
		"""From http://openpyxl.readthedocs.io/en/default/"""
		wb = load_workbook("./.temp/{0}/{0}_FINAL.xlsx".format(sample[:-15]))
		
		sample_names = []
		
		#Get samples list
		ws = wb["Index"]
		for row in ws.iter_rows('F2:AF2'):
			for cell in row:
				if cell.value:
					if len(cell.value) > 0:
						sample_names.append(cell.value)
		
		#get peak areas
		ws = wb["Statistics - Lit Only"]
		for row_index, row in enumerate(ws.iter_rows('A2:L{}'.format(ws.max_row+1))):
			if row_index > 1:
				for ind, samp in enumerate(sample_names):
					sample_data.setdefault(ws.cell("A{}".format(row_index+1)).value,[]).append((sample_names[ind],ws.cell("{}{}".format(get_column_letter(29+ind),row_index+1)).value))
					
		all_sample_names += sample_names
		batch_names[sample] = sample_names

	for compound in sample_data:
		temp_samples = []
		temp_output = []
		
		for samp in sample_data[compound]:
			if any(x[0] == samp[0] and x != samp for x in sample_data[compound]):		
				peak_data = [x[1] for x in sample_data[compound] if x[0] == samp[0]]
				
				if (sample_names[ind],max(peak_data)) not in temp_output:
					temp_output.append((sample_names[ind],max(peak_data)))
			else:
				temp_output.append(samp)
		
		if len(temp_output) > 0:
			sample_data[compound] = temp_output
			
	for compound in sample_data:
		temp_samples = []

	# where a compound is not present in samples, write "none"
	for compound in sample_data:
		for sample_name in all_sample_names:
			if not any(sample_name in data for data in sample_data[compound]):
				sample_data.setdefault(compound,[]).append((sample_name,None))
	
	point_dict = {}
	error_dict = {}
	outlier_dict = {}
	extremes_dict = {}
	raw_data_dict = {}
	
	any_outliers = False # Are there actually any outliers?
	any_extremes = False # Are there actually any extreme values?
	
	#print CSV to terminal
#	print ';'.join(['','Mean'] + ['']*(len(samples_to_compare)-1)  + ['Stdev'] + ['']*(len(samples_to_compare)-1) + ['Median'] + ['']*(len(samples_to_compare)-1)  + ['Diff'] + ['']*(len(samples_to_compare)-1) )
#	print ';'.join([''] + [x[1] for x in samples_to_compare] + [x[1] for x in samples_to_compare] + [x[1] for x in samples_to_compare] + [x[1] for x in samples_to_compare])
	
	for compound in sample_data:
			#point_list = []
			mean_list = []
			median_list = []
			error_list = []
			outlier_list = []
			extremes_list = []
			raw_data_list = []
			diff_list = []
			
			for batch in samples_to_compare:
				batch = batch[0]
					
				temp_data = []
				for sample_name in batch_names[batch]:
					for samp in sample_data[compound]:
						if samp[0] == sample_name:
							temp_data.append(samp[1])
				
				if show_outliers:
					#detect outliers
				#	print compound, batch
				#	print temp_data
			
					if outlier_mode == "mad":				
						x = outliers.mad_outliers(temp_data)
					elif outlier_mode == "quartiles":
						x = outliers.quartile_outliers(temp_data)
					elif outlier_mode == "2stdev":
						x = outliers.stdev_outlier(temp_data,2) #outlier classed as more than 2 stdev away from mean
					elif outlier_mode == "spss":
						x = outliers.spss_outliers(temp_data)

				#	print("{}: {}".format(outlier_mode, x))
									
					if type(x) == tuple and len(x) == 3:
						print x
						
						if len(x[0]) > 1:
							extremes_list.append([extr for extr in x[0]])
						elif len(x[0]) > 0:
							extremes_list.append(x[0][0])
							any_extremes = True
						else:
							extremes_list.append(float("nan"))
						
						if len(x[1]) > 1:
							outlier_list.append([outl for outl in x[1]])
						elif len(x[1]) > 0:
							outlier_list.append(x[1][0])
							any_outliers = True
						else:
							outlier_list.append(float("nan"))	
						
						temp_data = x[2]
						
						diff_list.append(abs((outliers.median_none(x[2])-mean_none(x[2]))/mean_none(x[2])))
					
					elif not (type(x) == float and numpy.isnan(x)):
						if len(x[0]) > 1:
							outlier_list.append([outl for outl in x[0]])
						elif len(x[0]) > 0:
							outlier_list.append(x[0][0])
							any_outliers = True
						else:
							outlier_list.append(float("nan"))	
						temp_data = x[1]
						
						diff_list.append(abs((outliers.median_none(x[1])-mean_none(x[1]))/mean_none(x[1])))
						
					else:
						outlier_list.append(float("nan"))	
						diff_list.append(float("nan"))
						
				print extremes_list, outlier_list, temp_data
				
				mean_list.append(mean_none(temp_data))
				median_list.append(outliers.median_none(temp_data))
					
				if use_mean:
					# compute mean and stdev
				#	mean_list.append(mean_none(temp_data))
					error_list.append(std_none(temp_data))
				elif use_median:
				#	median_list.append(outliers.median_none(temp_data))
					#error_list.append(outliers.iqr_none(temp_data))
					error_list.append((outliers.percentile_none(temp_data,25),outliers.percentile_none(temp_data,75)))
								
				raw_data_list.append(temp_data)
				
			#more CSV
			temp_output = [compound] + mean_list + error_list + median_list + diff_list
#			print ';'.join(map(str,temp_output))
			
			if use_mean:
				point_dict[compound] = mean_list
			elif use_median:
				point_dict[compound] = median_list
			
			error_dict[compound] = error_list
			
			if show_outliers:
				outlier_dict[compound] = outlier_list
				extremes_dict[compound] = extremes_list
			if show_raw_data:
				raw_data_dict[compound] = raw_data_list
	
	while len(colours) < len(sample_data):
		colours += colours
	while len(styles) < len(sample_data):
		styles += styles

	colours = colours[:len(sample_data)]
	styles = styles[:len(sample_data)]
	
	IDs = [x[1] for x in samples_to_compare]

	shift = 0.1   # Change these to increase distance between pairs of points
	curr_shift = 0.1
	
	# determine order of compounds on graph
	count_compounds = []
	
	for compound in sample_data:
		count = 0
		for sample in sample_data[compound]:
			if sample[1] != None:
				count += 1
		count_compounds.append((compound,count))
		
	count_compounds.sort(key=operator.itemgetter(1), reverse=True)
	
	# Plot points on graph
	fig,ax=plt.subplots()
	count = 0
	legend_elements = []
	
	for compound, mark, colour in zip(count_compounds,styles, colours):
		compound = compound[0]
		
		data = point_dict[compound]
		outlier_data = outlier_dict[compound]
		extremes_data = extremes_dict[compound]
		if show_raw_data:
			raw_data = raw_data_dict[compound]
		data_Err = error_dict[compound]
		
		legend_elements.append(Line2D([0], [0], marker=mark, color='w', label=compound,
						  markerfacecolor=colour, markersize=10,linestyle="None"))
						  
		for sample_index in range(len(data)):
			std = data_Err[sample_index]
			mean = data[sample_index]
					
			if use_mean:
				if std > 0 and mean > 0:
					if (std/mean)*100 < 10:
						data_Err[sample_index] = float('nan')
		
		x_vals = [curr_shift]
		for i in range(1,len(data)):
			x_vals.append(x_vals[-1]+(len(sample_data)*shift))
			
		#print x_vals			
		
		x_vals = numpy.array(x_vals)
			
		# plot outliers
		if show_outliers:
	#		print outlier_data
			for outlier_list, x_val in zip(outlier_data,x_vals):
	#			print outlier_list, x_val
				if type(outlier_list) == list:
						for outlier in outlier_list:
							plt.scatter(x_val, outlier, color = "k", marker = "x", label = '')
				else:
					plt.scatter(x_val, outlier_list, color = "k", marker = "x", label = '')
			
			#plot extremes
	#		print extremes_data
			for extremes_list, x_val in zip(extremes_data,x_vals):
	#			print extremes_list, x_val
				if type(extremes_list) == list:
						for extreme in extremes_list:
							plt.scatter(x_val, extreme, color = "b", marker = "*", label = '')
				else:
					plt.scatter(x_val, outlier_list, color = "k", marker = "*", label = '')
			
			
			
			
			#plt.scatter(x_vals, outlier_data, color = "k", marker = "x", label = '')
			
			
		# plot raw data
		if show_raw_data:
			for samplex, sampleys in zip(x_vals, raw_data):
				if len(sampleys) > 0:
					for sampley in sampleys:
						plt.scatter(samplex, sampley, color = "grey", marker = "x", label = '', alpha=0.3)
		
		if use_mean:
			# mean and stdev
			plt.scatter(x_vals, data, color = colour, marker=mark, label = compound)
			plt.errorbar(x_vals, data,yerr=data_Err, linestyle="None", color = colour, capsize=3)

		elif use_median:
			# median and IQR
			plt.scatter(x_vals, data, color = colour, marker=mark, label = compound)
			#print data_Err
			for x_val, median, percentiles in zip(x_vals, data, data_Err):
				lower_quartile = abs(median-percentiles[0])
				upper_quartile = abs(percentiles[1]-median)
				
				if median > 0.0:
				#	print ''
				#	print compound
				#	print upper_quartile
				#	print median
				#	print lower_quartile
					
					
					uplim = False
					lowlim = False
					
					#upper_lim = 0.35*(10**magnitude(median))
					#lower_lim = 0.35*(10**magnitude(median))
					upper_lim = 3500
					lower_lim = 3500
				#	print upper_lim
				#	print lower_lim
					
					if upper_quartile < upper_lim and lower_quartile > lower_lim:
							
				#			print "hiding upper marker"
							
							plotline, caplines, barline = ax.errorbar(
									x_val, median, 
									yerr=numpy.array([[lower_quartile],[upper_quartile]]),
									linestyle="None", color = colour,
									capsize=3, uplims=True
									)
							
							caplines[0].set_marker('_')
							caplines[0].set_markersize(7)
							
					elif lower_quartile < lower_lim and upper_quartile > upper_lim:
						
				#			print "hiding lower marker"
						
							plotline, caplines, barline = ax.errorbar(
									x_val, median, 
									yerr=numpy.array([[lower_quartile],[upper_quartile]]),
									linestyle="None", color = colour,
									capsize=3, lolims=True
									)
							
							caplines[0].set_marker('_')
							caplines[0].set_markersize(7)

					elif lower_quartile > lower_lim and upper_quartile > upper_lim:
						
				#		print "showing both markers"
						
						plotline, caplines, barline = ax.errorbar(
									x_val, median, 
									yerr=numpy.array([[lower_quartile],[upper_quartile]]),
									linestyle="None", color = colour,
									capsize=3,
								)
						caplines[0].set_marker('_')
						caplines[0].set_markersize(7)
						caplines[1].set_marker('_')
						caplines[1].set_markersize(7)
					
				#	else:
				#		print "hiding both markers"
					#raw_input(">")
			
		count += 1
		if count == len(sample_data):
		#if count == 5: # custom override
			break
		curr_shift += shift
	
	#gridlines and labels
	plt.grid(b=True, axis = "y", which='major', color='b', linestyle='-')
	plt.grid(b=True, axis = "y", which='minor', color='r', alpha=0.3, linestyle='--')
	ax.grid(axis = "x", which='major', linestyle='-')
		
	x_vals = [shift/2]
	min_x_vals = [0.1 + ((len(sample_data)*shift)/2)]

	for i in range(len(data)):
		x_vals.append(x_vals[-1]+(len(sample_data)*shift))
		min_x_vals.append(x_vals[-1]+((len(sample_data)*shift)/2))
		
	x_vals = numpy.array(x_vals)
	min_x_vals = numpy.array(min_x_vals)
	
	# Major tick labels
	plt.xticks(x_vals, IDs, size='small')#,ha='left')
	ax.xaxis.set_major_formatter(ticker.NullFormatter())

	# Minor tick labels
	ax.xaxis.set_minor_locator(ticker.FixedLocator(min_x_vals))
	ax.xaxis.set_minor_formatter(ticker.FixedFormatter(IDs))
	
	# x-axis limits	
	ax.set_xlim(left=x_vals[0]-(shift/2), right=(x_vals[-1]+(shift/2)))
	ax.set_axisbelow(True)
		
	plt.yscale('log') # logarithmic scale on y-axis

	plt.xlabel('Ammunition',fontsize=12,labelpad=10)
	plt.ylabel('Peak Area',fontsize=12,labelpad=10)
	
	#fig.set_size_inches(11.69,8.27) # A4 Landscape
	fig.set_size_inches(4,8.27)
	fig.tight_layout()    
	
#	plt.show()
#	sys.exit()
	
	#save chart
	for file_format in ["pdf","png"]:
#	for file_format in ["pdf"]:
		plt.savefig("./charts/{}_{}_chart.{}".format("_".join(IDs),metric,file_format), papertype = 'a4', orientation = 'landscape', format = file_format, dpi = 1200,bbox_inches="tight")
	
	
	#legend
	#print outlier_dict
	if show_outliers and any_outliers:
		legend_elements.append(Line2D([0], [0], marker="x", color="k", label="Outlier",
						  markerfacecolor="k", markersize=10,linestyle="None"))
	if show_raw_data:
		legend_elements.append(Line2D([0], [0], marker="x", color="k", label="Raw Data",
						  markerfacecolor="grey", alpha=0.3, markersize=10,linestyle="None"))
		
	plt.clf()
	plt.axis('off')	
	fig.set_size_inches(4,4)	
	plt.legend(handles=legend_elements, loc="center", scatterpoints=1,ncol=leg_cols,title="Legend")	
	fig.tight_layout()    
	
	#save legend
	for file_format in ["pdf","png"]:
#	for file_format in ["pdf"]:
		plt.savefig("./charts/{}_{}_legend.{}".format("_".join(IDs),metric,file_format), papertype = 'a4', orientation = 'portrait', format = file_format, dpi = 1200,bbox_inches="tight")


def magnitude(x):
	return int(math.log10(x))
			
if __name__ == '__main__':
	"""from https://stackoverflow.com/questions/14463277/how-to-disable-python-warnings"""
	with warnings.catch_warnings():
		warnings.simplefilter("ignore")

		#Command line switches
		parser = argparse.ArgumentParser()
		
		parser.add_argument("--info",help="Show program info.", action='store_true')
		parser.add_argument("-v","--verbose",help="Turns on verbosity for diagnostics", action='store_true')
		
		args = parser.parse_args()
		
		if args.info:
			parser.print_help()	#show help and info
			info()
			sys.exit(0) 
			
		time.sleep(0.1); bar.update(24)
		
		print("\nReady")
		comparison()	# Actual Comparison
		print("\nComplete.")
